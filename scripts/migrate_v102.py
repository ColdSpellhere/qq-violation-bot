#!/usr/bin/env python3
from __future__ import annotations

import argparse
import hashlib
import importlib.util
import json
import os
import re
import sqlite3
import sys
import tempfile
from contextlib import contextmanager
from dataclasses import dataclass
from datetime import datetime
from pathlib import Path
from typing import Iterator
from urllib.parse import quote
from zipfile import ZIP_DEFLATED, ZipFile

from openpyxl import load_workbook


PROJECT_DIR = Path(__file__).resolve().parents[1]
if str(PROJECT_DIR) not in sys.path:
    sys.path.insert(0, str(PROJECT_DIR))

def _load_pure_module(name: str, path: Path):
    spec = importlib.util.spec_from_file_location(name, path)
    if spec is None or spec.loader is None:
        raise RuntimeError(f"cannot load module: {path}")
    module = importlib.util.module_from_spec(spec)
    sys.modules[name] = module
    spec.loader.exec_module(module)
    return module


_POLICY_DIR = PROJECT_DIR / "plugins" / "violation_record"
_deduction_policy = _load_pure_module(
    "_v102_migration_deduction_policy", _POLICY_DIR / "deduction_policy.py"
)
_policy_schema = _load_pure_module(
    "_v102_migration_policy_schema", _POLICY_DIR / "policy_schema.py"
)

RULE_VERSION = _deduction_policy.RULE_VERSION
TERMINAL_STATUSES = _deduction_policy.TERMINAL_STATUSES
_insert_event = _deduction_policy._insert_event
_create_pending_action = _deduction_policy._create_pending_action
_start_cycle = _deduction_policy._start_cycle
raw_effective_record_summary = _deduction_policy.raw_effective_record_summary
REQUIRED_V102_INDEXES = _policy_schema.REQUIRED_V102_INDEXES
REQUIRED_V102_TABLES = _policy_schema.REQUIRED_V102_TABLES
V102_SCHEMA_VERSION = _policy_schema.V102_SCHEMA_VERSION
REPAIRABLE_SCHEMA_VERSIONS = frozenset({"v1.0.2beta-1", V102_SCHEMA_VERSION})
ensure_v102_schema = _policy_schema.ensure_v102_schema


MAIN_SHEETS = ("蜂巢", "蜂窝", "蜂箱")
EXCLUDED_SHEETS = ("低频小于三", "封存记录", "手动黑名单", "OOPZ")
REQUIRED_SHEETS = (*MAIN_SHEETS, *EXCLUDED_SHEETS)
LOW_FREQUENCY_SHEET = "低频小于三"
QQ_CELL_RE = re.compile(r"[（(](\d{5,12})[）)]")
QQ_CANDIDATE_RE = re.compile(r"\d{5,12}")
SHA256_RE = re.compile(r"[0-9a-fA-F]{64}")
BATCH_ID_RE = re.compile(r"[A-Za-z0-9][A-Za-z0-9._:-]{0,127}")
REQUIRED_UNIQUE_CONSTRAINTS = {
    "v102_policy_events": ("idempotency_key",),
    "v102_pending_actions": ("idempotency_key",),
    "v102_notification_outbox": (
        "event_id",
        "message_type",
        "reminder_slot",
    ),
}
REQUIRED_SNAPSHOT_COLUMNS = {
    "v102_policy_state": {
        "baseline_total_count",
        "baseline_current_count",
        "baseline_raw_total",
        "baseline_record_watermark",
        "baseline_locked",
        "baseline_last_effective_violation_time",
        "baseline_last_deduct_time",
        "baseline_last_final_warning_time",
    },
    "v102_baseline_audit": {
        "old_locked",
        "old_last_effective_violation_time",
        "old_last_deduct_time",
        "old_last_final_warning_time",
    },
    "v102_status_bridge_jobs": {"caused_by_record_id"},
}


class MigrationError(RuntimeError):
    pass


@dataclass(frozen=True)
class BaselineRow:
    group_area: str
    qq_number: str
    approved_current_count: int
    source_sheet: str
    source_row: int


@dataclass(frozen=True)
class BaselineData:
    rows: tuple[BaselineRow, ...]
    source_sha256: str


def file_sha256(path: Path) -> str:
    digest = hashlib.sha256()
    with path.open("rb") as stream:
        for chunk in iter(lambda: stream.read(1024 * 1024), b""):
            digest.update(chunk)
    return digest.hexdigest()


def _require_file(path: Path, label: str) -> Path:
    resolved = path.expanduser().resolve()
    if not resolved.is_file():
        raise MigrationError(f"{label}不存在或不是普通文件：{resolved}")
    return resolved


@contextmanager
def _baseline_workbook(path: Path) -> Iterator[object]:
    workbook = None
    temporary: tempfile.TemporaryDirectory[str] | None = None
    try:
        try:
            workbook = load_workbook(
                path, read_only=True, data_only=True, keep_links=False
            )
        except Exception as first_error:
            temporary = tempfile.TemporaryDirectory(prefix="v102-baseline-")
            safe_path = Path(temporary.name) / "baseline-no-styles.xlsx"
            try:
                with ZipFile(path, "r") as source, ZipFile(
                    safe_path, "w", ZIP_DEFLATED
                ) as target:
                    for item in source.infolist():
                        if item.filename in {
                            "xl/styles.xml",
                            "xl/theme/theme1.xml",
                        }:
                            continue
                        target.writestr(item, source.read(item.filename))
                workbook = load_workbook(
                    safe_path, read_only=True, data_only=True, keep_links=False
                )
            except Exception as second_error:
                raise MigrationError(
                    "基线工作簿无法只读解析："
                    f"{type(first_error).__name__}/{type(second_error).__name__}"
                ) from second_error
        yield workbook
    finally:
        if workbook is not None:
            workbook.close()
        if temporary is not None:
            temporary.cleanup()


def _parse_current_count(value: object, *, sheet: str, row: int) -> int:
    if isinstance(value, bool):
        raise MigrationError(f"{sheet}!C{row} 当前次数不是非负整数")
    if isinstance(value, int):
        result = value
    elif isinstance(value, float) and value.is_integer():
        result = int(value)
    elif isinstance(value, str) and value.strip().isdigit():
        result = int(value.strip())
    else:
        raise MigrationError(f"{sheet}!C{row} 当前次数不是非负整数")
    if result < 0:
        raise MigrationError(f"{sheet}!C{row} 当前次数不是非负整数")
    return result


def _extract_sheet_rows(workbook, sheet_name: str) -> list[BaselineRow]:
    sheet = workbook[sheet_name]
    rows: list[BaselineRow] = []
    for row_number, values in enumerate(sheet.iter_rows(values_only=True), 1):
        member_text = str(values[1] or "").strip() if len(values) > 1 else ""
        match = QQ_CELL_RE.search(member_text)
        if match is None:
            if (
                sheet_name in MAIN_SHEETS
                and len(values) > 2
                and values[2] not in (None, "")
                and QQ_CANDIDATE_RE.search(member_text)
            ):
                raise MigrationError(
                    f"{sheet_name}!B{row_number} QQ号未使用标准括号格式"
                )
            continue
        count_value = values[2] if len(values) > 2 else None
        rows.append(
            BaselineRow(
                group_area=sheet_name,
                qq_number=match.group(1),
                approved_current_count=_parse_current_count(
                    count_value, sheet=sheet_name, row=row_number
                ),
                source_sheet=sheet_name,
                source_row=row_number,
            )
        )
    return rows


def _extract_excluded_qqs(workbook) -> dict[str, set[str]]:
    excluded: dict[str, set[str]] = {}
    # Only standardized low-frequency rows conflict with the three main areas.
    # The other excluded sheets are never imported, and OOPZ can share QQs.
    for sheet_name in (LOW_FREQUENCY_SHEET,):
        sheet = workbook[sheet_name]
        for values in sheet.iter_rows(values_only=True):
            member_text = str(values[1] or "").strip() if len(values) > 1 else ""
            match = QQ_CELL_RE.search(member_text)
            if match:
                excluded.setdefault(match.group(1), set()).add(sheet_name)
    return excluded


def read_baseline(path: str | Path) -> BaselineData:
    source = _require_file(Path(path), "基线文件")
    source_sha256 = file_sha256(source)
    with _baseline_workbook(source) as workbook:
        missing = [name for name in REQUIRED_SHEETS if name not in workbook.sheetnames]
        if missing:
            raise MigrationError(f"基线工作簿缺少工作表：{', '.join(missing)}")
        rows: list[BaselineRow] = []
        for sheet_name in MAIN_SHEETS:
            rows.extend(_extract_sheet_rows(workbook, sheet_name))
        if not rows:
            raise MigrationError("三个主群域没有可迁移成员")

        seen: dict[tuple[str, str], BaselineRow] = {}
        duplicates: list[str] = []
        for row in rows:
            key = (row.group_area, row.qq_number)
            if key in seen:
                duplicates.append(
                    f"{row.group_area}/{row.qq_number}"
                    f"({seen[key].source_row},{row.source_row})"
                )
            else:
                seen[key] = row
        if duplicates:
            raise MigrationError(f"主群域存在重复成员：{', '.join(duplicates)}")

        excluded = _extract_excluded_qqs(workbook)
        conflicts = [
            f"{row.group_area}/{row.qq_number} -> {'/'.join(sorted(excluded[row.qq_number]))}"
            for row in rows
            if row.qq_number in excluded
        ]
        if conflicts:
            raise MigrationError(f"主群域成员同时出现在排除表：{', '.join(conflicts)}")

    return BaselineData(tuple(rows), source_sha256)


def _read_only_connection(path: Path) -> sqlite3.Connection:
    uri = f"file:{quote(str(path), safe='/')}?mode=ro"
    conn = sqlite3.connect(uri, uri=True)
    conn.row_factory = sqlite3.Row
    conn.execute("PRAGMA foreign_keys = ON")
    conn.execute("PRAGMA busy_timeout = 5000")
    return conn


def _write_connection(path: Path) -> sqlite3.Connection:
    conn = sqlite3.connect(path, isolation_level=None)
    conn.row_factory = sqlite3.Row
    conn.execute("PRAGMA foreign_keys = ON")
    conn.execute("PRAGMA busy_timeout = 5000")
    return conn


def _table_names(conn: sqlite3.Connection) -> set[str]:
    return {
        row["name"]
        for row in conn.execute("SELECT name FROM sqlite_master WHERE type='table'")
    }


def _index_names(conn: sqlite3.Connection) -> set[str]:
    return {
        row["name"]
        for row in conn.execute("SELECT name FROM sqlite_master WHERE type='index'")
    }


def _has_unique_constraint(
    conn: sqlite3.Connection, table: str, columns: tuple[str, ...]
) -> bool:
    table_identifier = table.replace('"', '""')
    for index in conn.execute(f'PRAGMA index_list("{table_identifier}")'):
        if int(index["unique"] or 0) != 1:
            continue
        index_identifier = str(index["name"]).replace('"', '""')
        indexed_columns = tuple(
            str(row["name"])
            for row in conn.execute(f'PRAGMA index_info("{index_identifier}")')
        )
        if indexed_columns == columns:
            return True
    return False


def _validate_legacy_database(conn: sqlite3.Connection) -> None:
    required = {
        "members",
        "member_group_states",
        "violation_records",
        "consultation_records",
        "operation_logs",
    }
    missing = required - _table_names(conn)
    if missing:
        raise MigrationError(f"数据库缺少业务表：{', '.join(sorted(missing))}")
    integrity = conn.execute("PRAGMA integrity_check").fetchone()[0]
    if integrity != "ok":
        raise MigrationError(f"数据库完整性检查失败：{integrity}")
    foreign_errors = conn.execute("PRAGMA foreign_key_check").fetchall()
    if foreign_errors:
        raise MigrationError(f"数据库存在外键错误：{len(foreign_errors)}")


def _member_and_state(
    conn: sqlite3.Connection, qq_number: str, group_area: str
) -> tuple[sqlite3.Row | None, sqlite3.Row | None]:
    member = conn.execute(
        "SELECT * FROM members WHERE qq_number=?", (qq_number,)
    ).fetchone()
    if member is None:
        return None, None
    state = conn.execute(
        """
        SELECT * FROM member_group_states
        WHERE member_id=? AND group_area=?
        """,
        (member["id"], group_area),
    ).fetchone()
    return member, state


def _old_adjustment(
    conn: sqlite3.Connection, member_id: int | None, group_area: str
) -> int:
    if member_id is None or "v102_policy_state" not in _table_names(conn):
        return 0
    row = conn.execute(
        """
        SELECT baseline_adjustment FROM v102_policy_state
        WHERE member_id=? AND group_area=?
        """,
        (member_id, group_area),
    ).fetchone()
    return int(row["baseline_adjustment"] or 0) if row else 0


def _raw_total(
    conn: sqlite3.Connection, member_id: int | None, group_area: str
) -> int:
    if member_id is None:
        return 0
    total, _ = raw_effective_record_summary(conn, member_id, group_area)
    return total


def _initial_cycle(
    *, status: str, current: int, last_final_warning_time: str | None
) -> tuple[str | None, str | None]:
    if status in TERMINAL_STATUSES:
        return None, "terminal_status"
    if status == "最后警告":
        if not last_final_warning_time:
            raise MigrationError("最后警告成员缺少 last_final_warning_time")
        try:
            datetime.fromisoformat(last_final_warning_time)
        except ValueError as exc:
            raise MigrationError("最后警告时间格式无效") from exc
        return "final_warning", None
    if status == "已质询" or current >= 3:
        return "slow", None
    if current > 0:
        return "normal", None
    return None, "zero_count"


def _change_for_row(
    conn: sqlite3.Connection, row: BaselineRow
) -> dict[str, object]:
    member, state = _member_and_state(conn, row.qq_number, row.group_area)
    member_id = int(member["id"]) if member else None
    old_total = int(state["total_count"] or 0) if state else 0
    old_deduct = int(state["deduct_count"] or 0) if state else 0
    old_current = int(state["current_count_cache"] or 0) if state else 0
    status = str(state["status"] or "正常") if state else "正常"
    last_final_warning = state["last_final_warning_time"] if state else None
    raw_total = _raw_total(conn, member_id, row.group_area)
    new_total = old_deduct + row.approved_current_count
    new_adjustment = new_total - raw_total
    cycle_type, no_cycle_reason = _initial_cycle(
        status=status,
        current=row.approved_current_count,
        last_final_warning_time=last_final_warning,
    )
    return {
        "group_area": row.group_area,
        "qq_number": row.qq_number,
        "source_sheet": row.source_sheet,
        "source_row": row.source_row,
        "member_id": member_id,
        "was_created": member is None or state is None,
        "old_total_count": old_total,
        "old_deduct_count": old_deduct,
        "old_current_count": old_current,
        "old_baseline_adjustment": _old_adjustment(
            conn, member_id, row.group_area
        ),
        "approved_current_count": row.approved_current_count,
        "raw_total_at_cutover": raw_total,
        "new_total_count": new_total,
        "new_baseline_adjustment": new_adjustment,
        "status": status,
        "last_final_warning_time": last_final_warning,
        "initial_cycle_type": cycle_type,
        "no_cycle_reason": no_cycle_reason,
    }


def _build_plan(
    conn: sqlite3.Connection, baseline: BaselineData
) -> dict[str, object]:
    _validate_legacy_database(conn)
    watermark = int(
        conn.execute("SELECT COALESCE(MAX(id), 0) FROM violation_records").fetchone()[0]
    )
    changes = [_change_for_row(conn, row) for row in baseline.rows]
    cycle_counts: dict[str, int] = {}
    for change in changes:
        key = str(change["initial_cycle_type"] or change["no_cycle_reason"])
        cycle_counts[key] = cycle_counts.get(key, 0) + 1
    return {
        "source_sha256": baseline.source_sha256,
        "cutover_record_watermark": watermark,
        "summary": {
            "baseline_rows": len(changes),
            "by_area": {
                area: sum(1 for item in changes if item["group_area"] == area)
                for area in MAIN_SHEETS
            },
            "initialization": cycle_counts,
        },
        "changes": changes,
    }


def dry_run(database: str | Path, baseline: str | Path) -> dict[str, object]:
    database_path = _require_file(Path(database), "数据库")
    baseline_data = read_baseline(baseline)
    conn = _read_only_connection(database_path)
    try:
        return _build_plan(conn, baseline_data)
    finally:
        conn.close()


def _validate_backup_sha256(value: str) -> str:
    normalized = str(value or "").strip().lower()
    if not SHA256_RE.fullmatch(normalized):
        raise MigrationError("backup SHA-256 必须是 64 位十六进制值")
    return normalized


def _business_data_fingerprint(conn: sqlite3.Connection) -> str:
    digest = hashlib.sha256()
    tables = sorted(
        str(row["name"])
        for row in conn.execute(
            """
            SELECT name FROM sqlite_master
            WHERE type='table' AND name NOT LIKE 'v102_%'
              AND (name NOT LIKE 'sqlite_%' OR name='sqlite_sequence')
            """
        )
    )
    digest.update(
        json.dumps(tables, ensure_ascii=False, separators=(",", ":")).encode()
    )
    digest.update(b"\n")
    for table in tables:
        identifier = table.replace('"', '""')
        table_sql = conn.execute(
            "SELECT sql FROM sqlite_master WHERE type='table' AND name=?", (table,)
        ).fetchone()[0]
        digest.update(str(table_sql or "").encode())
        digest.update(b"\n")
        columns = [
            str(row["name"])
            for row in conn.execute(f'PRAGMA table_info("{identifier}")')
        ]
        digest.update(
            json.dumps(columns, ensure_ascii=False, separators=(",", ":")).encode()
        )
        digest.update(b"\n")
        order_by = ", ".join(
            '"' + column.replace('"', '""') + '"' for column in columns
        )
        query = f'SELECT * FROM "{identifier}"'
        if order_by:
            query += f" ORDER BY {order_by}"
        for row in conn.execute(query):
            values = [
                {"bytes": value.hex()} if isinstance(value, bytes) else value
                for value in row
            ]
            digest.update(
                json.dumps(
                    values,
                    ensure_ascii=False,
                    separators=(",", ":"),
                    default=str,
                ).encode()
            )
            digest.update(b"\n")
    schema_objects = conn.execute(
        """
        SELECT type, name, tbl_name, sql FROM sqlite_master
        WHERE type IN ('index', 'trigger', 'view')
          AND tbl_name NOT LIKE 'v102_%'
          AND name NOT LIKE 'sqlite_autoindex_%'
        ORDER BY type, name
        """
    ).fetchall()
    for row in schema_objects:
        digest.update(
            json.dumps(
                list(row), ensure_ascii=False, separators=(",", ":"), default=str
            ).encode()
        )
        digest.update(b"\n")
    return digest.hexdigest()


def _validate_pre_cutover_backup(
    database_path: Path,
    snapshot_database: str | Path,
    expected_sha256: str,
) -> tuple[Path, str]:
    snapshot_path = _require_file(Path(snapshot_database), "切换前数据库快照")
    try:
        same_file = os.path.samefile(snapshot_path, database_path)
    except OSError as exc:
        raise MigrationError("无法确认切换前数据库快照是否为独立文件") from exc
    if same_file:
        raise MigrationError("切换前数据库快照不能与目标数据库是同一文件")
    actual_digest = file_sha256(snapshot_path)
    if actual_digest != expected_sha256:
        raise MigrationError(
            "切换前数据库快照摘要不匹配："
            f"expected={expected_sha256} actual={actual_digest}"
        )
    return snapshot_path, actual_digest


def _time_text(value: str | None = None) -> str:
    if value is None:
        return datetime.now().replace(microsecond=0).isoformat(sep=" ")
    try:
        return datetime.fromisoformat(value).replace(microsecond=0).isoformat(sep=" ")
    except ValueError as exc:
        raise MigrationError(f"时间格式无效：{value}") from exc


def _default_batch_id(cutover_at: str, source_sha256: str) -> str:
    stamp = datetime.fromisoformat(cutover_at).strftime("%Y%m%d_%H%M%S")
    return f"v102-{stamp}-{source_sha256[:12]}"


def _validate_batch_id(value: str) -> str:
    normalized = str(value or "").strip()
    if not BATCH_ID_RE.fullmatch(normalized):
        raise MigrationError(
            "batch_id 必须为 1-128 位字母、数字、点、下划线、冒号或短横线"
        )
    return normalized


def _ensure_member_and_state(
    conn: sqlite3.Connection,
    *,
    qq_number: str,
    group_area: str,
    at: str,
) -> tuple[int, sqlite3.Row, bool]:
    member = conn.execute(
        "SELECT * FROM members WHERE qq_number=?", (qq_number,)
    ).fetchone()
    member_created = member is None
    if member is None:
        cursor = conn.execute(
            """
            INSERT INTO members(
                qq_number, qq_nickname, aliases, created_at, updated_at
            ) VALUES(?, NULL, '[]', ?, ?)
            """,
            (qq_number, at, at),
        )
        member_id = int(cursor.lastrowid)
    else:
        member_id = int(member["id"])

    state = conn.execute(
        """
        SELECT * FROM member_group_states
        WHERE member_id=? AND group_area=?
        """,
        (member_id, group_area),
    ).fetchone()
    state_created = state is None
    if state is None:
        conn.execute(
            """
            INSERT INTO member_group_states(
                member_id, group_area, created_at, updated_at
            ) VALUES(?, ?, ?, ?)
            """,
            (member_id, group_area, at, at),
        )
        state = conn.execute(
            """
            SELECT * FROM member_group_states
            WHERE member_id=? AND group_area=?
            """,
            (member_id, group_area),
        ).fetchone()
    return member_id, state, member_created or state_created


def _assert_empty_policy_data(conn: sqlite3.Connection) -> None:
    checks = (
        "v102_policy_events",
        "v102_policy_cycles",
        "v102_policy_state",
        "v102_pending_actions",
        "v102_notification_outbox",
        "v102_migration_checkpoints",
        "v102_baseline_audit",
    )
    populated = [
        table
        for table in checks
        if int(conn.execute(f"SELECT COUNT(*) FROM {table}").fetchone()[0]) > 0
    ]
    if populated:
        raise MigrationError(f"v102 已应用或存在策略数据：{', '.join(populated)}")


def _insert_operation_log(
    conn: sqlite3.Connection,
    *,
    member_id: int,
    group_area: str,
    operation_type: str,
    before: dict[str, object],
    after: dict[str, object],
    batch_id: str,
    at: str,
) -> None:
    conn.execute(
        """
        INSERT INTO operation_logs(
            group_area, operation_type, source, target_member_id,
            before_json, after_json, created_at, remark
        ) VALUES(?, ?, '自动', ?, ?, ?, ?, ?)
        """,
        (
            group_area,
            operation_type,
            member_id,
            json.dumps(before, ensure_ascii=False, sort_keys=True),
            json.dumps(after, ensure_ascii=False, sort_keys=True),
            at,
            f"batch_id={batch_id}",
        ),
    )


def _apply_row(
    conn: sqlite3.Connection,
    *,
    row: BaselineRow,
    batch_id: str,
    cutover_at: str,
    cutover_record_watermark: int,
) -> None:
    member_id, state, was_created = _ensure_member_and_state(
        conn,
        qq_number=row.qq_number,
        group_area=row.group_area,
        at=cutover_at,
    )
    old_total = int(state["total_count"] or 0)
    old_deduct = int(state["deduct_count"] or 0)
    old_current = int(state["current_count_cache"] or 0)
    raw_total, _ = raw_effective_record_summary(conn, member_id, row.group_area)
    new_total = old_deduct + row.approved_current_count
    adjustment = new_total - raw_total
    status = str(state["status"] or "正常")
    cycle_type, no_cycle_reason = _initial_cycle(
        status=status,
        current=row.approved_current_count,
        last_final_warning_time=state["last_final_warning_time"],
    )

    conn.execute(
        """
        INSERT INTO v102_policy_state(
            member_id, group_area, policy_tag, slow_level,
            v102_operation_count, baseline_adjustment,
            baseline_total_count, baseline_deduct_count,
            baseline_current_count, baseline_raw_total,
            baseline_record_watermark, baseline_locked, baseline_status,
            baseline_last_effective_violation_time,
            baseline_last_deduct_time, baseline_last_final_warning_time,
            baseline_initialized_at, no_cycle_reason,
            created_at, updated_at
        ) VALUES(?, ?, 'none', 0, 0, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
        """,
        (
            member_id,
            row.group_area,
            adjustment,
            old_total,
            old_deduct,
            old_current,
            raw_total,
            cutover_record_watermark,
            1 if int(state["locked"] or 0) else 0,
            status,
            state["last_effective_violation_time"],
            state["last_deduct_time"],
            state["last_final_warning_time"],
            cutover_at,
            no_cycle_reason,
            cutover_at,
            cutover_at,
        ),
    )
    conn.execute(
        """
        UPDATE member_group_states
        SET total_count=?, current_count_cache=?, updated_at=?
        WHERE member_id=? AND group_area=?
        """,
        (
            new_total,
            row.approved_current_count,
            cutover_at,
            member_id,
            row.group_area,
        ),
    )
    conn.execute(
        """
        INSERT INTO v102_baseline_audit(
            batch_id, member_id, group_area, old_total_count,
            old_deduct_count, old_current_count, old_baseline_adjustment,
            old_locked, old_last_effective_violation_time,
            old_last_deduct_time, old_last_final_warning_time,
            approved_current_count, new_total_count,
            new_baseline_adjustment, was_created, source_sheet,
            source_row, created_at
        ) VALUES(?, ?, ?, ?, ?, ?, 0, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
        """,
        (
            batch_id,
            member_id,
            row.group_area,
            old_total,
            old_deduct,
            old_current,
            1 if int(state["locked"] or 0) else 0,
            state["last_effective_violation_time"],
            state["last_deduct_time"],
            state["last_final_warning_time"],
            row.approved_current_count,
            new_total,
            adjustment,
            1 if was_created else 0,
            row.source_sheet,
            row.source_row,
            cutover_at,
        ),
    )
    baseline_event_id, created = _insert_event(
        conn,
        member_id=member_id,
        group_area=row.group_area,
        event_type="baseline_migrated",
        effective_time=cutover_at,
        event_priority=0,
        source_sequence=row.source_row,
        ingest_time=cutover_at,
        idempotency_key=(
            f"migration:{batch_id}:baseline:{row.group_area}:{row.qq_number}"
        ),
        payload={
            "batch_id": batch_id,
            "approved_current_count": row.approved_current_count,
            "baseline_deduct_count": old_deduct,
            "baseline_adjustment": adjustment,
            "baseline_status": status,
        },
    )
    if not created:
        raise MigrationError(
            f"迁移事件幂等冲突：{row.group_area}/{row.qq_number}"
        )

    if cycle_type:
        start_at = (
            str(state["last_final_warning_time"])
            if cycle_type == "final_warning"
            else cutover_at
        )
        cycle_id = _start_cycle(
            conn,
            member_id=member_id,
            group_area=row.group_area,
            cycle_type=cycle_type,
            start_at=start_at,
            caused_by_event_id=baseline_event_id,
            fixed_sequence=1 if cycle_type in {"stop", "final_warning"} else 0,
        )
        if cycle_type == "final_warning" and cycle_id is not None:
            historical = conn.execute(
                """
                SELECT COUNT(*) AS count, MAX(id) AS last_id,
                       MAX(violation_time) AS last_time
                FROM violation_records
                WHERE member_id=? AND group_area=? AND id<=?
                  AND is_withdrawn=0 AND is_test=0 AND is_countable=1
                  AND action LIKE '%禁言%'
                  AND violation_time>?
                """,
                (
                    member_id,
                    row.group_area,
                    cutover_record_watermark,
                    start_at,
                ),
            ).fetchone()
            if int(historical["count"] or 0) > 0:
                event_id, _ = _insert_event(
                    conn,
                    member_id=member_id,
                    group_area=row.group_area,
                    event_type="historical_final_warning_violation_detected",
                    effective_time=str(historical["last_time"]),
                    event_priority=10,
                    source_sequence=int(historical["last_id"] or 0),
                    ingest_time=cutover_at,
                    idempotency_key=(
                        f"migration:{batch_id}:final-warning-history:"
                        f"{row.group_area}:{row.qq_number}"
                    ),
                    caused_by_event_id=baseline_event_id,
                    payload={
                        "batch_id": batch_id,
                        "violation_count": int(historical["count"]),
                        "last_violation_time": str(historical["last_time"]),
                    },
                )
                _create_pending_action(
                    conn,
                    member_id=member_id,
                    group_area=row.group_area,
                    action_type="remove_member",
                    reason="最后警告后已有历史禁言，请管理判断是否移出",
                    caused_by_event_id=event_id,
                    at=cutover_at,
                )
                conn.execute(
                    """
                    UPDATE v102_policy_cycles
                    SET status='pending_decision', updated_at=? WHERE id=?
                    """,
                    (cutover_at, cycle_id),
                )

    _insert_operation_log(
        conn,
        member_id=member_id,
        group_area=row.group_area,
        operation_type="v1.0.2beta基线迁移",
        before={
            "total_count": old_total,
            "deduct_count": old_deduct,
            "current_count": old_current,
        },
        after={
            "total_count": new_total,
            "deduct_count": old_deduct,
            "current_count": row.approved_current_count,
            "baseline_adjustment": adjustment,
            "initial_cycle_type": cycle_type,
            "no_cycle_reason": no_cycle_reason,
        },
        batch_id=batch_id,
        at=cutover_at,
    )


def _validate_applied_transaction(
    conn: sqlite3.Connection, *, batch_id: str, expected_rows: int
) -> None:
    audit_count = int(
        conn.execute(
            "SELECT COUNT(*) FROM v102_baseline_audit WHERE batch_id=?",
            (batch_id,),
        ).fetchone()[0]
    )
    if audit_count != expected_rows:
        raise MigrationError(
            f"迁移审计数量不一致：expected={expected_rows} actual={audit_count}"
        )
    bad = conn.execute(
        """
        SELECT COUNT(*)
        FROM v102_baseline_audit a
        JOIN member_group_states s
          ON s.member_id=a.member_id AND s.group_area=a.group_area
        JOIN v102_policy_state p
          ON p.member_id=a.member_id AND p.group_area=a.group_area
        WHERE a.batch_id=?
          AND (
              s.total_count != a.new_total_count
              OR s.deduct_count != a.old_deduct_count
              OR s.current_count_cache != a.approved_current_count
              OR p.baseline_adjustment != a.new_baseline_adjustment
              OR p.v102_operation_count != 0
          )
        """,
        (batch_id,),
    ).fetchone()[0]
    if bad:
        raise MigrationError(f"迁移后投影不一致：{bad}")
    foreign_errors = conn.execute("PRAGMA foreign_key_check").fetchall()
    if foreign_errors:
        raise MigrationError(f"迁移产生外键错误：{len(foreign_errors)}")


def apply_migration(
    database: str | Path,
    baseline: str | Path,
    *,
    snapshot_database: str | Path,
    backup_sha256: str,
    cutover_at: str | None = None,
    batch_id: str | None = None,
) -> dict[str, object]:
    database_path = _require_file(Path(database), "数据库")
    backup_digest = _validate_backup_sha256(backup_sha256)
    snapshot_path, backup_digest = _validate_pre_cutover_backup(
        database_path, snapshot_database, backup_digest
    )
    baseline_data = read_baseline(baseline)
    at = _time_text(cutover_at)
    migration_id = _validate_batch_id(
        batch_id or _default_batch_id(at, baseline_data.source_sha256)
    )

    conn = _write_connection(database_path)
    try:
        conn.execute("BEGIN IMMEDIATE")
        _validate_legacy_database(conn)
        snapshot = _read_only_connection(snapshot_path)
        try:
            _validate_legacy_database(snapshot)
            if _business_data_fingerprint(snapshot) != _business_data_fingerprint(conn):
                raise MigrationError("切换前数据库快照与目标数据库业务数据不一致")
        finally:
            snapshot.close()
        if baseline_data.source_sha256 != file_sha256(
            _require_file(Path(baseline), "基线文件")
        ):
            raise MigrationError("基线文件在预检后发生变化")
        ensure_v102_schema(conn)
        _assert_empty_policy_data(conn)
        watermark = int(
            conn.execute(
                "SELECT COALESCE(MAX(id), 0) FROM violation_records"
            ).fetchone()[0]
        )
        conn.execute(
            """
            INSERT INTO v102_migration_checkpoints(
                batch_id, schema_version, cutover_at,
                cutover_record_watermark, source_sha256,
                backup_sha256, status, created_at, updated_at
            ) VALUES(?, ?, ?, ?, ?, ?, 'applied', ?, ?)
            """,
            (
                migration_id,
                V102_SCHEMA_VERSION,
                at,
                watermark,
                baseline_data.source_sha256,
                backup_digest,
                at,
                at,
            ),
        )
        for row in baseline_data.rows:
            _apply_row(
                conn,
                row=row,
                batch_id=migration_id,
                cutover_at=at,
                cutover_record_watermark=watermark,
            )
        _validate_applied_transaction(
            conn, batch_id=migration_id, expected_rows=len(baseline_data.rows)
        )
        conn.commit()
    except Exception:
        if conn.in_transaction:
            conn.rollback()
        raise
    finally:
        conn.close()

    return {
        "status": "applied",
        "batch_id": migration_id,
        "cutover_at": at,
        "cutover_record_watermark": watermark,
        "source_sha256": baseline_data.source_sha256,
        "backup_sha256": backup_digest,
        "baseline_rows": len(baseline_data.rows),
    }


def _rollback_reason(status: str, current: int) -> str | None:
    if status in TERMINAL_STATUSES:
        return "terminal_status"
    if current <= 0:
        return "zero_count"
    return None


def _status_after_manual_inputs(
    conn: sqlite3.Connection,
    *,
    member_id: int,
    group_area: str,
    baseline_status: str,
    baseline_last_final_warning_time: str | None,
) -> tuple[str, str | None]:
    status = baseline_status
    last_final_warning_time = baseline_last_final_warning_time
    timeline: list[tuple[str, int, int, int, str]] = []
    rows = conn.execute(
        """
        SELECT e.payload_json, e.effective_time, e.event_priority,
               e.source_sequence, e.id,
               j.operation_log_id AS bridge_operation_log_id
        FROM v102_policy_events e
        LEFT JOIN (
            SELECT applied_event_id, MIN(operation_log_id) AS operation_log_id
            FROM v102_status_bridge_jobs
            WHERE applied_event_id IS NOT NULL
            GROUP BY applied_event_id
        ) j ON j.applied_event_id=e.id
        WHERE e.member_id=? AND e.group_area=? AND e.event_type='status_changed'
          AND e.replay_generation=0 AND e.is_effective=1
        """,
        (member_id, group_area),
    ).fetchall()
    for row in rows:
        payload = json.loads(row["payload_json"] or "{}")
        if payload.get("status"):
            sequence = (
                int(row["bridge_operation_log_id"])
                if row["bridge_operation_log_id"] is not None
                else int(row["source_sequence"] or row["id"])
            )
            timeline.append(
                (
                    str(row["effective_time"]),
                    int(row["event_priority"]),
                    sequence,
                    int(row["id"]),
                    str(payload["status"]),
                )
            )
    jobs = conn.execute(
        """
        SELECT id, operation_log_id, target_status, effective_at
        FROM v102_status_bridge_jobs
        WHERE member_id=? AND group_area=?
          AND job_status IN ('pending', 'processing', 'applied', 'failed')
          AND applied_event_id IS NULL
          AND (
              caused_by_record_id IS NULL
              OR EXISTS (
                  SELECT 1 FROM violation_records r
                  WHERE r.id=v102_status_bridge_jobs.caused_by_record_id
                    AND r.is_withdrawn=0
              )
          )
        """,
        (member_id, group_area),
    ).fetchall()
    for job in jobs:
        if job["target_status"]:
            timeline.append(
                (
                    str(job["effective_at"]),
                    30,
                    int(job["operation_log_id"]),
                    int(job["id"]),
                    str(job["target_status"]),
                )
            )
    for effective_at, _, _, _, target_status in sorted(timeline):
        status = target_status
        if status == "最后警告":
            last_final_warning_time = effective_at
    return status, last_final_warning_time


def _attest_legacy_repair_snapshot(
    conn: sqlite3.Connection,
    *,
    audits: list[sqlite3.Row],
    old_members: dict[int, dict[str, object]],
    old_states: dict[tuple[int, str], dict[str, object]],
    snapshot_violation_rows: tuple[tuple[object, ...], ...],
    watermark: int,
    snapshot_has_v102_tables: bool,
) -> None:
    def fail(detail: str) -> None:
        raise MigrationError(f"切换前数据库快照证明失败：{detail}")

    if snapshot_has_v102_tables:
        fail("快照已包含 v102 表，不是可证明的迁移前数据库")
    current_violation_rows = tuple(
        tuple(row)
        for row in conn.execute(
            "SELECT * FROM violation_records WHERE id<=? ORDER BY id", (watermark,)
        )
    )
    if current_violation_rows != snapshot_violation_rows:
        fail("水位内违规记录与当前数据库不一致")

    stable_state_fields = (
        "status",
        "locked",
        "last_effective_violation_time",
        "last_deduct_time",
        "last_final_warning_time",
    )
    for audit in audits:
        member_id = int(audit["member_id"])
        group_area = str(audit["group_area"])
        key = (member_id, group_area)
        old_state = old_states.get(key)
        was_created = bool(int(audit["was_created"] or 0))
        old_counts = (
            int(audit["old_total_count"]),
            int(audit["old_deduct_count"]),
            int(audit["old_current_count"]),
        )
        if was_created:
            if old_state is not None or old_counts != (0, 0, 0):
                fail(f"新增范围存在性不一致：{member_id}/{group_area}")
            continue
        if old_state is None:
            fail(f"原有范围在快照中缺失：{member_id}/{group_area}")
        old_member = old_members.get(member_id)
        current_member = conn.execute(
            "SELECT * FROM members WHERE id=?", (member_id,)
        ).fetchone()
        if (
            old_member is None
            or current_member is None
            or str(old_member["qq_number"]) != str(current_member["qq_number"])
        ):
            fail(f"成员身份不一致：{member_id}/{group_area}")
        snapshot_counts = (
            int(old_state["total_count"] or 0),
            int(old_state["deduct_count"] or 0),
            int(old_state["current_count_cache"] or 0),
        )
        if snapshot_counts != old_counts:
            fail(f"迁移前次数与审计不一致：{member_id}/{group_area}")
        current_state = conn.execute(
            """
            SELECT * FROM member_group_states
            WHERE member_id=? AND group_area=?
            """,
            (member_id, group_area),
        ).fetchone()
        if current_state is None:
            fail(f"当前数据库范围缺失：{member_id}/{group_area}")
        if tuple(old_state[field] for field in stable_state_fields) != tuple(
            current_state[field] for field in stable_state_fields
        ):
            fail(f"状态或计时无法由当前迁移事实证明：{member_id}/{group_area}")


def repair_runtime_snapshots(
    database: str | Path,
    snapshot_database: str | Path,
    *,
    batch_id: str,
    backup_sha256: str,
) -> dict[str, object]:
    database_path = _require_file(Path(database), "数据库")
    snapshot_path = _require_file(Path(snapshot_database), "切换前数据库快照")
    migration_id = _validate_batch_id(batch_id)
    expected_digest = _validate_backup_sha256(backup_sha256)
    actual_digest = file_sha256(snapshot_path)
    if actual_digest != expected_digest:
        raise MigrationError(
            "切换前数据库快照 SHA-256 不匹配："
            f"expected={expected_digest} actual={actual_digest}"
        )

    snapshot = _read_only_connection(snapshot_path)
    try:
        _validate_legacy_database(snapshot)
        snapshot_has_v102_tables = any(
            name.startswith("v102_") for name in _table_names(snapshot)
        )
        snapshot_watermark = int(
            snapshot.execute(
                "SELECT COALESCE(MAX(id), 0) FROM violation_records"
            ).fetchone()[0]
        )
        old_states = {
            (int(row["member_id"]), str(row["group_area"])): dict(row)
            for row in snapshot.execute(
                "SELECT * FROM member_group_states"
            ).fetchall()
        }
        old_members = {
            int(row["id"]): dict(row)
            for row in snapshot.execute("SELECT * FROM members").fetchall()
        }
        snapshot_violation_rows = tuple(
            tuple(row)
            for row in snapshot.execute(
                "SELECT * FROM violation_records WHERE id<=? ORDER BY id",
                (snapshot_watermark,),
            )
        )
    finally:
        snapshot.close()

    conn = _write_connection(database_path)
    try:
        conn.execute("BEGIN IMMEDIATE")
        ensure_v102_schema(conn)
        checkpoint = conn.execute(
            "SELECT * FROM v102_migration_checkpoints WHERE batch_id=?",
            (migration_id,),
        ).fetchone()
        if checkpoint is None or checkpoint["status"] != "applied":
            raise MigrationError("只能修复已应用的迁移批次")
        if str(checkpoint["schema_version"]) not in REPAIRABLE_SCHEMA_VERSIONS:
            raise MigrationError(
                f"迁移批次版本不可修复：{checkpoint['schema_version']}"
            )
        if str(checkpoint["backup_sha256"]).lower() != actual_digest:
            raise MigrationError("迁移检查点记录的备份 SHA-256 与快照文件不一致")
        if int(checkpoint["cutover_record_watermark"] or 0) != snapshot_watermark:
            raise MigrationError(
                "迁移检查点水位与切换前数据库快照不一致："
                f"checkpoint={checkpoint['cutover_record_watermark']} "
                f"snapshot={snapshot_watermark}"
            )
        audits = conn.execute(
            """
            SELECT * FROM v102_baseline_audit
            WHERE batch_id=? ORDER BY id
            """,
            (migration_id,),
        ).fetchall()
        if not audits:
            raise MigrationError("迁移批次没有基线审计记录")
        _attest_legacy_repair_snapshot(
            conn,
            audits=list(audits),
            old_members=old_members,
            old_states=old_states,
            snapshot_violation_rows=snapshot_violation_rows,
            watermark=snapshot_watermark,
            snapshot_has_v102_tables=snapshot_has_v102_tables,
        )
        repaired = 0
        for audit in audits:
            key = (int(audit["member_id"]), str(audit["group_area"]))
            old = old_states.get(key)
            old_locked = int(old["locked"] or 0) if old is not None else 0
            old_effective = (
                old["last_effective_violation_time"] if old is not None else None
            )
            old_deduct_time = old["last_deduct_time"] if old is not None else None
            old_final = (
                old["last_final_warning_time"] if old is not None else None
            )
            conn.execute(
                """
                UPDATE v102_baseline_audit
                SET old_locked=?, old_last_effective_violation_time=?,
                    old_last_deduct_time=?, old_last_final_warning_time=?
                WHERE id=?
                """,
                (
                    old_locked,
                    old_effective,
                    old_deduct_time,
                    old_final,
                    audit["id"],
                ),
            )
            updated = conn.execute(
                """
                UPDATE v102_policy_state
                SET baseline_total_count=?, baseline_deduct_count=?,
                    baseline_current_count=?, baseline_raw_total=?,
                    baseline_record_watermark=?, baseline_locked=?,
                    baseline_status=?,
                    baseline_last_effective_violation_time=?,
                    baseline_last_deduct_time=?,
                    baseline_last_final_warning_time=?
                WHERE member_id=? AND group_area=?
                """,
                (
                    int(audit["old_total_count"]),
                    int(audit["old_deduct_count"]),
                    int(audit["old_current_count"]),
                    int(audit["new_total_count"])
                    - int(audit["new_baseline_adjustment"]),
                    int(checkpoint["cutover_record_watermark"]),
                    old_locked,
                    str(old["status"] if old is not None else "正常"),
                    old_effective,
                    old_deduct_time,
                    old_final,
                    audit["member_id"],
                    audit["group_area"],
                ),
            )
            if updated.rowcount != 1:
                raise MigrationError(
                    f"策略快照缺失：{audit['member_id']}/{audit['group_area']}"
                )
            repaired += 1
        foreign_errors = conn.execute("PRAGMA foreign_key_check").fetchall()
        if foreign_errors:
            raise MigrationError(f"快照修复产生外键错误：{len(foreign_errors)}")
        repaired_at = _time_text()
        conn.execute(
            """
            UPDATE v102_migration_checkpoints
            SET schema_version=?, updated_at=? WHERE batch_id=?
            """,
            (V102_SCHEMA_VERSION, repaired_at, migration_id),
        )
        conn.commit()
    except Exception:
        if conn.in_transaction:
            conn.rollback()
        raise
    finally:
        conn.close()
    return {
        "status": "snapshots_repaired",
        "batch_id": migration_id,
        "backup_sha256": actual_digest,
        "repaired_scopes": repaired,
    }


def logical_rollback(
    database: str | Path,
    batch_id: str,
    *,
    rolled_back_at: str | None = None,
) -> dict[str, object]:
    database_path = _require_file(Path(database), "数据库")
    migration_id = _validate_batch_id(batch_id)
    at = _time_text(rolled_back_at)
    conn = _write_connection(database_path)
    try:
        conn.execute("BEGIN IMMEDIATE")
        _validate_legacy_database(conn)
        if not REQUIRED_V102_TABLES <= _table_names(conn):
            raise MigrationError("v102 schema 不完整，不能逻辑回滚")
        missing_indexes = REQUIRED_V102_INDEXES - _index_names(conn)
        if missing_indexes:
            raise MigrationError(
                "v102 索引不完整，不能逻辑回滚："
                + ", ".join(sorted(missing_indexes))
            )
        for table, required_columns in REQUIRED_SNAPSHOT_COLUMNS.items():
            columns = {
                row["name"] for row in conn.execute(f"PRAGMA table_info({table})")
            }
            missing_columns = required_columns - columns
            if missing_columns:
                raise MigrationError(
                    "v102 快照字段不完整，不能逻辑回滚："
                    f"{table}({', '.join(sorted(missing_columns))})"
                )
        checkpoint = conn.execute(
            "SELECT * FROM v102_migration_checkpoints WHERE batch_id=?",
            (migration_id,),
        ).fetchone()
        if checkpoint is None:
            raise MigrationError(f"找不到迁移批次：{migration_id}")
        if checkpoint["status"] != "applied":
            raise MigrationError(f"迁移批次不是已应用状态：{checkpoint['status']}")
        if str(checkpoint["schema_version"]) != V102_SCHEMA_VERSION:
            raise MigrationError(
                "迁移批次尚未完成快照修复，禁止逻辑回滚："
                f"expected={V102_SCHEMA_VERSION} actual={checkpoint['schema_version']}"
            )
        invalid_snapshots = _policy_schema.invalid_v102_baseline_snapshots(
            conn, batch_id=migration_id
        )
        if invalid_snapshots:
            raise MigrationError(
                f"迁移快照语义校验失败，禁止逻辑回滚：{invalid_snapshots}"
            )
        audits = conn.execute(
            """
            SELECT * FROM v102_baseline_audit
            WHERE batch_id=? ORDER BY id
            """,
            (migration_id,),
        ).fetchall()
        if not audits:
            raise MigrationError("迁移批次没有基线审计记录")
        audits_by_scope = {
            (int(row["member_id"]), str(row["group_area"])): row
            for row in audits
        }
        policies = conn.execute(
            """
            SELECT * FROM v102_policy_state
            WHERE baseline_initialized_at>=?
               OR EXISTS (
                   SELECT 1 FROM v102_baseline_audit a
                   WHERE a.batch_id=?
                     AND a.member_id=v102_policy_state.member_id
                     AND a.group_area=v102_policy_state.group_area
               )
            ORDER BY member_id, group_area
            """,
            (checkpoint["cutover_at"], migration_id),
        ).fetchall()
        if not policies:
            raise MigrationError("迁移批次没有可回滚策略状态")
        policy_scopes = {
            (int(row["member_id"]), str(row["group_area"])) for row in policies
        }
        missing_audit_scopes = set(audits_by_scope) - policy_scopes
        if missing_audit_scopes:
            raise MigrationError(
                "迁移快照策略状态缺失，禁止逻辑回滚："
                f"{len(missing_audit_scopes)}"
            )

        for policy in policies:
            member_id = int(policy["member_id"])
            group_area = str(policy["group_area"])
            audit = audits_by_scope.get((member_id, group_area))
            state = conn.execute(
                """
                SELECT * FROM member_group_states
                WHERE member_id=? AND group_area=?
                """,
                (member_id, group_area),
            ).fetchone()
            if state is None:
                raise MigrationError(
                    f"回滚目标状态缺失：{member_id}/{group_area}"
                )
            raw_now, last_time = raw_effective_record_summary(
                conn, member_id, group_area
            )
            if audit is not None:
                old_total = int(audit["old_total_count"])
                old_deduct = int(audit["old_deduct_count"])
                old_locked = int(audit["old_locked"] or 0)
                old_last_effective = audit["old_last_effective_violation_time"]
                old_last_deduct = audit["old_last_deduct_time"]
                raw_at_snapshot = int(audit["new_total_count"]) - int(
                    audit["new_baseline_adjustment"]
                )
                record_watermark = int(checkpoint["cutover_record_watermark"])
                initialized_at = str(checkpoint["cutover_at"])
            else:
                old_total = int(policy["baseline_total_count"] or 0)
                old_deduct = int(policy["baseline_deduct_count"] or 0)
                old_locked = int(policy["baseline_locked"] or 0)
                old_last_effective = policy[
                    "baseline_last_effective_violation_time"
                ]
                old_last_deduct = policy["baseline_last_deduct_time"]
                raw_at_snapshot = int(policy["baseline_raw_total"] or 0)
                record_watermark = int(policy["baseline_record_watermark"] or 0)
                initialized_at = str(
                    policy["baseline_initialized_at"] or checkpoint["cutover_at"]
                )
            restored_total = max(
                0,
                old_total + raw_now - raw_at_snapshot,
            )
            restored_deduct = old_deduct
            restored_current = max(0, restored_total - restored_deduct)
            restored_adjustment = restored_total - raw_now
            raw_changed = conn.execute(
                """
                SELECT 1 FROM violation_records
                WHERE member_id=? AND group_area=?
                  AND (id>? OR (id<=? AND updated_at>?))
                LIMIT 1
                """,
                (
                    member_id,
                    group_area,
                    record_watermark,
                    record_watermark,
                    initialized_at,
                ),
            ).fetchone()
            restored_last_effective = (
                last_time if raw_changed is not None else old_last_effective
            )
            restored_last_deduct = (
                state["last_deduct_time"]
                if raw_changed is not None
                else old_last_deduct
            )
            restored_status, restored_last_final_warning = _status_after_manual_inputs(
                conn,
                member_id=member_id,
                group_area=group_area,
                baseline_status=str(policy["baseline_status"] or "正常"),
                baseline_last_final_warning_time=policy[
                    "baseline_last_final_warning_time"
                ],
            )
            no_cycle_reason = _rollback_reason(
                restored_status, restored_current
            )
            restored_locked = (
                1
                if restored_status in TERMINAL_STATUSES
                else (
                    0
                    if restored_status != str(policy["baseline_status"] or "正常")
                    else old_locked
                )
            )

            conn.execute(
                """
                UPDATE v102_policy_state
                SET policy_tag='none', slow_level=0,
                    v102_operation_count=0, baseline_adjustment=?,
                    baseline_deduct_count=?, active_cycle_id=NULL,
                    no_cycle_reason=?, pending_action_type=NULL,
                    last_reason='logical_rollback',
                    state_version=state_version+1, updated_at=?
                WHERE member_id=? AND group_area=?
                """,
                (
                    restored_adjustment,
                    restored_deduct,
                    no_cycle_reason,
                    at,
                    member_id,
                    group_area,
                ),
            )
            conn.execute(
                """
                UPDATE member_group_states
                SET status=?, locked=?, total_count=?, deduct_count=?, current_count_cache=?,
                    last_effective_violation_time=?, last_deduct_time=?,
                    last_final_warning_time=?, updated_at=?
                WHERE member_id=? AND group_area=?
                """,
                (
                    restored_status,
                    restored_locked,
                    restored_total,
                    restored_deduct,
                    restored_current,
                    restored_last_effective,
                    restored_last_deduct,
                    restored_last_final_warning,
                    at,
                    member_id,
                    group_area,
                ),
            )
            conn.execute(
                """
                UPDATE v102_policy_cycles
                SET status='cancelled', closed_reason='logical_rollback',
                    updated_at=?
                WHERE member_id=? AND group_area=?
                  AND status IN ('active', 'pending_decision')
                """,
                (at, member_id, group_area),
            )
            conn.execute(
                """
                UPDATE v102_pending_actions
                SET status='cancelled', updated_at=?
                WHERE member_id=? AND group_area=? AND status='pending'
                """,
                (at, member_id, group_area),
            )
            event_ids = [
                int(row[0])
                for row in conn.execute(
                    """
                    SELECT id FROM v102_policy_events
                    WHERE member_id=? AND group_area=? AND is_effective=1
                    """,
                    (member_id, group_area),
                )
            ]
            if event_ids:
                placeholders = ",".join("?" for _ in event_ids)
                conn.execute(
                    f"""
                    UPDATE v102_notification_attempts
                    SET status='cancelled', finished_at=?,
                        detail='logical_rollback', updated_at=?
                    WHERE status='sending'
                      AND outbox_id IN (
                          SELECT id FROM v102_notification_outbox
                          WHERE event_id IN ({placeholders})
                            AND status='sending'
                      )
                    """,
                    (at, at, *event_ids),
                )
                conn.execute(
                    f"""
                    UPDATE v102_notification_outbox
                    SET status='cancelled', last_error='logical_rollback',
                        updated_at=?
                    WHERE event_id IN ({placeholders})
                      AND status IN ('pending', 'sending', 'failed')
                    """,
                    (at, *event_ids),
                )
            conn.execute(
                """
                UPDATE v102_policy_events SET is_effective=0
                WHERE member_id=? AND group_area=? AND is_effective=1
                """,
                (member_id, group_area),
            )
            _insert_operation_log(
                conn,
                member_id=member_id,
                group_area=group_area,
                operation_type="v1.0.2beta逻辑回滚",
                before={
                    "total_count": int(state["total_count"] or 0),
                    "deduct_count": int(state["deduct_count"] or 0),
                    "current_count": int(state["current_count_cache"] or 0),
                    "status": str(state["status"]),
                },
                after={
                    "total_count": restored_total,
                    "deduct_count": restored_deduct,
                    "current_count": restored_current,
                    "preserved_raw_records": raw_now,
                    "status": restored_status,
                },
                batch_id=migration_id,
                at=at,
            )

        conn.execute(
            """
            UPDATE v102_migration_checkpoints
            SET status='rolled_back', updated_at=? WHERE batch_id=?
            """,
            (at, migration_id),
        )
        foreign_errors = conn.execute("PRAGMA foreign_key_check").fetchall()
        if foreign_errors:
            raise MigrationError(f"回滚产生外键错误：{len(foreign_errors)}")
        conn.commit()
    except Exception:
        if conn.in_transaction:
            conn.rollback()
        raise
    finally:
        conn.close()

    return {
        "status": "rolled_back",
        "batch_id": migration_id,
        "rolled_back_at": at,
        "restored_scopes": len(policies),
    }


def verify_database(database: str | Path) -> dict[str, object]:
    database_path = _require_file(Path(database), "数据库")
    conn = _read_only_connection(database_path)
    try:
        tables = _table_names(conn)
        indexes = _index_names(conn)
        missing_tables = sorted(REQUIRED_V102_TABLES - tables)
        missing_indexes = sorted(REQUIRED_V102_INDEXES - indexes)
        missing_columns = {
            table: sorted(
                required
                - {
                    row["name"]
                    for row in conn.execute(f"PRAGMA table_info({table})")
                }
            )
            for table, required in REQUIRED_SNAPSHOT_COLUMNS.items()
            if table in tables
        }
        missing_columns = {
            table: columns
            for table, columns in missing_columns.items()
            if columns
        }
        integrity = conn.execute("PRAGMA integrity_check").fetchone()[0]
        foreign_errors = [list(row) for row in conn.execute("PRAGMA foreign_key_check")]
        errors: list[str] = []
        if integrity != "ok":
            errors.append(f"integrity_check={integrity}")
        if foreign_errors:
            errors.append(f"foreign_key_errors={len(foreign_errors)}")
        if missing_tables:
            errors.append(f"missing_tables={','.join(missing_tables)}")
        if missing_indexes:
            errors.append(f"missing_indexes={','.join(missing_indexes)}")
        if missing_columns:
            errors.append(
                "missing_columns="
                + ",".join(
                    f"{table}({','.join(columns)})"
                    for table, columns in sorted(missing_columns.items())
                )
            )

        equation_errors = 0
        active_cycle_duplicates = 0
        invalid_operation_counts = 0
        invalid_active_links = 0
        invalid_policy_tags = 0
        invalid_baseline_snapshots = 0
        checkpoint_schema_versions: list[str] = []
        if "v102_migration_checkpoints" in tables:
            checkpoint_schema_versions = sorted(
                {
                    str(row[0])
                    for row in conn.execute(
                        """
                        SELECT schema_version FROM v102_migration_checkpoints
                        WHERE status IN ('applied', 'rolled_back')
                        """
                    )
                }
            )
            invalid_versions = [
                version
                for version in checkpoint_schema_versions
                if version != V102_SCHEMA_VERSION
            ]
            if invalid_versions:
                errors.append(
                    "checkpoint_schema_versions=" + ",".join(invalid_versions)
                )
        if "v102_policy_state" in tables:
            invalid_policy_tags = int(
                conn.execute(
                    """
                    SELECT COUNT(*) FROM v102_policy_state
                    WHERE policy_tag NOT IN ('none', 'slow', 'stop')
                    """
                ).fetchone()[0]
            )
        missing_unique_constraints = [
            f"{table}({','.join(columns)})"
            for table, columns in REQUIRED_UNIQUE_CONSTRAINTS.items()
            if table not in tables or not _has_unique_constraint(conn, table, columns)
        ]
        if not missing_tables and not missing_columns:
            states = conn.execute(
                """
                SELECT p.*, s.total_count, s.deduct_count,
                       s.current_count_cache
                FROM v102_policy_state p
                JOIN member_group_states s
                  ON s.member_id=p.member_id AND s.group_area=p.group_area
                """
            ).fetchall()
            for state in states:
                raw_total, _ = raw_effective_record_summary(
                    conn, int(state["member_id"]), str(state["group_area"])
                )
                expected_total = max(
                    0, raw_total + int(state["baseline_adjustment"] or 0)
                )
                expected_current = max(
                    0, expected_total - int(state["deduct_count"] or 0)
                )
                if (
                    int(state["total_count"] or 0) != expected_total
                    or int(state["current_count_cache"] or 0)
                    != expected_current
                ):
                    equation_errors += 1
                if not 0 <= int(state["v102_operation_count"] or 0) <= 5:
                    invalid_operation_counts += 1
                if state["active_cycle_id"] is not None:
                    active = conn.execute(
                        """
                        SELECT 1 FROM v102_policy_cycles
                        WHERE id=? AND member_id=? AND group_area=?
                          AND status IN ('active', 'pending_decision')
                        """,
                        (
                            state["active_cycle_id"],
                            state["member_id"],
                            state["group_area"],
                        ),
                    ).fetchone()
                    if active is None:
                        invalid_active_links += 1
            active_cycle_duplicates = int(
                conn.execute(
                    """
                    SELECT COUNT(*) FROM (
                        SELECT member_id, group_area
                        FROM v102_policy_cycles
                        WHERE status IN ('active', 'pending_decision')
                        GROUP BY member_id, group_area HAVING COUNT(*) > 1
                    )
                    """
                ).fetchone()[0]
            )
            invalid_baseline_snapshots = int(
                conn.execute(
                    """
                    SELECT COUNT(*)
                    FROM v102_baseline_audit a
                    JOIN v102_migration_checkpoints c ON c.batch_id=a.batch_id
                    JOIN v102_policy_state p
                      ON p.member_id=a.member_id AND p.group_area=a.group_area
                    WHERE p.baseline_total_count!=a.old_total_count
                       OR p.baseline_deduct_count!=a.old_deduct_count
                       OR p.baseline_current_count!=a.old_current_count
                       OR p.baseline_raw_total!=(
                           a.new_total_count-a.new_baseline_adjustment
                       )
                       OR p.baseline_record_watermark!=c.cutover_record_watermark
                       OR p.baseline_locked!=a.old_locked
                       OR p.baseline_last_effective_violation_time
                          IS NOT a.old_last_effective_violation_time
                       OR p.baseline_last_deduct_time
                          IS NOT a.old_last_deduct_time
                       OR p.baseline_last_final_warning_time
                          IS NOT a.old_last_final_warning_time
                    """
                ).fetchone()[0]
            )
        if equation_errors:
            errors.append(f"count_equation_errors={equation_errors}")
        if active_cycle_duplicates:
            errors.append(f"active_cycle_duplicates={active_cycle_duplicates}")
        if invalid_operation_counts:
            errors.append(f"invalid_operation_counts={invalid_operation_counts}")
        if invalid_active_links:
            errors.append(f"invalid_active_links={invalid_active_links}")
        if invalid_policy_tags:
            errors.append(f"invalid_policy_tags={invalid_policy_tags}")
        if invalid_baseline_snapshots:
            errors.append(
                f"invalid_baseline_snapshots={invalid_baseline_snapshots}"
            )
        if missing_unique_constraints:
            errors.append(
                "missing_unique_constraints=" + ",".join(missing_unique_constraints)
            )
        return {
            "ok": not errors,
            "integrity_check": integrity,
            "foreign_key_errors": foreign_errors,
            "missing_tables": missing_tables,
            "missing_indexes": missing_indexes,
            "missing_columns": missing_columns,
            "count_equation_errors": equation_errors,
            "active_cycle_duplicates": active_cycle_duplicates,
            "invalid_operation_counts": invalid_operation_counts,
            "invalid_active_links": invalid_active_links,
            "invalid_policy_tags": invalid_policy_tags,
            "invalid_baseline_snapshots": invalid_baseline_snapshots,
            "checkpoint_schema_versions": checkpoint_schema_versions,
            "missing_unique_constraints": missing_unique_constraints,
            "errors": errors,
        }
    finally:
        conn.close()


def build_parser() -> argparse.ArgumentParser:
    parser = argparse.ArgumentParser(
        description="Guarded v1.0.2beta baseline migration"
    )
    parser.add_argument("--database", required=True, type=Path)
    parser.add_argument("--snapshot-database", type=Path)
    parser.add_argument("--baseline", type=Path)
    parser.add_argument("--backup-sha256")
    parser.add_argument("--batch-id")
    parser.add_argument("--cutover-at")
    modes = parser.add_mutually_exclusive_group(required=True)
    modes.add_argument("--dry-run", action="store_true")
    modes.add_argument("--apply", action="store_true")
    modes.add_argument("--logical-rollback", action="store_true")
    modes.add_argument("--repair-snapshots", action="store_true")
    modes.add_argument("--verify", action="store_true")
    return parser


def main(argv: list[str] | None = None) -> int:
    parser = build_parser()
    args = parser.parse_args(argv)
    try:
        if args.dry_run:
            if args.baseline is None:
                raise MigrationError("--dry-run 必须提供 --baseline")
            result = dry_run(args.database, args.baseline)
        elif args.apply:
            if args.baseline is None:
                raise MigrationError("--apply 必须提供 --baseline")
            if args.backup_sha256 is None:
                raise MigrationError("--apply 必须提供 --backup-sha256")
            if args.snapshot_database is None:
                raise MigrationError("--apply 必须提供 --snapshot-database")
            result = apply_migration(
                args.database,
                args.baseline,
                snapshot_database=args.snapshot_database,
                backup_sha256=args.backup_sha256,
                cutover_at=args.cutover_at,
                batch_id=args.batch_id,
            )
        elif args.logical_rollback:
            if not args.batch_id:
                raise MigrationError("--logical-rollback 必须提供 --batch-id")
            result = logical_rollback(args.database, args.batch_id)
        elif args.repair_snapshots:
            if args.snapshot_database is None:
                raise MigrationError("--repair-snapshots 必须提供 --snapshot-database")
            if not args.batch_id:
                raise MigrationError("--repair-snapshots 必须提供 --batch-id")
            if args.backup_sha256 is None:
                raise MigrationError("--repair-snapshots 必须提供 --backup-sha256")
            result = repair_runtime_snapshots(
                args.database,
                args.snapshot_database,
                batch_id=args.batch_id,
                backup_sha256=args.backup_sha256,
            )
        else:
            result = verify_database(args.database)
            if not result["ok"]:
                print(json.dumps(result, ensure_ascii=False, indent=2, sort_keys=True))
                return 2
    except (MigrationError, sqlite3.Error) as exc:
        print(f"迁移失败：{exc}", file=sys.stderr)
        return 2
    print(json.dumps(result, ensure_ascii=False, indent=2, sort_keys=True))
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
