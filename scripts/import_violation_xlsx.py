#!/usr/bin/env python3
from __future__ import annotations

import argparse
import csv
import json
import re
import sys
import unicodedata
from collections import defaultdict
from dataclasses import dataclass
from datetime import date, datetime, time
from pathlib import Path
from zipfile import ZIP_DEFLATED, ZipFile

from dotenv import load_dotenv
from openpyxl import load_workbook
from openpyxl.utils.datetime import from_excel

PROJECT_DIR = Path(__file__).resolve().parents[1]
if str(PROJECT_DIR) not in sys.path:
    sys.path.insert(0, str(PROJECT_DIR))

load_dotenv(PROJECT_DIR / ".env")

from plugins.violation_record.config import CONFIG  # noqa: E402
from plugins.violation_record.admin_resolver import resolve_admin_by_name  # noqa: E402
from plugins.violation_record.db import backup_database, connect, dump_json, ensure_schema_extensions, now_str  # noqa: E402
from plugins.violation_record.validators import normalize_time  # noqa: E402


SOURCE_AREAS = {
    "蜂巢": ("蜂巢", 0, "明确分区工作表"),
    "蜂窝": ("蜂窝", 0, "明确分区工作表"),
    "蜂箱": ("蜂箱", 0, "明确分区工作表"),
    "低频小于三": ("蜂巢", 1, "按旧表结构推断为蜂巢低频历史数据"),
    "暂存": ("蜂巢", 1, "按旧表结构推断为蜂巢暂存历史数据；只导入有可解析时间的记录"),
}

CONTROL_CHARS = dict.fromkeys(
    map(
        ord,
        [
            "\u200b",
            "\u200c",
            "\u200d",
            "\u202a",
            "\u202b",
            "\u202c",
            "\u202d",
            "\u202e",
            "\u2066",
            "\u2067",
            "\u2068",
            "\u2069",
            "\ufeff",
        ],
    ),
    None,
)


@dataclass(frozen=True)
class ParsedMember:
    qq_number: str
    qq_nickname: str
    raw_text: str


@dataclass(frozen=True)
class ParsedRecord:
    area: str
    area_inferred: int
    area_note: str
    source_sheet: str
    source_row: int
    source_col: int
    member: ParsedMember
    sheet_current_count: int | None
    violation_time: str
    judgement: str
    action: str
    is_countable: int
    count_delta: int
    raw_record_text: str | None
    handler_name: str | None


def clean_text(value: object) -> str:
    if value is None:
        return ""
    text = str(value).replace("\r\n", "\n").replace("\r", "\n")
    text = text.translate(CONTROL_CHARS)
    text = unicodedata.normalize("NFKC", text)
    return "\n".join(part.strip() for part in text.splitlines()).strip()


def parse_qq(value: object) -> list[str]:
    text = unicodedata.normalize("NFKC", clean_text(value))
    numbers = re.findall(r"\d{5,12}", text)
    cleaned: list[str] = []
    for number in numbers:
        if number.startswith("0") and len(number) > 10:
            number = number.lstrip("0") or "0"
        cleaned.append(number)
    return cleaned


def parse_member(value: object) -> ParsedMember | None:
    raw = clean_text(value)
    numbers = parse_qq(raw)
    if not raw or not numbers:
        return None
    qq_number = numbers[-1]
    nickname = raw
    for number in numbers:
        nickname = re.sub(rf"[（(]?\s*0*{re.escape(number)}\s*[）)]?", "", nickname)
    nickname = re.sub(r"\d{5,12}", "", nickname)
    nickname = nickname.replace("（", "").replace("）", "").replace("(", "").replace(")", "")
    nickname = " ".join(part.strip() for part in nickname.splitlines() if part.strip()).strip()
    return ParsedMember(qq_number=qq_number, qq_nickname=nickname or "未知昵称", raw_text=raw)


def parse_current_count(value: object) -> int | None:
    if value in (None, ""):
        return None
    if isinstance(value, (int, float)):
        return max(0, int(value))
    text = clean_text(value)
    if not text:
        return None
    match = re.search(r"\d+", text)
    return int(match.group(0)) if match else None


def parse_excel_datetime(value: object) -> str | None:
    if value in (None, ""):
        return None
    dt: datetime | None = None
    if isinstance(value, datetime):
        dt = value
    elif isinstance(value, date):
        dt = datetime.combine(value, time.min)
    elif isinstance(value, (int, float)):
        if 20000 <= float(value) <= 70000:
            try:
                dt = from_excel(value)
            except (TypeError, ValueError, OverflowError):
                return None
    else:
        text = clean_text(value)
        text = re.sub(r"^(\d{4}/\d{1,2}/\d{1,2})/(\d{1,2}:\d{1,2})$", r"\1 \2", text)
        parsed = normalize_time(text)
        if parsed:
            return parsed
        date_only = re.fullmatch(r"(\d{4})[/-](\d{1,2})[/-](\d{1,2})", text)
        if date_only:
            try:
                return datetime(
                    int(date_only.group(1)),
                    int(date_only.group(2)),
                    int(date_only.group(3)),
                    0,
                    0,
                ).strftime("%Y-%m-%d %H:%M:%S")
            except ValueError:
                return None
        short_date = re.fullmatch(r"(\d{1,2})[/-](\d{1,2})", text)
        if short_date:
            try:
                return datetime(datetime.now().year, int(short_date.group(1)), int(short_date.group(2)), 0, 0).strftime("%Y-%m-%d %H:%M:%S")
            except ValueError:
                return None
    if not dt:
        return None
    return dt.replace(second=0, microsecond=0).strftime("%Y-%m-%d %H:%M:%S")


def split_record_text(value: object) -> tuple[str, str, int, int, str | None]:
    raw = clean_text(value) or None
    if not raw:
        return "历史违规记录", "导入记录", 1, 1, None
    text = raw.replace("，", ",")
    if "警告" in text and not any(word in text for word in ("最后警告", "拉黑", "移出", "退群")):
        return text, "警告", 0, 0, raw
    action_words = ["禁言", "口球", "质询", "最后警告", "退群", "移出", "拉黑", "踢出", "撤回"]
    matched = next((word for word in action_words if word in text), None)
    if matched:
        return text, text if matched in {"禁言", "口球"} else matched, 1, 1, raw
    return text, "导入记录", 1, 1, raw


def style_safe_workbook_path(path: Path) -> Path:
    try:
        wb = load_workbook(path, read_only=True, data_only=True)
        wb.close()
        return path
    except Exception:
        tmp = Path("/tmp") / f"{path.stem}_no_styles{path.suffix}"
        with ZipFile(path, "r") as zin, ZipFile(tmp, "w", ZIP_DEFLATED) as zout:
            for item in zin.infolist():
                if item.filename in {"xl/styles.xml", "xl/theme/theme1.xml"}:
                    continue
                zout.writestr(item, zin.read(item.filename))
        return tmp


def get_or_create_member(conn, member: ParsedMember) -> int:
    ts = now_str()
    conn.execute(
        """
        INSERT INTO members(qq_number, qq_nickname, aliases, created_at, updated_at)
        VALUES(?, ?, '[]', ?, ?)
        ON CONFLICT(qq_number) DO UPDATE SET
            qq_nickname=CASE
                WHEN excluded.qq_nickname IS NOT NULL AND excluded.qq_nickname != '' THEN excluded.qq_nickname
                ELSE members.qq_nickname
            END,
            updated_at=excluded.updated_at
        """,
        (member.qq_number, member.qq_nickname, ts, ts),
    )
    return int(conn.execute("SELECT id FROM members WHERE qq_number=?", (member.qq_number,)).fetchone()["id"])


def ensure_state(conn, member_id: int, area: str) -> int:
    ts = now_str()
    conn.execute(
        """
        INSERT INTO member_group_states(member_id, group_area, created_at, updated_at)
        VALUES(?, ?, ?, ?)
        ON CONFLICT(member_id, group_area) DO NOTHING
        """,
        (member_id, area, ts, ts),
    )
    return int(conn.execute("SELECT id FROM member_group_states WHERE member_id=? AND group_area=?", (member_id, area)).fetchone()["id"])


def parse_record_sheets(workbook_path: Path) -> tuple[list[ParsedRecord], dict[str, object]]:
    wb = load_workbook(workbook_path, read_only=True, data_only=True)
    records: list[ParsedRecord] = []
    report: dict[str, object] = {
        "workbook_used": str(workbook_path),
        "sheets": {},
        "skipped": [],
    }
    try:
        for sheet, (area, inferred, note) in SOURCE_AREAS.items():
            if sheet not in wb.sheetnames:
                report["skipped"].append({"sheet": sheet, "reason": "工作表不存在"})
                continue
            ws = wb[sheet]
            sheet_report = {
                "area": area,
                "area_inferred": bool(inferred),
                "members_seen": 0,
                "records_seen": 0,
                "records_parsed": 0,
                "rows_without_time": 0,
                "rows_without_member": 0,
            }
            report["sheets"][sheet] = sheet_report
            rows = list(ws.iter_rows(values_only=True))
            if len(rows) < 4:
                continue
            current_admins = list(rows[0])
            historical_admins = list(rows[1]) if sheet in {"低频小于三", "暂存"} and len(rows) > 1 else []
            start_index = 3 if sheet in {"蜂巢", "蜂窝", "蜂箱"} else 3
            row_index = start_index
            while row_index < len(rows):
                member_row = list(rows[row_index])
                time_row = list(rows[row_index + 1]) if row_index + 1 < len(rows) else []
                excel_row = row_index + 1
                member = parse_member(member_row[1] if len(member_row) > 1 else None)
                if not member:
                    if any(v not in (None, "") for v in member_row):
                        sheet_report["rows_without_member"] += 1
                    row_index += 2
                    continue
                sheet_report["members_seen"] += 1
                current_count = parse_current_count(member_row[2] if len(member_row) > 2 else None)
                max_cols = max(len(member_row), len(time_row))
                for col_index in range(4, max_cols + 1):
                    value = member_row[col_index - 1] if col_index - 1 < len(member_row) else None
                    time_value = time_row[col_index - 1] if col_index - 1 < len(time_row) else None
                    raw_record = clean_text(value) or None
                    has_content = bool(raw_record) or time_value not in (None, "")
                    if not has_content:
                        continue
                    sheet_report["records_seen"] += 1
                    violation_time = parse_excel_datetime(time_value)
                    if not violation_time:
                        sheet_report["rows_without_time"] += 1
                        report["skipped"].append(
                            {
                                "sheet": sheet,
                                "row": excel_row,
                                "col": col_index,
                                "member": member.raw_text,
                                "reason": "没有可解析时间",
                                "raw_record_text": raw_record,
                                "raw_time": clean_text(time_value),
                            }
                        )
                        continue
                    judgement, action, is_countable, count_delta, raw_text = split_record_text(value)
                    handler_name = None
                    if col_index - 1 < len(current_admins):
                        handler_name = clean_text(current_admins[col_index - 1]) or None
                    if not handler_name and col_index - 1 < len(historical_admins):
                        handler_name = clean_text(historical_admins[col_index - 1]) or None
                    records.append(
                        ParsedRecord(
                            area=area,
                            area_inferred=inferred,
                            area_note=note,
                            source_sheet=sheet,
                            source_row=excel_row,
                            source_col=col_index,
                            member=member,
                            sheet_current_count=current_count,
                            violation_time=violation_time,
                            judgement=judgement,
                            action=action,
                            is_countable=is_countable,
                            count_delta=count_delta,
                            raw_record_text=raw_text,
                            handler_name=handler_name,
                        )
                    )
                    sheet_report["records_parsed"] += 1
                row_index += 2
    finally:
        wb.close()
    return records, report


def parse_test_sheet(workbook_path: Path) -> tuple[list[dict[str, object]], list[dict[str, object]]]:
    wb = load_workbook(workbook_path, read_only=True, data_only=True)
    updates: list[dict[str, object]] = []
    skipped: list[dict[str, object]] = []
    try:
        if "测试" not in wb.sheetnames:
            return updates, skipped
        ws = wb["测试"]
        for row_index, row in enumerate(ws.iter_rows(min_row=3, values_only=True), start=3):
            qq_values = parse_qq(row[0] if len(row) > 0 else None)
            nickname = clean_text(row[1] if len(row) > 1 else None) or "未知昵称"
            consultation_text = clean_text(row[4] if len(row) > 4 else None)
            opinion = clean_text(row[5] if len(row) > 5 else None)
            blacklist = clean_text(row[6] if len(row) > 6 else None)
            remark = clean_text(row[7] if len(row) > 7 else None)
            if not qq_values:
                continue
            area = next((item for item in ("蜂巢", "蜂窝", "蜂箱") if item in consultation_text), None)
            if not area:
                skipped.append({"sheet": "测试", "row": row_index, "qq": "/".join(qq_values), "reason": "缺少明确分区"})
                continue
            time_text = consultation_text
            for item in ("蜂巢", "蜂窝", "蜂箱"):
                time_text = time_text.replace(item, "")
            consultation_time = parse_excel_datetime(time_text) or normalize_time(time_text)
            if not consultation_time:
                skipped.append({"sheet": "测试", "row": row_index, "qq": "/".join(qq_values), "reason": "质询时间无法解析"})
                continue
            updates.append(
                {
                    "row": row_index,
                    "qq_number": qq_values[0],
                    "qq_nickname": nickname,
                    "area": area,
                    "consultation_time": consultation_time,
                    "result": opinion or ("已拉黑" if blacklist == "是" else "导入质询记录"),
                    "status_after": "已拉黑" if blacklist == "是" else "已质询",
                    "locked": 1 if blacklist == "是" else 0,
                    "remark": remark or "无",
                    "raw_record_text": consultation_text,
                }
            )
    finally:
        wb.close()
    return updates, skipped


def import_record_remark(item: ParsedRecord) -> str:
    parts = [f"源表：{item.source_sheet}", item.area_note]
    if item.handler_name:
        parts.append(f"历史处理人：{item.handler_name}")
    return "；".join(part for part in parts if part)


def resolve_import_handler_id(handler_name: str | None) -> int | None:
    if not handler_name:
        return None
    status, admin = resolve_admin_by_name(handler_name)
    if status == "ok":
        return int(admin["id"])
    return None


def record_exists(conn, member_id: int, item: ParsedRecord, remark: str) -> bool:
    existing = conn.execute(
        """
        SELECT 1 FROM violation_records
        WHERE member_id=? AND group_area=? AND violation_time=? AND judgement=? AND action=? AND remark=?
            AND is_withdrawn=0
        LIMIT 1
        """,
        (member_id, item.area, item.violation_time, item.judgement, item.action, remark),
    ).fetchone()
    return existing is not None


def consultation_exists(conn, member_id: int, update: dict[str, object]) -> bool:
    existing = conn.execute(
        """
        SELECT 1
        FROM consultation_records
        WHERE member_id=? AND group_area=? AND consultation_type='质询'
            AND consultation_time=? AND result=? AND status_after=?
        LIMIT 1
        """,
        (
            member_id,
            update["area"],
            update["consultation_time"],
            update["result"],
            update["status_after"],
        ),
    ).fetchone()
    return existing is not None


def sync_imported_states(conn, member_area_info: dict[tuple[int, str], dict[str, object]]) -> list[dict[str, object]]:
    synced: list[dict[str, object]] = []
    sync_time = now_str()
    for (member_id, area), info in member_area_info.items():
        state_id = ensure_state(conn, member_id, area)
        effective_total = int(
            conn.execute(
                """
                SELECT COALESCE(SUM(count_delta), 0) AS c, MAX(violation_time) AS last_time
                FROM violation_records
                WHERE member_id=? AND group_area=? AND is_withdrawn=0 AND is_test=0 AND is_countable=1
                """,
                (member_id, area),
            ).fetchone()["c"]
            or 0
        )
        last_time = conn.execute(
            """
            SELECT MAX(violation_time) AS t
            FROM violation_records
            WHERE member_id=? AND group_area=? AND is_withdrawn=0 AND is_test=0 AND is_countable=1
            """,
            (member_id, area),
        ).fetchone()["t"]
        imported_counts = [int(v) for v in info.get("sheet_current_counts", []) if v is not None]
        desired_current = max(imported_counts) if imported_counts else effective_total
        desired_current = max(0, min(desired_current, effective_total)) if effective_total else 0
        deduct_count = max(0, effective_total - desired_current)
        conn.execute(
            """
            UPDATE member_group_states
            SET total_count=?, deduct_count=?, current_count_cache=?, last_effective_violation_time=?,
                last_deduct_time=?, updated_at=?
            WHERE id=?
            """,
            (
                effective_total,
                deduct_count,
                desired_current,
                last_time,
                sync_time,
                sync_time,
                state_id,
            ),
        )
        synced.append(
            {
                "member_id": member_id,
                "area": area,
                "effective_total": effective_total,
                "deduct_count": deduct_count,
                "current_count": desired_current,
                "source_sheets": sorted(set(info.get("source_sheets", []))),
            }
        )
    return synced


def import_records(source: Path, dry_run: bool = False) -> dict[str, object]:
    workbook_path = style_safe_workbook_path(source)
    records, parse_report = parse_record_sheets(workbook_path)
    test_updates, test_skipped = parse_test_sheet(workbook_path)
    source_file = source.name
    batch_id = datetime.now().strftime("%Y%m%d_%H%M%S")
    report_dir = PROJECT_DIR / "import_reports"
    report_dir.mkdir(parents=True, exist_ok=True)
    backup_path = None if dry_run else backup_database("before_xlsx_import")
    summary: dict[str, object] = {
        "source": str(source),
        "workbook_loaded": str(workbook_path),
        "batch_id": batch_id,
        "dry_run": dry_run,
        "backup_path": str(backup_path) if backup_path else None,
        "parsed_records": len(records),
        "inserted_records": 0,
        "duplicate_records": 0,
        "members_touched": 0,
        "states_synced": 0,
        "consultations_inserted": 0,
        "consultations_duplicate": 0,
        "parse_report": parse_report,
        "test_sheet_skipped": test_skipped,
    }
    if dry_run:
        return summary

    member_area_info: dict[tuple[int, str], dict[str, object]] = defaultdict(
        lambda: {"sheet_current_counts": [], "source_sheets": [], "area_inferred": 0, "area_note": ""}
    )
    touched_members: set[int] = set()
    imported_rows_for_csv: list[dict[str, object]] = []
    ts = now_str()
    with connect() as conn:
        ensure_schema_extensions(conn)
        conn.execute("CREATE INDEX IF NOT EXISTS idx_violation_member_area_time ON violation_records(member_id, group_area, violation_time)")
        conn.execute("CREATE INDEX IF NOT EXISTS idx_consultation_member_area_time ON consultation_records(member_id, group_area, consultation_time)")
        for item in records:
            member_id = get_or_create_member(conn, item.member)
            touched_members.add(member_id)
            ensure_state(conn, member_id, item.area)
            info = member_area_info[(member_id, item.area)]
            info["sheet_current_counts"].append(item.sheet_current_count)
            info["source_sheets"].append(item.source_sheet)
            info["area_inferred"] = max(int(info["area_inferred"]), item.area_inferred)
            info["area_note"] = item.area_note if item.area_inferred else info["area_note"]
            remark = import_record_remark(item)
            if record_exists(conn, member_id, item, remark):
                summary["duplicate_records"] = int(summary["duplicate_records"]) + 1
                continue
            handler_admin_id = resolve_import_handler_id(item.handler_name)
            conn.execute(
                """
                INSERT INTO violation_records(
                    member_id, group_area, violation_time, judgement, action, handler_admin_id, recorder_admin_id,
                    remark, is_countable, count_delta, is_withdrawn, is_test, created_at, updated_at
                )
                VALUES(?, ?, ?, ?, ?, ?, NULL, ?, ?, ?, 0, 0, ?, ?)
                """,
                (
                    member_id,
                    item.area,
                    item.violation_time,
                    item.judgement,
                    item.action,
                    handler_admin_id,
                    remark,
                    item.is_countable,
                    item.count_delta,
                    ts,
                    ts,
                ),
            )
            summary["inserted_records"] = int(summary["inserted_records"]) + 1
            imported_rows_for_csv.append(
                {
                    "group_area": item.area,
                    "qq_number": item.member.qq_number,
                    "qq_nickname": item.member.qq_nickname,
                    "violation_time": item.violation_time,
                    "judgement": item.judgement,
                    "action": item.action,
                    "sheet": item.source_sheet,
                    "row": item.source_row,
                    "col": item.source_col,
                    "handler_name": item.handler_name,
                    "area_inferred": item.area_inferred,
                }
            )

        synced = sync_imported_states(conn, member_area_info)
        summary["states_synced"] = len(synced)
        summary["members_touched"] = len(touched_members)

        for update in test_updates:
            member = ParsedMember(str(update["qq_number"]), str(update["qq_nickname"]), f"{update['qq_nickname']}\n{update['qq_number']}")
            member_id = get_or_create_member(conn, member)
            touched_members.add(member_id)
            if consultation_exists(conn, member_id, update):
                summary["consultations_duplicate"] = int(summary["consultations_duplicate"]) + 1
                continue
            before = dict(conn.execute("SELECT * FROM member_group_states WHERE member_id=? AND group_area=?", (member_id, update["area"])).fetchone() or {})
            ensure_state(conn, member_id, str(update["area"]))
            conn.execute(
                """
                INSERT INTO consultation_records(
                    member_id, group_area, consultation_type, consultation_time, consultant_admin_id,
                    result, status_after, remark, created_at, updated_at
                )
                VALUES(?, ?, '质询', ?, NULL, ?, ?, ?, ?, ?)
                """,
                (
                    member_id,
                    update["area"],
                    update["consultation_time"],
                    update["result"],
                    update["status_after"],
                    update["remark"],
                    ts,
                    ts,
                ),
            )
            conn.execute(
                """
                UPDATE member_group_states
                SET status=?, locked=?, updated_at=?
                WHERE member_id=? AND group_area=?
                """,
                (update["status_after"], update["locked"], ts, member_id, update["area"]),
            )
            after = dict(conn.execute("SELECT * FROM member_group_states WHERE member_id=? AND group_area=?", (member_id, update["area"])).fetchone() or {})
            conn.execute(
                """
                INSERT INTO operation_logs(group_area, operation_type, source, operator_qq, operator_nickname,
                    target_member_id, before_json, after_json, message_id, created_at, remark)
                VALUES(?, '导入质询状态', '后台', NULL, 'Excel导入', ?, ?, ?, NULL, ?, ?)
                """,
                (update["area"], member_id, dump_json(before), dump_json(after), ts, f"来源：测试表第{update['row']}行"),
            )
            summary["consultations_inserted"] = int(summary["consultations_inserted"]) + 1

        conn.execute(
            """
            INSERT INTO operation_logs(group_area, operation_type, source, operator_qq, operator_nickname,
                target_member_id, before_json, after_json, message_id, created_at, remark)
            VALUES(NULL, '导入历史违规表', '后台', NULL, 'Excel导入', NULL, NULL, ?, NULL, ?, ?)
            """,
            (dump_json(summary), now_str(), f"source={source_file}; batch_id={batch_id}"),
        )

    report_path = report_dir / f"import_report_{batch_id}.json"
    csv_path = report_dir / f"imported_violation_records_{batch_id}.csv"
    report_path.write_text(json.dumps(summary, ensure_ascii=False, indent=2), encoding="utf-8")
    with csv_path.open("w", newline="", encoding="utf-8-sig") as f:
        if imported_rows_for_csv:
            writer = csv.DictWriter(f, fieldnames=list(imported_rows_for_csv[0].keys()))
            writer.writeheader()
            writer.writerows(imported_rows_for_csv)
    summary["report_path"] = str(report_path)
    summary["imported_csv_path"] = str(csv_path)
    return summary


def main() -> int:
    parser = argparse.ArgumentParser(description="Import legacy violation xlsx into the bot SQLite database.")
    parser.add_argument("xlsx", nargs="?", default="/opt/import/蜂巢违规记录.xlsx")
    parser.add_argument("--dry-run", action="store_true")
    args = parser.parse_args()
    source = Path(args.xlsx).expanduser().resolve()
    if not source.exists():
        print(f"file not found: {source}", file=sys.stderr)
        return 2
    print(f"database: {CONFIG.database_path}")
    result = import_records(source, dry_run=args.dry_run)
    print(json.dumps(result, ensure_ascii=False, indent=2))
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
