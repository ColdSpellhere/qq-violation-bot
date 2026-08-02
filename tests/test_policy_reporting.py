from __future__ import annotations

import tempfile
import unittest
from dataclasses import replace
from datetime import datetime, timedelta
from pathlib import Path
from unittest.mock import patch

from openpyxl import load_workbook

from plugins.violation_record import db, exporter
from plugins.violation_record.config import CONFIG
from plugins.violation_record.policy_schema import ensure_v102_schema


class PolicyReportingTests(unittest.TestCase):
    def setUp(self) -> None:
        self.temp = tempfile.TemporaryDirectory()
        root = Path(self.temp.name)
        database_path = root / "business.db"
        config = replace(
            CONFIG,
            database_path=database_path,
            database_url=f"sqlite:///{database_path}",
        )
        self.patches = (
            patch.object(db, "CONFIG", config),
            patch.object(exporter, "EXPORT_DIR", root / "exports"),
        )
        for item in self.patches:
            item.start()
        db.init_db()
        now = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
        with db.connect() as conn:
            ensure_v102_schema(conn)
            operation = conn.execute(
                """
                INSERT INTO members(
                    qq_number, qq_nickname, aliases, created_at, updated_at
                ) VALUES('123456', '小明', '[]', ?, ?)
                """,
                (now, now),
            )
            member_id = int(conn.execute("SELECT id FROM members").fetchone()["id"])
            conn.execute(
                """
                INSERT INTO member_group_states(
                    member_id, group_area, status, total_count, deduct_count,
                    current_count_cache, created_at, updated_at
                ) VALUES(?, '蜂巢', '正常', 1, 0, 1, ?, ?)
                """,
                (member_id, now, now),
            )
            for index, event_type in enumerate(
                ("cycle_started", "slow_entered", "cycle_settled"), 1
            ):
                conn.execute(
                    """
                    INSERT INTO v102_policy_events(
                        member_id, group_area, event_type, effective_time,
                        event_priority, source_sequence, ingest_time,
                        payload_json, rule_version, idempotency_key, created_at
                    ) VALUES(?, '蜂巢', ?, ?, 40, ?, ?, ?,
                             'v1.0.2beta', ?, ?)
                    """,
                    (
                        member_id,
                        event_type,
                        now,
                        index,
                        now,
                        f'{{"sequence":{index}}}',
                        f"report-event-{index}",
                        now,
                    ),
                )
            conn.execute(
                """
                INSERT INTO operation_logs(
                    group_area, operation_type, source, operator_qq,
                    operator_nickname, target_member_id, before_json,
                    after_json, created_at, remark
                ) VALUES('蜂巢', '测试操作', '手动', '90001', '管理员',
                         ?, '{}', '{}', ?, '测试')
                """,
                (member_id, now),
            )
            event_id = int(
                conn.execute(
                    "SELECT id FROM v102_policy_events ORDER BY id LIMIT 1"
                ).fetchone()["id"]
            )
            old = (datetime.now() - timedelta(days=8)).strftime(
                "%Y-%m-%d %H:%M:%S"
            )
            pending = conn.execute(
                """
                INSERT INTO v102_pending_actions(
                    member_id, group_area, action_type, status, reason,
                    caused_by_event_id, idempotency_key,
                    created_at, updated_at
                ) VALUES(?, '蜂巢', 'stop_suggestion', 'resolved', '待管理确认', ?,
                         'report-pending', ?, ?)
                """,
                (member_id, event_id, old, now),
            )
            outbox = conn.execute(
                """
                INSERT INTO v102_notification_outbox(
                    event_id, pending_action_id, member_id, group_area,
                    message_type, reminder_slot, message_text,
                    scheduled_at, status, attempt_count,
                    created_at, updated_at
                ) VALUES(?, ?, ?, '蜂巢', 'pending_reminder', 'report-slot',
                         '测试通知', ?, 'sent', 1, ?, ?)
                """,
                (event_id, pending.lastrowid, member_id, now, now, now),
            )
            conn.execute(
                """
                INSERT INTO v102_notification_attempts(
                    outbox_id, attempt_number, status, started_at,
                    finished_at, detail, created_at, updated_at
                ) VALUES(?, 1, 'sent', ?, ?, NULL, ?, ?)
                """,
                (outbox.lastrowid, now, now, now, now),
            )
            conn.execute(
                """
                INSERT INTO v102_status_bridge_jobs(
                    operation_log_id, member_id, group_area, target_status,
                    effective_at, idempotency_key, job_status, attempt_count,
                    applied_event_id, created_at, updated_at
                ) VALUES(?, ?, '蜂巢', '已质询', ?, 'report-status-job',
                         'applied', 1, ?, ?, ?)
                """,
                (operation.lastrowid, member_id, now, event_id, now, now),
            )

    def tearDown(self) -> None:
        for item in reversed(self.patches):
            item.stop()
        self.temp.cleanup()

    def test_weekly_report_preserves_existing_sheets_and_adds_complete_policy_log(self) -> None:
        path = exporter.weekly_report("xlsx")
        workbook = load_workbook(path, read_only=True, data_only=True)

        self.assertEqual(workbook.sheetnames[:2], ["本周更新", "当前统计"])
        self.assertIn("减数策略日志", workbook.sheetnames)
        sheet = workbook["减数策略日志"]
        rows = list(sheet.iter_rows(values_only=True))
        headers = rows[0]
        self.assertIn("事件编号", headers)
        self.assertIn("规则版本", headers)
        self.assertIn("通知状态", headers)
        self.assertEqual(len(rows) - 1, 3)
        event_types = {row[headers.index("事件类型")] for row in rows[1:]}
        self.assertEqual(
            event_types,
            {"cycle_started", "slow_entered", "cycle_settled"},
        )
        self.assertIn("减数待办", workbook.sheetnames)
        pending_rows = list(
            workbook["减数待办"].iter_rows(values_only=True)
        )
        self.assertEqual(len(pending_rows) - 1, 1)

        self.assertIn("通知发送历史", workbook.sheetnames)
        attempt_rows = list(
            workbook["通知发送历史"].iter_rows(values_only=True)
        )
        self.assertEqual(len(attempt_rows) - 1, 1)
        self.assertIn("发送结果", attempt_rows[0])
        self.assertIn("尝试序号", attempt_rows[0])

        self.assertIn("状态联动作业", workbook.sheetnames)
        status_job_rows = list(
            workbook["状态联动作业"].iter_rows(values_only=True)
        )
        self.assertEqual(len(status_job_rows) - 1, 1)
        self.assertIn("作业状态", status_job_rows[0])
        self.assertIn("操作人QQ", status_job_rows[0])
        workbook.close()


if __name__ == "__main__":
    unittest.main()
