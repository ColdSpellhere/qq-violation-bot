from __future__ import annotations

import asyncio
import sqlite3
import tempfile
import unittest
from contextlib import closing
from datetime import datetime, timedelta
from pathlib import Path
from types import SimpleNamespace
from zoneinfo import ZoneInfo

from openpyxl import load_workbook

from plugins.feature_control.state import FeatureController, FeatureState


MONITOR_GROUP_ID = 900_000_000_000_100_001
REPORT_GROUP_ID = 900_000_000_000_100_002
OTHER_GROUP_ID = 900_000_000_000_100_003
NORMAL_CHAT_GROUP_ID = 900_000_000_000_100_004
USER_A = 900_000_000_000_200_001
USER_B = 900_000_000_000_200_002
USER_C = 900_000_000_000_200_003
FIXED_TIME = datetime(2026, 8, 30, 14, 5, 6)
DISPLAY_TIMEZONE = ZoneInfo("Asia/Shanghai")


def _member(
    user_id: int,
    *,
    nickname: str = "",
    card: str = "",
) -> dict[str, object]:
    return {
        "group_id": MONITOR_GROUP_ID,
        "user_id": user_id,
        "nickname": nickname,
        "card": card,
        "role": "member",
    }


class FakeBot:
    def __init__(
        self,
        members: object,
        *,
        upload_failures: int = 0,
        send_failures: int = 0,
        reported_member_count: int | None = None,
        send_delay: float = 0.0,
    ) -> None:
        self.members = members
        self.upload_failures = upload_failures
        self.send_failures = send_failures
        self.reported_member_count = reported_member_count
        self.send_delay = send_delay
        self.api_calls: list[tuple[str, dict[str, object]]] = []
        self.group_messages: list[dict[str, object]] = []

    async def call_api(self, api: str, **kwargs: object) -> object:
        self.api_calls.append((api, kwargs))
        if api == "get_group_member_list":
            return self.members
        if api == "get_group_info":
            count = self.reported_member_count
            if count is None:
                count = len(self.members) if isinstance(self.members, list) else 0
            return {
                "group_id": MONITOR_GROUP_ID,
                "member_count": count,
            }
        if api == "get_group_member_info":
            target = str(kwargs.get("user_id", ""))
            if isinstance(self.members, list):
                for member in self.members:
                    if str(member.get("user_id", "")) == target:
                        return member
            raise RuntimeError("synthetic member not found")
        if api == "upload_group_file":
            if self.upload_failures:
                self.upload_failures -= 1
                raise RuntimeError("synthetic upload failure")
            return {"status": "ok"}
        raise AssertionError(f"unexpected OneBot API: {api}")

    async def send_group_msg(self, **kwargs: object) -> dict[str, int]:
        if self.send_delay:
            await asyncio.sleep(self.send_delay)
        if self.send_failures:
            self.send_failures -= 1
            raise RuntimeError("synthetic send failure")
        self.group_messages.append(dict(kwargs))
        return {"message_id": len(self.group_messages)}


def _feature_defaults() -> FeatureState:
    return FeatureState(
        business_enabled=True,
        chat_enabled=True,
        group_chat_enabled=True,
        private_chat_enabled=False,
        group_chat_allowed_group_ids=(MONITOR_GROUP_ID, NORMAL_CHAT_GROUP_ID),
        private_chat_allowed_user_ids=(),
    )


def _service(
    store: object,
    output_dir: Path,
    *,
    report_group_id: int = REPORT_GROUP_ID,
    runtime_enabled=None,
):
    from plugins.hive_member_monitor.service import HiveMemberMonitorService

    config = SimpleNamespace(
        hive_member_monitor_enabled=True,
        hive_member_monitor_group_id=MONITOR_GROUP_ID,
        hive_member_report_group_id=report_group_id,
    )
    return HiveMemberMonitorService(
        config=config,
        store=store,
        output_dir=output_dir,
        clock=lambda: FIXED_TIME,
        runtime_enabled=runtime_enabled,
    )


class MonitorOnlyIsolationTests(unittest.TestCase):
    def test_monitor_only_exclusion_wins_over_runtime_chat_allowlist(self) -> None:
        with tempfile.TemporaryDirectory() as directory:
            controller = FeatureController(
                Path(directory) / "runtime_features.json",
                _feature_defaults(),
                excluded_group_chat_ids=(MONITOR_GROUP_ID,),
            )

            self.assertFalse(controller.group_chat_allowed(MONITOR_GROUP_ID))
            self.assertTrue(controller.group_chat_allowed(NORMAL_CHAT_GROUP_ID))


class MemberNormalizationAndExportTests(unittest.TestCase):
    def test_member_normalization_uses_card_then_nickname_then_qq(self) -> None:
        from plugins.hive_member_monitor.store import normalize_members

        normalized = normalize_members(
            [
                _member(USER_B, nickname="昵称乙", card="  群名片乙  "),
                _member(USER_A, nickname="  昵称甲  "),
                _member(USER_C),
                _member(USER_B, nickname="昵称乙", card="群名片乙"),
                {"user_id": "not-a-number", "nickname": "无效成员"},
            ]
        )

        self.assertEqual(
            [str(USER_A), str(USER_B), str(USER_C)],
            [item.user_id for item in normalized],
        )
        self.assertEqual(
            ["昵称甲", "群名片乙", str(USER_C)],
            [item.qq_name for item in normalized],
        )

    def test_legacy_member_schema_backfills_episode_watermark_idempotently(self) -> None:
        from plugins.hive_member_monitor.store import MemberSnapshotStore

        with tempfile.TemporaryDirectory() as directory:
            database_path = Path(directory) / "legacy-members.db"
            with closing(sqlite3.connect(database_path)) as connection:
                connection.executescript(
                    """
                    CREATE TABLE hive_monitor_members (
                        group_id INTEGER NOT NULL,
                        user_id TEXT NOT NULL,
                        nickname TEXT NOT NULL,
                        card TEXT NOT NULL,
                        qq_name TEXT NOT NULL,
                        role TEXT NOT NULL,
                        active INTEGER NOT NULL,
                        missing_count INTEGER NOT NULL,
                        first_seen_at TEXT NOT NULL,
                        last_seen_at TEXT NOT NULL,
                        left_at TEXT,
                        PRIMARY KEY(group_id, user_id)
                    );
                    """
                )
                connection.execute(
                    "INSERT INTO hive_monitor_members VALUES(?,?,?,?,?,?,1,0,?,?,NULL)",
                    (
                        MONITOR_GROUP_ID,
                        str(USER_A),
                        "甲",
                        "",
                        "甲",
                        "member",
                        "2026-08-30 14:05:00",
                        "2026-08-30 14:07:00",
                    ),
                )
                connection.commit()

            first = MemberSnapshotStore(database_path)
            second = MemberSnapshotStore(database_path)

            self.assertEqual(
                "2026-08-30 14:05:00",
                first.member_episode_started_at(MONITOR_GROUP_ID, USER_A),
            )
            self.assertEqual(
                "2026-08-30 14:05:00",
                second.member_episode_started_at(MONITOR_GROUP_ID, USER_A),
            )

    def test_full_sync_preserves_episode_watermark_but_rejoin_moves_it(self) -> None:
        from plugins.hive_member_monitor.store import MemberSnapshotStore

        with tempfile.TemporaryDirectory() as directory:
            store = MemberSnapshotStore(Path(directory) / "members.db")
            store.replace_snapshot(
                MONITOR_GROUP_ID,
                [_member(USER_A, nickname="甲")],
                now="2026-08-30 14:05:00",
            )
            store.reconcile_snapshot(
                MONITOR_GROUP_ID,
                [_member(USER_A, nickname="甲")],
                now="2026-08-30 14:07:00",
            )
            self.assertEqual(
                "2026-08-30 14:05:00",
                store.member_episode_started_at(MONITOR_GROUP_ID, USER_A),
            )

            store.mark_member_left(
                MONITOR_GROUP_ID,
                USER_A,
                now="2026-08-30 14:08:00",
            )
            store.upsert_member(
                MONITOR_GROUP_ID,
                _member(USER_A, nickname="重新入群的甲"),
                now="2026-08-30 15:00:00",
            )
            self.assertEqual(
                "2026-08-30 15:00:00",
                store.member_episode_started_at(MONITOR_GROUP_ID, USER_A),
            )

    def test_excel_has_exact_columns_stable_order_and_text_qq_ids(self) -> None:
        from plugins.hive_member_monitor.exporter import export_member_list
        from plugins.hive_member_monitor.store import normalize_members

        with tempfile.TemporaryDirectory() as directory:
            members = normalize_members(
                [
                    _member(USER_B, nickname="乙"),
                    _member(USER_A, nickname="甲"),
                ]
            )
            path = export_member_list(
                members,
                output_dir=Path(directory),
                now=FIXED_TIME,
            )

            self.assertEqual(
                "蜂巢群员名单_2026-08-30_14-05-06.xlsx",
                path.name,
            )
            workbook = load_workbook(path, read_only=False, data_only=True)
            self.addCleanup(workbook.close)
            sheet = workbook.active
            rows = list(sheet.iter_rows(values_only=True))

            self.assertEqual(("QQ号", "QQ名字"), rows[0])
            self.assertEqual(
                [(str(USER_A), "甲"), (str(USER_B), "乙")],
                rows[1:],
            )
            for cell in (sheet["A2"], sheet["A3"]):
                self.assertIsInstance(cell.value, str)
                self.assertEqual("s", cell.data_type)
                self.assertEqual("@", cell.number_format)

    def test_excel_treats_untrusted_names_as_safe_text(self) -> None:
        from plugins.hive_member_monitor.exporter import export_member_list
        from plugins.hive_member_monitor.store import normalize_members

        with tempfile.TemporaryDirectory() as directory:
            members = normalize_members(
                [
                    _member(
                        USER_A,
                        card=" \t=HYPERLINK(\"https://invalid.example\",\"点击\")",
                    ),
                    _member(USER_B, nickname="带控制字符\x07的昵称"),
                ]
            )
            path = export_member_list(
                members,
                output_dir=Path(directory),
                now=FIXED_TIME,
            )

            workbook = load_workbook(path, read_only=False, data_only=False)
            self.addCleanup(workbook.close)
            sheet = workbook.active

            self.assertEqual("s", sheet["B2"].data_type)
            self.assertTrue(str(sheet["B2"].value).startswith("'"))
            self.assertNotIn("\x07", str(sheet["B3"].value))


class FirstExportDeliveryTests(unittest.IsolatedAsyncioTestCase):
    async def test_delivery_state_is_bound_to_report_group(self) -> None:
        from plugins.hive_member_monitor.store import MemberSnapshotStore

        with tempfile.TemporaryDirectory() as directory:
            root = Path(directory)
            store = MemberSnapshotStore(root / "members.db")
            members = [_member(USER_A, nickname="甲")]

            await _service(store, root / "exports").sync_once(FakeBot(members))
            changed_target_bot = FakeBot(members)
            await _service(
                store,
                root / "exports",
                report_group_id=OTHER_GROUP_ID,
            ).sync_once(changed_target_bot)

            uploads = [
                call
                for call in changed_target_bot.api_calls
                if call[0] == "upload_group_file"
            ]
            self.assertEqual(1, len(uploads))
            self.assertEqual(str(OTHER_GROUP_ID), uploads[0][1]["group_id"])

    async def test_first_successful_fetch_uploads_once_across_restart(self) -> None:
        from plugins.hive_member_monitor.store import MemberSnapshotStore

        with tempfile.TemporaryDirectory() as directory:
            root = Path(directory)
            database_path = root / "members.db"
            members = [
                _member(USER_A, nickname="甲"),
                _member(USER_B, nickname="乙"),
            ]
            first_bot = FakeBot(members)
            first_store = MemberSnapshotStore(database_path)

            await _service(first_store, root / "exports").sync_once(first_bot)

            uploads = [call for call in first_bot.api_calls if call[0] == "upload_group_file"]
            self.assertEqual(1, len(uploads))
            self.assertEqual(str(REPORT_GROUP_ID), uploads[0][1]["group_id"])
            self.assertEqual(
                "蜂巢群员名单_2026-08-30_14-05-06.xlsx",
                uploads[0][1]["name"],
            )
            self.assertTrue(first_store.initial_export_delivered(MONITOR_GROUP_ID))
            with closing(sqlite3.connect(database_path)) as connection:
                metadata = connection.execute(
                    "SELECT initial_export_report_group_id,"
                    "initial_export_file_name,initial_export_sha256 "
                    "FROM hive_monitor_group_state WHERE group_id=?",
                    (MONITOR_GROUP_ID,),
                ).fetchone()
            self.assertEqual(REPORT_GROUP_ID, metadata[0])
            self.assertEqual(
                "蜂巢群员名单_2026-08-30_14-05-06.xlsx",
                metadata[1],
            )
            self.assertRegex(str(metadata[2]), r"^[0-9a-f]{64}$")

            restarted_bot = FakeBot(members)
            restarted_store = MemberSnapshotStore(database_path)
            await _service(restarted_store, root / "exports").sync_once(restarted_bot)

            self.assertEqual(
                [],
                [call for call in restarted_bot.api_calls if call[0] == "upload_group_file"],
            )
            self.assertTrue(restarted_store.initial_export_delivered(MONITOR_GROUP_ID))

    async def test_failed_upload_remains_pending_and_retries(self) -> None:
        from plugins.hive_member_monitor.store import MemberSnapshotStore

        with tempfile.TemporaryDirectory() as directory:
            root = Path(directory)
            store = MemberSnapshotStore(root / "members.db")
            bot = FakeBot([_member(USER_A, nickname="甲")], upload_failures=1)
            service = _service(store, root / "exports")

            with self.assertRaisesRegex(RuntimeError, "synthetic upload failure"):
                await service.sync_once(bot)
            self.assertFalse(store.initial_export_delivered(MONITOR_GROUP_ID))

            await service.sync_once(bot)

            uploads = [call for call in bot.api_calls if call[0] == "upload_group_file"]
            self.assertEqual(2, len(uploads))
            self.assertTrue(store.initial_export_delivered(MONITOR_GROUP_ID))


class DepartureMonitoringTests(unittest.IsolatedAsyncioTestCase):
    async def test_decrease_rolls_back_member_state_when_outbox_insert_fails(self) -> None:
        from plugins.hive_member_monitor.store import MemberSnapshotStore

        with tempfile.TemporaryDirectory() as directory:
            root = Path(directory)
            database_path = root / "members.db"
            store = MemberSnapshotStore(database_path)
            store.replace_snapshot(
                MONITOR_GROUP_ID,
                [_member(USER_A, nickname="甲")],
                now=FIXED_TIME,
            )
            with closing(sqlite3.connect(database_path)) as connection:
                connection.execute(
                    """
                    CREATE TRIGGER fail_single_departure_outbox_insert
                    BEFORE INSERT ON hive_monitor_departure_outbox
                    BEGIN
                        SELECT RAISE(ABORT, 'forced departure outbox failure');
                    END
                    """
                )
                connection.commit()

            with self.assertRaisesRegex(
                sqlite3.DatabaseError,
                "forced departure outbox failure",
            ):
                await _service(
                    store,
                    root / "exports",
                ).handle_group_decrease(
                    FakeBot([]),
                    group_id=MONITOR_GROUP_ID,
                    user_id=USER_A,
                    sub_type="leave",
                    event_time=1_788_084_306,
                )

            self.assertTrue(store.member_active(MONITOR_GROUP_ID, USER_A))
            self.assertEqual(
                [],
                store.list_pending_departures(group_id=MONITOR_GROUP_ID),
            )

    async def test_reconcile_rolls_back_all_departures_when_one_outbox_insert_fails(
        self,
    ) -> None:
        from plugins.hive_member_monitor.store import MemberSnapshotStore

        with tempfile.TemporaryDirectory() as directory:
            root = Path(directory)
            database_path = root / "members.db"
            store = MemberSnapshotStore(database_path)
            base_user_id = 830_000_000_000_000_000
            baseline = [
                _member(base_user_id + index, nickname=f"成员{index}")
                for index in range(100)
            ]
            current = baseline[:-2]
            store.replace_snapshot(MONITOR_GROUP_ID, baseline, now=FIXED_TIME)
            store.mark_initial_export_delivered(
                MONITOR_GROUP_ID,
                REPORT_GROUP_ID,
                now=FIXED_TIME,
            )
            bot = FakeBot(current, reported_member_count=len(current))
            service = _service(store, root / "exports")

            await service.sync_once(bot)
            failed_user_id = str(baseline[-1]["user_id"])
            with closing(sqlite3.connect(database_path)) as connection:
                connection.execute(
                    f"""
                    CREATE TRIGGER fail_one_reconcile_outbox_insert
                    BEFORE INSERT ON hive_monitor_departure_outbox
                    WHEN NEW.user_id = '{failed_user_id}'
                    BEGIN
                        SELECT RAISE(ABORT, 'forced reconcile outbox failure');
                    END
                    """
                )
                connection.commit()

            with self.assertRaisesRegex(
                sqlite3.DatabaseError,
                "forced reconcile outbox failure",
            ):
                await service.sync_once(bot)

            self.assertEqual(100, store.member_count(MONITOR_GROUP_ID))
            self.assertEqual(
                [],
                store.list_pending_departures(group_id=MONITOR_GROUP_ID),
            )

    async def test_group_increase_reactivates_member_without_report(self) -> None:
        from plugins.hive_member_monitor.store import MemberSnapshotStore

        with tempfile.TemporaryDirectory() as directory:
            root = Path(directory)
            store = MemberSnapshotStore(root / "members.db")
            store.replace_snapshot(
                MONITOR_GROUP_ID,
                [_member(USER_A, nickname="旧昵称")],
                now=FIXED_TIME,
            )
            store.mark_member_left(MONITOR_GROUP_ID, USER_A, now=FIXED_TIME)
            bot = FakeBot([_member(USER_A, card="重新入群")])

            changed = await _service(
                store,
                root / "exports",
            ).handle_group_increase(
                bot,
                group_id=MONITOR_GROUP_ID,
                user_id=USER_A,
            )

            self.assertTrue(changed)
            self.assertEqual(1, store.member_count(MONITOR_GROUP_ID))
            self.assertEqual(
                "重新入群",
                store.get_member(MONITOR_GROUP_ID, USER_A).qq_name,
            )
            self.assertEqual([], bot.group_messages)

    async def test_kick_me_is_not_reported_as_member_departure(self) -> None:
        from plugins.hive_member_monitor.store import MemberSnapshotStore

        with tempfile.TemporaryDirectory() as directory:
            root = Path(directory)
            store = MemberSnapshotStore(root / "members.db")
            store.replace_snapshot(
                MONITOR_GROUP_ID,
                [_member(USER_A, nickname="甲")],
                now=FIXED_TIME,
            )
            bot = FakeBot([])

            changed = await _service(
                store,
                root / "exports",
            ).handle_group_decrease(
                bot,
                group_id=MONITOR_GROUP_ID,
                user_id=USER_A,
                sub_type="kick_me",
                event_time=1_788_084_306,
            )

            self.assertFalse(changed)
            self.assertEqual(1, store.member_count(MONITOR_GROUP_ID))
            self.assertEqual([], bot.group_messages)

    async def test_departure_name_cannot_inject_onebot_cq_segments(self) -> None:
        from plugins.hive_member_monitor.store import MemberSnapshotStore

        with tempfile.TemporaryDirectory() as directory:
            root = Path(directory)
            store = MemberSnapshotStore(root / "members.db")
            store.replace_snapshot(
                MONITOR_GROUP_ID,
                [_member(USER_A, card="恶意名片[CQ:at,qq=all]")],
                now=FIXED_TIME,
            )
            bot = FakeBot([])

            await _service(store, root / "exports").handle_group_decrease(
                bot,
                group_id=MONITOR_GROUP_ID,
                user_id=USER_A,
                sub_type="leave",
                event_time=1_788_084_306,
            )

            message = bot.group_messages[0]["message"]
            self.assertEqual("text", getattr(message, "type", None))
            self.assertIn(
                "恶意名片[CQ:at,qq=all]",
                str(getattr(message, "data", {}).get("text", "")),
            )

    async def test_concurrent_services_claim_each_departure_once(self) -> None:
        from plugins.hive_member_monitor.store import MemberSnapshotStore

        with tempfile.TemporaryDirectory() as directory:
            root = Path(directory)
            database_path = root / "members.db"
            first_store = MemberSnapshotStore(database_path)
            first_store.ensure_departure_event(
                event_key="concurrent-event",
                group_id=MONITOR_GROUP_ID,
                user_id=USER_A,
                qq_name="甲",
                sub_type="leave",
                event_time=1_788_084_306,
                source="OneBot V11 group_decrease",
                now=FIXED_TIME,
            )
            second_store = MemberSnapshotStore(database_path)
            bot = FakeBot([], send_delay=0.02)

            results = await asyncio.gather(
                _service(first_store, root / "exports").deliver_pending_departures(bot),
                _service(second_store, root / "exports").deliver_pending_departures(bot),
            )

            self.assertEqual(1, sum(results))
            self.assertEqual(1, len(bot.group_messages))

    async def test_live_sender_keeps_claim_beyond_original_120_second_window(self) -> None:
        from plugins.hive_member_monitor.service import HiveMemberMonitorService
        from plugins.hive_member_monitor.store import MemberSnapshotStore

        class BlockingBot:
            def __init__(self) -> None:
                self.first_send_started = asyncio.Event()
                self.release_first_send = asyncio.Event()
                self.group_messages: list[dict[str, object]] = []

            async def send_group_msg(self, **kwargs: object) -> dict[str, int]:
                self.group_messages.append(dict(kwargs))
                if len(self.group_messages) == 1:
                    self.first_send_started.set()
                    await self.release_first_send.wait()
                return {"message_id": len(self.group_messages)}

        with tempfile.TemporaryDirectory() as directory:
            root = Path(directory)
            database_path = root / "members.db"
            first_store = MemberSnapshotStore(database_path)
            first_store.ensure_departure_event(
                event_key="long-running-send",
                group_id=MONITOR_GROUP_ID,
                user_id=USER_A,
                qq_name="甲",
                sub_type="leave",
                event_time=1_788_084_306,
                source="OneBot V11 group_decrease",
                now=FIXED_TIME,
            )
            second_store = MemberSnapshotStore(database_path)
            config = SimpleNamespace(
                hive_member_monitor_enabled=True,
                hive_member_monitor_group_id=MONITOR_GROUP_ID,
                hive_member_report_group_id=REPORT_GROUP_ID,
            )
            first_service = HiveMemberMonitorService(
                config=config,
                store=first_store,
                output_dir=root / "exports",
                clock=lambda: FIXED_TIME,
            )
            second_service = HiveMemberMonitorService(
                config=config,
                store=second_store,
                output_dir=root / "exports",
                clock=lambda: FIXED_TIME + timedelta(seconds=121),
            )
            bot = BlockingBot()

            first_task = asyncio.create_task(
                first_service.deliver_pending_departures(bot)
            )
            await asyncio.wait_for(bot.first_send_started.wait(), timeout=1)
            try:
                second_result = await second_service.deliver_pending_departures(bot)
            finally:
                bot.release_first_send.set()
            first_result = await asyncio.wait_for(first_task, timeout=1)

            self.assertEqual(1, len(bot.group_messages))
            self.assertEqual(1, first_result)
            self.assertEqual(0, second_result)

    async def test_delayed_notice_does_not_duplicate_reconciled_departure(self) -> None:
        from plugins.hive_member_monitor.store import MemberSnapshotStore

        with tempfile.TemporaryDirectory() as directory:
            root = Path(directory)
            store = MemberSnapshotStore(root / "members.db")
            store.replace_snapshot(
                MONITOR_GROUP_ID,
                [
                    _member(USER_A, nickname="甲"),
                    _member(USER_B, nickname="乙"),
                    _member(USER_C, nickname="丙"),
                ],
                now=FIXED_TIME,
            )
            store.mark_initial_export_delivered(
                MONITOR_GROUP_ID,
                REPORT_GROUP_ID,
                now=FIXED_TIME,
            )
            bot = FakeBot(
                [
                    _member(USER_A, nickname="甲"),
                    _member(USER_B, nickname="乙"),
                ],
                reported_member_count=2,
            )
            service = _service(store, root / "exports")

            await service.sync_once(bot)
            await service.sync_once(bot)
            self.assertEqual(1, len(bot.group_messages))

            await service.handle_group_decrease(
                bot,
                group_id=MONITOR_GROUP_ID,
                user_id=USER_C,
                sub_type="leave",
                event_time=1_788_084_600,
            )

            self.assertEqual(1, len(bot.group_messages))

    async def test_replayed_old_event_does_not_mark_rejoined_member_inactive(self) -> None:
        from plugins.hive_member_monitor.store import MemberSnapshotStore

        with tempfile.TemporaryDirectory() as directory:
            root = Path(directory)
            store = MemberSnapshotStore(root / "members.db")
            store.replace_snapshot(
                MONITOR_GROUP_ID,
                [_member(USER_A, nickname="甲")],
                now=FIXED_TIME,
            )
            bot = FakeBot([])
            service = _service(store, root / "exports")
            event = {
                "group_id": MONITOR_GROUP_ID,
                "user_id": USER_A,
                "sub_type": "leave",
                "event_time": 1_788_084_306,
            }

            await service.handle_group_decrease(bot, **event)
            store.upsert_member(
                MONITOR_GROUP_ID,
                _member(USER_A, nickname="重新入群的甲"),
                now="2026-08-30 15:00:00",
            )
            self.assertEqual(1, store.member_count(MONITOR_GROUP_ID))

            await service.handle_group_decrease(bot, **event)

            self.assertEqual(1, store.member_count(MONITOR_GROUP_ID))
            self.assertEqual(1, len(bot.group_messages))

    async def test_delayed_cross_source_event_before_rejoin_watermark_is_ignored(self) -> None:
        from plugins.hive_member_monitor.store import MemberSnapshotStore

        with tempfile.TemporaryDirectory() as directory:
            root = Path(directory)
            store = MemberSnapshotStore(root / "members.db")
            store.replace_snapshot(
                MONITOR_GROUP_ID,
                [
                    _member(USER_A, nickname="甲"),
                    _member(USER_B, nickname="乙"),
                    _member(USER_C, nickname="丙"),
                ],
                now=FIXED_TIME,
            )
            store.mark_initial_export_delivered(
                MONITOR_GROUP_ID,
                REPORT_GROUP_ID,
                now=FIXED_TIME,
            )
            bot = FakeBot(
                [
                    _member(USER_A, nickname="甲"),
                    _member(USER_B, nickname="乙"),
                ],
                reported_member_count=2,
            )
            service = _service(store, root / "exports")
            await service.sync_once(bot)
            await service.sync_once(bot)
            self.assertEqual(1, len(bot.group_messages))

            rejoined_at = datetime(2026, 8, 30, 15, 0, 0)
            store.upsert_member(
                MONITOR_GROUP_ID,
                _member(USER_C, nickname="重新入群的丙"),
                now=rejoined_at,
            )
            delayed_old_event_time = int(
                datetime(
                    2026,
                    8,
                    30,
                    14,
                    10,
                    0,
                    tzinfo=DISPLAY_TIMEZONE,
                ).timestamp()
            )

            changed = await service.handle_group_decrease(
                bot,
                group_id=MONITOR_GROUP_ID,
                user_id=USER_C,
                sub_type="leave",
                event_time=delayed_old_event_time,
            )

            self.assertFalse(changed)
            self.assertEqual(3, store.member_count(MONITOR_GROUP_ID))
            self.assertEqual(1, len(bot.group_messages))

    async def test_real_departure_during_full_sync_is_newer_than_episode_watermark(
        self,
    ) -> None:
        from plugins.hive_member_monitor.service import HiveMemberMonitorService
        from plugins.hive_member_monitor.store import MemberSnapshotStore

        with tempfile.TemporaryDirectory() as directory:
            root = Path(directory)
            store = MemberSnapshotStore(root / "members.db")
            baseline_time = datetime(2026, 8, 30, 14, 5, 0)
            sync_commit_time = datetime(2026, 8, 30, 14, 7, 0)
            store.replace_snapshot(
                MONITOR_GROUP_ID,
                [_member(USER_A, nickname="甲")],
                now=baseline_time,
            )
            store.mark_initial_export_delivered(
                MONITOR_GROUP_ID,
                REPORT_GROUP_ID,
                now=baseline_time,
            )
            config = SimpleNamespace(
                hive_member_monitor_enabled=True,
                hive_member_monitor_group_id=MONITOR_GROUP_ID,
                hive_member_report_group_id=REPORT_GROUP_ID,
            )
            service = HiveMemberMonitorService(
                config=config,
                store=store,
                output_dir=root / "exports",
                clock=lambda: sync_commit_time,
            )
            bot = FakeBot([_member(USER_A, nickname="甲")])

            await service.sync_once(bot)
            event_time = int(
                datetime(
                    2026,
                    8,
                    30,
                    14,
                    6,
                    0,
                    tzinfo=DISPLAY_TIMEZONE,
                ).timestamp()
            )
            changed = await service.handle_group_decrease(
                bot,
                group_id=MONITOR_GROUP_ID,
                user_id=USER_A,
                sub_type="leave",
                event_time=event_time,
            )

            self.assertTrue(changed)
            self.assertFalse(store.member_active(MONITOR_GROUP_ID, USER_A))
            self.assertEqual(1, len(bot.group_messages))

    async def test_failed_departure_send_stays_pending_and_keeps_operator(self) -> None:
        from plugins.hive_member_monitor.store import MemberSnapshotStore

        with tempfile.TemporaryDirectory() as directory:
            root = Path(directory)
            store = MemberSnapshotStore(root / "members.db")
            store.replace_snapshot(
                MONITOR_GROUP_ID,
                [_member(USER_A, nickname="甲")],
                now=FIXED_TIME,
            )
            bot = FakeBot([], send_failures=1)
            service = _service(store, root / "exports")

            with self.assertRaisesRegex(RuntimeError, "synthetic send failure"):
                await service.handle_group_decrease(
                    bot,
                    group_id=MONITOR_GROUP_ID,
                    user_id=USER_A,
                    sub_type="kick",
                    event_time=1_788_084_306,
                    operator_id=USER_B,
                )

            pending = store.list_pending_departures(group_id=MONITOR_GROUP_ID)
            self.assertEqual(1, len(pending))
            self.assertEqual(str(USER_B), pending[0].operator_id)
            self.assertEqual(1, await service.deliver_pending_departures(bot))
            self.assertEqual(1, len(bot.group_messages))
            self.assertIn(str(USER_B), str(bot.group_messages[0]["message"]))
            self.assertIn("被管理员移出", str(bot.group_messages[0]["message"]))

    async def test_duplicate_departure_is_reported_once_across_restart(self) -> None:
        from plugins.hive_member_monitor.store import MemberSnapshotStore

        with tempfile.TemporaryDirectory() as directory:
            root = Path(directory)
            database_path = root / "members.db"
            store = MemberSnapshotStore(database_path)
            store.replace_snapshot(
                MONITOR_GROUP_ID,
                [_member(USER_A, nickname="甲", card="甲的群名片")],
                now=FIXED_TIME,
            )
            first_bot = FakeBot([])
            first_service = _service(store, root / "exports")

            await first_service.handle_group_decrease(
                first_bot,
                group_id=MONITOR_GROUP_ID,
                user_id=USER_A,
                sub_type="leave",
                event_time=1_788_084_306,
            )

            restarted_store = MemberSnapshotStore(database_path)
            restarted_bot = FakeBot([])
            restarted_service = _service(restarted_store, root / "exports")
            await restarted_service.handle_group_decrease(
                restarted_bot,
                group_id=MONITOR_GROUP_ID,
                user_id=USER_A,
                sub_type="leave",
                event_time=1_788_084_306,
            )

            self.assertEqual(1, len(first_bot.group_messages))
            self.assertEqual([], restarted_bot.group_messages)
            sent = first_bot.group_messages[0]
            self.assertEqual(REPORT_GROUP_ID, sent["group_id"])
            text = str(sent["message"])
            for expected in (
                "群员退群日志",
                str(MONITOR_GROUP_ID),
                str(USER_A),
                "甲的群名片",
                "主动退群",
                "OneBot",
            ):
                self.assertIn(expected, text)

    async def test_decrease_from_other_group_is_ignored(self) -> None:
        from plugins.hive_member_monitor.store import MemberSnapshotStore

        with tempfile.TemporaryDirectory() as directory:
            root = Path(directory)
            store = MemberSnapshotStore(root / "members.db")
            bot = FakeBot([])

            await _service(store, root / "exports").handle_group_decrease(
                bot,
                group_id=OTHER_GROUP_ID,
                user_id=USER_A,
                sub_type="leave",
                event_time=1_788_084_306,
            )

            self.assertEqual([], bot.group_messages)
            self.assertEqual([], bot.api_calls)


class SnapshotSafetyTests(unittest.IsolatedAsyncioTestCase):
    async def test_repeated_identical_large_difference_eventually_reconciles(self) -> None:
        from plugins.hive_member_monitor.store import MemberSnapshotStore

        with tempfile.TemporaryDirectory() as directory:
            root = Path(directory)
            database_path = root / "members.db"
            store = MemberSnapshotStore(database_path)
            base_user_id = 840_000_000_000_000_000
            baseline = [
                _member(base_user_id + index, nickname=f"成员{index}")
                for index in range(100)
            ]
            current = baseline[:70]
            store.replace_snapshot(MONITOR_GROUP_ID, baseline, now=FIXED_TIME)
            store.mark_initial_export_delivered(
                MONITOR_GROUP_ID,
                REPORT_GROUP_ID,
                now=FIXED_TIME,
            )
            bot = FakeBot(current, reported_member_count=len(current))
            service = _service(store, root / "exports")

            for _ in range(2):
                try:
                    await service.sync_once(bot)
                except ValueError as exc:
                    self.assertIn("departure fuse", str(exc))
                self.assertEqual(100, store.member_count(MONITOR_GROUP_ID))
                self.assertEqual(
                    [],
                    store.list_pending_departures(group_id=MONITOR_GROUP_ID),
                )

            try:
                await service.sync_once(bot)
            except ValueError as exc:
                self.fail(
                    "the third identical complete snapshot must pass conservative "
                    f"confirmation instead of remaining fused: {exc}"
                )

            self.assertEqual(70, store.member_count(MONITOR_GROUP_ID))
            with closing(sqlite3.connect(database_path)) as connection:
                outbox_count = connection.execute(
                    "SELECT COUNT(*) FROM hive_monitor_departure_outbox "
                    "WHERE group_id=?",
                    (MONITOR_GROUP_ID,),
                ).fetchone()[0]
            self.assertEqual(30, outbox_count)

    async def test_runtime_disable_during_fetch_prevents_mutation_and_upload(self) -> None:
        from plugins.hive_member_monitor.store import MemberSnapshotStore

        enabled = True

        class DisablingBot(FakeBot):
            async def call_api(self, api: str, **kwargs: object) -> object:
                nonlocal enabled
                result = await super().call_api(api, **kwargs)
                if api == "get_group_info":
                    enabled = False
                return result

        with tempfile.TemporaryDirectory() as directory:
            root = Path(directory)
            store = MemberSnapshotStore(root / "members.db")
            bot = DisablingBot([_member(USER_A, nickname="甲")])
            service = _service(
                store,
                root / "exports",
                runtime_enabled=lambda: enabled,
            )

            self.assertEqual(0, await service.sync_once(bot))
            self.assertEqual(0, store.member_count(MONITOR_GROUP_ID))
            self.assertEqual(
                [],
                [call for call in bot.api_calls if call[0] == "upload_group_file"],
            )

    async def test_first_snapshot_supports_more_than_sqlite_variable_limit(self) -> None:
        from plugins.hive_member_monitor.store import MemberSnapshotStore

        class LimitedVariableStore(MemberSnapshotStore):
            def _connect(self):
                connection = super()._connect()
                setlimit = getattr(connection, "setlimit", None)
                if setlimit is not None:
                    setlimit(sqlite3.SQLITE_LIMIT_VARIABLE_NUMBER, 999)
                return connection

        with tempfile.TemporaryDirectory() as directory:
            store = LimitedVariableStore(Path(directory) / "members.db")
            members = [
                _member(800_000_000_000_000_000 + index, nickname=f"成员{index}")
                for index in range(2_894)
            ]

            delta = store.replace_snapshot(
                MONITOR_GROUP_ID,
                members,
                now=FIXED_TIME,
            )

            self.assertEqual(2_894, store.member_count(MONITOR_GROUP_ID))
            self.assertEqual(2_894, len(delta.joined))
            self.assertEqual((), delta.departed)

    async def test_initial_export_requires_full_list_count_match(self) -> None:
        from plugins.hive_member_monitor.store import MemberSnapshotStore

        with tempfile.TemporaryDirectory() as directory:
            root = Path(directory)
            store = MemberSnapshotStore(root / "members.db")
            bot = FakeBot(
                [_member(USER_A, nickname="甲")],
                reported_member_count=2,
            )

            with self.assertRaisesRegex(ValueError, "member count"):
                await _service(store, root / "exports").sync_once(bot)

            self.assertEqual(0, store.member_count(MONITOR_GROUP_ID))
            self.assertFalse(store.initial_export_delivered(MONITOR_GROUP_ID))
            self.assertEqual(
                [],
                [call for call in bot.api_calls if call[0] == "upload_group_file"],
            )

    async def test_stable_truncated_list_never_creates_mass_departures(self) -> None:
        from plugins.hive_member_monitor.store import MemberSnapshotStore

        with tempfile.TemporaryDirectory() as directory:
            root = Path(directory)
            store = MemberSnapshotStore(root / "members.db")
            all_members = [
                _member(810_000_000_000_000_000 + index, nickname=f"成员{index}")
                for index in range(2_894)
            ]
            store.replace_snapshot(MONITOR_GROUP_ID, all_members, now=FIXED_TIME)
            store.mark_initial_export_delivered(
                MONITOR_GROUP_ID,
                REPORT_GROUP_ID,
                now=FIXED_TIME,
            )
            bot = FakeBot(
                all_members[:2_000],
                reported_member_count=2_894,
            )
            service = _service(store, root / "exports")

            for _ in range(2):
                with self.assertRaisesRegex(ValueError, "member count"):
                    await service.sync_once(bot)

            self.assertEqual(2_894, store.member_count(MONITOR_GROUP_ID))
            self.assertEqual([], bot.group_messages)

    async def test_mass_difference_fuse_blocks_even_matching_reported_count(self) -> None:
        from plugins.hive_member_monitor.store import MemberSnapshotStore

        with tempfile.TemporaryDirectory() as directory:
            root = Path(directory)
            store = MemberSnapshotStore(root / "members.db")
            all_members = [
                _member(820_000_000_000_000_000 + index, nickname=f"成员{index}")
                for index in range(100)
            ]
            store.replace_snapshot(MONITOR_GROUP_ID, all_members, now=FIXED_TIME)
            store.mark_initial_export_delivered(
                MONITOR_GROUP_ID,
                REPORT_GROUP_ID,
                now=FIXED_TIME,
            )
            bot = FakeBot(all_members[:70], reported_member_count=70)

            with self.assertRaisesRegex(ValueError, "departure fuse"):
                await _service(store, root / "exports").sync_once(bot)

            self.assertEqual(100, store.member_count(MONITOR_GROUP_ID))
            self.assertEqual([], bot.group_messages)

    async def test_reconcile_requires_two_valid_absences_before_reporting(self) -> None:
        from plugins.hive_member_monitor.store import MemberSnapshotStore

        with tempfile.TemporaryDirectory() as directory:
            root = Path(directory)
            store = MemberSnapshotStore(root / "members.db")
            store.replace_snapshot(
                MONITOR_GROUP_ID,
                [
                    _member(USER_A, nickname="甲"),
                    _member(USER_B, nickname="乙"),
                    _member(USER_C, nickname="丙"),
                ],
                now=FIXED_TIME,
            )
            store.mark_initial_export_delivered(
                MONITOR_GROUP_ID,
                REPORT_GROUP_ID,
                now=FIXED_TIME,
            )
            bot = FakeBot(
                [
                    _member(USER_A, nickname="甲"),
                    _member(USER_B, nickname="乙"),
                ]
            )
            service = _service(store, root / "exports")

            await service.sync_once(bot)
            self.assertEqual([], bot.group_messages)
            self.assertIsNotNone(store.get_member(MONITOR_GROUP_ID, USER_C))

            await service.sync_once(bot)
            self.assertEqual(1, len(bot.group_messages))
            self.assertIn("名单差异复核确认退群", str(bot.group_messages[0]["message"]))
            self.assertIn(str(USER_C), str(bot.group_messages[0]["message"]))

    async def test_implausibly_truncated_list_does_not_mutate_snapshot(self) -> None:
        from plugins.hive_member_monitor.store import MemberSnapshotStore

        with tempfile.TemporaryDirectory() as directory:
            root = Path(directory)
            store = MemberSnapshotStore(root / "members.db")
            baseline = [
                _member(USER_A, nickname="甲"),
                _member(USER_B, nickname="乙"),
                _member(USER_C, nickname="丙"),
                _member(USER_C + 1, nickname="丁"),
            ]
            store.replace_snapshot(MONITOR_GROUP_ID, baseline, now=FIXED_TIME)
            store.mark_initial_export_delivered(
                MONITOR_GROUP_ID,
                REPORT_GROUP_ID,
                now=FIXED_TIME,
            )
            service = _service(store, root / "exports")

            with self.assertRaisesRegex(ValueError, "departure fuse"):
                await service.sync_once(FakeBot([baseline[0]]))

            self.assertEqual(
                4,
                len(store.list_members(MONITOR_GROUP_ID)),
            )

    async def test_empty_or_invalid_member_payload_never_replaces_snapshot(self) -> None:
        from plugins.hive_member_monitor.store import MemberSnapshotStore

        with tempfile.TemporaryDirectory() as directory:
            root = Path(directory)
            store = MemberSnapshotStore(root / "members.db")
            service = _service(store, root / "exports")
            baseline = [
                _member(USER_A, nickname="甲"),
                _member(USER_B, nickname="乙"),
            ]
            await service.sync_once(FakeBot(baseline))
            expected = [(str(USER_A), "甲"), (str(USER_B), "乙")]

            for payload in ([], {}, None):
                with self.subTest(payload=payload):
                    bot = FakeBot(payload)
                    with self.assertRaises((TypeError, ValueError)):
                        await service.sync_once(bot)
                    self.assertEqual(
                        expected,
                        [
                            (item.user_id, item.qq_name)
                            for item in store.list_members(MONITOR_GROUP_ID)
                        ],
                    )
                    self.assertEqual([], bot.group_messages)


if __name__ == "__main__":
    unittest.main()
