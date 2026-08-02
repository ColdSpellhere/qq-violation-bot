from __future__ import annotations

import hashlib
import json
import os
import shutil
import sqlite3
import subprocess
import sys
import tempfile
import unittest
from contextlib import contextmanager
from pathlib import Path
from typing import Iterator

from openpyxl import Workbook, load_workbook

from scripts import migrate_v102


CUTOVER = "2026-08-02 17:00:00"
BATCH_ID = "v102-test-batch"


def file_sha256(path: Path) -> str:
    digest = hashlib.sha256()
    with path.open("rb") as stream:
        for chunk in iter(lambda: stream.read(1024 * 1024), b""):
            digest.update(chunk)
    return digest.hexdigest()


def create_workbook(
    path: Path,
    *,
    duplicate_scope: bool = False,
    excluded_conflict: bool = False,
    invalid_count: bool = False,
) -> None:
    workbook = Workbook()
    workbook.remove(workbook.active)
    for name in migrate_v102.REQUIRED_SHEETS:
        sheet = workbook.create_sheet(name)
        sheet.cell(1, 1, "管理员标记色")
        sheet.cell(2, 2, "名称与QQ号")
        sheet.cell(2, 3, "当前次数")

    rows = {
        "蜂巢": [
            ("表中新昵称\n（10001）", 1),
            ("新成员昵称\n（30003）", 0),
            ("累计成员\n（40004）", 3),
            ("终态成员\n（50005）", 1),
        ],
        "蜂窝": [("跨域成员\n（10001）", 1)],
        "蜂箱": [
            ("最后警告成员\n（60006）", 2),
            ("已质询成员\n（70007）", 0),
        ],
    }
    for sheet_name, values in rows.items():
        sheet = workbook[sheet_name]
        for offset, (member, count) in enumerate(values):
            row = 5 + offset * 2
            sheet.cell(row, 2, member)
            sheet.cell(row, 3, count)

    if duplicate_scope:
        workbook["蜂巢"].cell(13, 2, "重复成员\n（10001）")
        workbook["蜂巢"].cell(13, 3, 1)
    if invalid_count:
        workbook["蜂巢"].cell(5, 3, 1.5)

    low_frequency_qq = "10001" if excluded_conflict else "80008"
    workbook["低频小于三"].cell(4, 2, f"低频成员\n（{low_frequency_qq}）")
    workbook["低频小于三"].cell(4, 3, 1)
    workbook["封存记录"].cell(4, 2, "封存成员\n（90009）")
    workbook["封存记录"].cell(4, 3, 1)
    workbook["手动黑名单"].cell(3, 1, 99999)
    workbook["OOPZ"].cell(5, 2, 88888)
    workbook.save(path)
    workbook.close()


def create_database(path: Path) -> None:
    conn = sqlite3.connect(path)
    conn.executescript(
        """
        PRAGMA foreign_keys = ON;
        CREATE TABLE members (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            qq_number TEXT UNIQUE NOT NULL,
            qq_nickname TEXT,
            aliases TEXT DEFAULT '[]',
            created_at TEXT NOT NULL,
            updated_at TEXT NOT NULL
        );
        CREATE TABLE admins (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            qq_number TEXT UNIQUE NOT NULL,
            nickname TEXT NOT NULL,
            aliases TEXT DEFAULT '[]',
            is_active INTEGER NOT NULL DEFAULT 1,
            created_at TEXT NOT NULL,
            updated_at TEXT NOT NULL
        );
        CREATE TABLE member_group_states (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            member_id INTEGER NOT NULL,
            group_area TEXT NOT NULL,
            status TEXT NOT NULL DEFAULT '正常',
            locked INTEGER NOT NULL DEFAULT 0,
            total_count INTEGER NOT NULL DEFAULT 0,
            deduct_count INTEGER NOT NULL DEFAULT 0,
            current_count_cache INTEGER NOT NULL DEFAULT 0,
            last_effective_violation_time TEXT,
            last_deduct_time TEXT,
            last_final_warning_time TEXT,
            created_at TEXT NOT NULL,
            updated_at TEXT NOT NULL,
            UNIQUE(member_id, group_area),
            FOREIGN KEY(member_id) REFERENCES members(id)
        );
        CREATE TABLE violation_records (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            member_id INTEGER NOT NULL,
            group_area TEXT NOT NULL,
            violation_time TEXT NOT NULL,
            judgement TEXT NOT NULL,
            action TEXT NOT NULL,
            handler_admin_id INTEGER,
            recorder_admin_id INTEGER,
            remark TEXT DEFAULT '无',
            is_countable INTEGER NOT NULL DEFAULT 1,
            count_delta INTEGER NOT NULL DEFAULT 1,
            is_withdrawn INTEGER NOT NULL DEFAULT 0,
            withdrawn_reason TEXT,
            is_test INTEGER NOT NULL DEFAULT 0,
            created_at TEXT NOT NULL,
            updated_at TEXT NOT NULL,
            FOREIGN KEY(member_id) REFERENCES members(id)
        );
        CREATE TABLE consultation_records (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            member_id INTEGER NOT NULL,
            group_area TEXT NOT NULL,
            consultation_type TEXT NOT NULL,
            consultation_time TEXT NOT NULL,
            consultant_admin_id INTEGER,
            result TEXT NOT NULL,
            status_after TEXT NOT NULL,
            remark TEXT DEFAULT '无',
            created_at TEXT NOT NULL,
            updated_at TEXT NOT NULL,
            FOREIGN KEY(member_id) REFERENCES members(id)
        );
        CREATE TABLE operation_logs (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            group_area TEXT,
            operation_type TEXT NOT NULL,
            source TEXT NOT NULL,
            operator_qq TEXT,
            operator_nickname TEXT,
            target_member_id INTEGER,
            before_json TEXT,
            after_json TEXT,
            message_id TEXT,
            created_at TEXT NOT NULL,
            remark TEXT
        );

        INSERT INTO members(
            id, qq_number, qq_nickname, aliases, created_at, updated_at
        ) VALUES
            (1, '10001', '旧昵称', '[]', '2026-01-01 00:00:00', '2026-01-01 00:00:00'),
            (4, '40004', '累计成员旧称', '[]', '2026-01-01 00:00:00', '2026-01-01 00:00:00'),
            (5, '50005', '终态成员旧称', '[]', '2026-01-01 00:00:00', '2026-01-01 00:00:00'),
            (6, '60006', '最后警告成员旧称', '[]', '2026-01-01 00:00:00', '2026-01-01 00:00:00'),
            (7, '70007', '已质询成员旧称', '[]', '2026-01-01 00:00:00', '2026-01-01 00:00:00');

        INSERT INTO member_group_states(
            member_id, group_area, status, locked, total_count, deduct_count,
            current_count_cache, last_final_warning_time, created_at, updated_at
        ) VALUES
            (1, '蜂巢', '正常', 0, 8, 2, 6, NULL, '2026-01-01 00:00:00', '2026-01-01 00:00:00'),
            (4, '蜂巢', '正常', 0, 0, 0, 0, NULL, '2026-01-01 00:00:00', '2026-01-01 00:00:00'),
            (5, '蜂巢', '已移出', 1, 0, 0, 0, NULL, '2026-01-01 00:00:00', '2026-01-01 00:00:00'),
            (6, '蜂箱', '最后警告', 0, 0, 0, 0, '2026-07-01 00:00:00', '2026-01-01 00:00:00', '2026-01-01 00:00:00'),
            (7, '蜂箱', '已质询', 0, 0, 0, 0, NULL, '2026-01-01 00:00:00', '2026-01-01 00:00:00');

        INSERT INTO violation_records(
            id, member_id, group_area, violation_time, judgement, action,
            is_countable, count_delta, created_at, updated_at
        ) VALUES(
            11, 1, '蜂巢', '2026-06-01 00:00:00', '历史违规',
            '禁言10分钟', 1, 1, '2026-06-01 00:00:00', '2026-06-01 00:00:00'
        );
        """
    )
    conn.commit()
    conn.close()


class MigrationTestCase(unittest.TestCase):
    def setUp(self) -> None:
        self.temp = tempfile.TemporaryDirectory()
        self.root = Path(self.temp.name)
        self.database = self.root / "business.db"
        self.baseline = self.root / "baseline.xlsx"
        self._snapshot_sequence = 0
        self.last_backup_sha256: str | None = None
        create_database(self.database)
        create_workbook(self.baseline)

    def tearDown(self) -> None:
        self.temp.cleanup()

    @contextmanager
    def connect(self) -> Iterator[sqlite3.Connection]:
        conn = sqlite3.connect(self.database)
        conn.row_factory = sqlite3.Row
        conn.execute("PRAGMA foreign_keys = ON")
        try:
            yield conn
        finally:
            conn.close()

    def create_pre_cutover_snapshot(self) -> tuple[Path, str]:
        self._snapshot_sequence += 1
        snapshot = self.root / f"pre-cutover-{self._snapshot_sequence}.db"
        shutil.copy2(self.database, snapshot)
        digest = file_sha256(snapshot)
        self.last_backup_sha256 = digest
        return snapshot, digest

    def apply_v102(self, *, batch_id: str = BATCH_ID) -> dict[str, object]:
        snapshot, digest = self.create_pre_cutover_snapshot()
        return migrate_v102.apply_migration(
            self.database,
            self.baseline,
            snapshot_database=snapshot,
            backup_sha256=digest,
            cutover_at=CUTOVER,
            batch_id=batch_id,
        )


class BaselineExtractionTests(MigrationTestCase):
    def test_extracts_only_main_scopes_and_allows_same_qq_across_areas(self) -> None:
        baseline = migrate_v102.read_baseline(self.baseline)

        self.assertEqual(len(baseline.rows), 7)
        self.assertEqual(
            [(row.group_area, row.qq_number) for row in baseline.rows].count(
                ("蜂巢", "10001")
            ),
            1,
        )
        self.assertIn(
            ("蜂窝", "10001"),
            [(row.group_area, row.qq_number) for row in baseline.rows],
        )
        self.assertNotIn("80008", {row.qq_number for row in baseline.rows})
        self.assertNotIn("90009", {row.qq_number for row in baseline.rows})
        self.assertNotIn("99999", {row.qq_number for row in baseline.rows})
        self.assertNotIn("88888", {row.qq_number for row in baseline.rows})
        self.assertEqual(baseline.source_sha256, file_sha256(self.baseline))

    def test_duplicate_member_in_same_area_fails_closed(self) -> None:
        create_workbook(self.baseline, duplicate_scope=True)

        with self.assertRaisesRegex(migrate_v102.MigrationError, "重复"):
            migrate_v102.read_baseline(self.baseline)

    def test_member_present_in_low_frequency_and_main_fails_closed(self) -> None:
        create_workbook(self.baseline, excluded_conflict=True)

        with self.assertRaisesRegex(migrate_v102.MigrationError, "排除表"):
            migrate_v102.read_baseline(self.baseline)

    def test_nonstandard_stale_low_frequency_rows_are_ignored(self) -> None:
        create_workbook(self.baseline)
        workbook = load_workbook(self.baseline)
        workbook["低频小于三"].cell(4, 2, "旧残留 10001")
        workbook.save(self.baseline)
        workbook.close()

        baseline = migrate_v102.read_baseline(self.baseline)

        self.assertIn(
            ("蜂巢", "10001"),
            {(row.group_area, row.qq_number) for row in baseline.rows},
        )

    def test_non_integer_current_count_fails_closed(self) -> None:
        create_workbook(self.baseline, invalid_count=True)

        with self.assertRaisesRegex(migrate_v102.MigrationError, "当前次数"):
            migrate_v102.read_baseline(self.baseline)


class DryRunTests(MigrationTestCase):
    def test_dry_run_is_read_only_deterministic_and_reports_exact_adjustments(self) -> None:
        database_before = file_sha256(self.database)

        first = migrate_v102.dry_run(self.database, self.baseline)
        second = migrate_v102.dry_run(self.database, self.baseline)

        self.assertEqual(first, second)
        self.assertEqual(file_sha256(self.database), database_before)
        self.assertEqual(first["summary"]["baseline_rows"], 7)
        self.assertEqual(first["cutover_record_watermark"], 11)
        item = next(
            row
            for row in first["changes"]
            if row["group_area"] == "蜂巢" and row["qq_number"] == "10001"
        )
        self.assertEqual(item["old_total_count"], 8)
        self.assertEqual(item["old_deduct_count"], 2)
        self.assertEqual(item["approved_current_count"], 1)
        self.assertEqual(item["new_total_count"], 3)
        self.assertEqual(item["raw_total_at_cutover"], 1)
        self.assertEqual(item["new_baseline_adjustment"], 2)
        with self.connect() as conn:
            table = conn.execute(
                "SELECT 1 FROM sqlite_master WHERE name='v102_policy_state'"
            ).fetchone()
        self.assertIsNone(table)


class ApplyAndVerifyTests(MigrationTestCase):
    def test_snapshot_repair_is_bound_to_exact_pre_cutover_database(self) -> None:
        old_effective = "2026-05-01 08:00:00"
        old_deduct = "2026-05-15 08:00:00"
        with self.connect() as conn:
            conn.execute(
                """
                UPDATE member_group_states
                SET last_effective_violation_time=?, last_deduct_time=?
                WHERE member_id=1 AND group_area='蜂巢'
                """,
                (old_effective, old_deduct),
            )
            conn.commit()
        snapshot = self.root / "pre-cutover.db"
        shutil.copy2(self.database, snapshot)
        digest = file_sha256(snapshot)
        migrate_v102.apply_migration(
            self.database,
            self.baseline,
            snapshot_database=snapshot,
            backup_sha256=digest,
            cutover_at=CUTOVER,
            batch_id=BATCH_ID,
        )
        with self.connect() as conn:
            conn.execute(
                """
                UPDATE v102_baseline_audit
                SET old_locked=0, old_last_effective_violation_time=NULL,
                    old_last_deduct_time=NULL,
                    old_last_final_warning_time=NULL
                """
            )
            conn.execute(
                """
                UPDATE v102_policy_state
                SET baseline_total_count=0, baseline_current_count=0,
                    baseline_raw_total=0, baseline_record_watermark=0,
                    baseline_locked=0,
                    baseline_last_effective_violation_time=NULL,
                    baseline_last_deduct_time=NULL,
                    baseline_last_final_warning_time=NULL
                """
            )
            conn.execute(
                """
                UPDATE v102_migration_checkpoints
                SET schema_version='v1.0.2beta-1'
                WHERE batch_id=?
                """,
                (BATCH_ID,),
            )
            conn.commit()

        result = migrate_v102.repair_runtime_snapshots(
            self.database,
            snapshot,
            batch_id=BATCH_ID,
            backup_sha256=digest,
        )

        self.assertEqual(result["repaired_scopes"], 7)
        with self.connect() as conn:
            audit = conn.execute(
                """
                SELECT old_last_effective_violation_time, old_last_deduct_time
                FROM v102_baseline_audit
                WHERE member_id=1 AND group_area='蜂巢'
                """
            ).fetchone()
            policy = conn.execute(
                """
                SELECT baseline_total_count, baseline_deduct_count,
                       baseline_current_count, baseline_raw_total,
                       baseline_record_watermark,
                       baseline_last_effective_violation_time,
                       baseline_last_deduct_time
                FROM v102_policy_state
                WHERE member_id=1 AND group_area='蜂巢'
                """
            ).fetchone()
            checkpoint_version = conn.execute(
                """
                SELECT schema_version FROM v102_migration_checkpoints
                WHERE batch_id=?
                """,
                (BATCH_ID,),
            ).fetchone()[0]
        self.assertEqual(tuple(audit), (old_effective, old_deduct))
        self.assertEqual(
            tuple(policy),
            (8, 2, 6, 1, 11, old_effective, old_deduct),
        )
        self.assertEqual(checkpoint_version, "v1.0.2beta-2")

    def test_snapshot_repair_rejects_checkpoint_watermark_mismatch(self) -> None:
        snapshot, digest = self.create_pre_cutover_snapshot()
        migrate_v102.apply_migration(
            self.database,
            self.baseline,
            snapshot_database=snapshot,
            backup_sha256=digest,
            cutover_at=CUTOVER,
            batch_id=BATCH_ID,
        )
        with self.connect() as conn:
            conn.execute(
                """
                UPDATE v102_migration_checkpoints
                SET cutover_record_watermark=999
                WHERE batch_id=?
                """,
                (BATCH_ID,),
            )
            conn.commit()
        database_before = file_sha256(self.database)

        with self.assertRaisesRegex(migrate_v102.MigrationError, "水位"):
            migrate_v102.repair_runtime_snapshots(
                self.database,
                snapshot,
                batch_id=BATCH_ID,
                backup_sha256=digest,
            )

        self.assertEqual(file_sha256(self.database), database_before)

    def test_snapshot_repair_rejects_unattested_legacy_status_snapshot(self) -> None:
        snapshot, digest = self.create_pre_cutover_snapshot()
        migrate_v102.apply_migration(
            self.database,
            self.baseline,
            snapshot_database=snapshot,
            backup_sha256=digest,
            cutover_at=CUTOVER,
            batch_id=BATCH_ID,
        )
        forged = self.root / "forged-pre-cutover.db"
        shutil.copy2(snapshot, forged)
        with sqlite3.connect(forged) as conn:
            conn.execute(
                """
                UPDATE member_group_states
                SET status='已质询', last_deduct_time='2099-01-01 00:00:00'
                WHERE member_id=1 AND group_area='蜂巢'
                """
            )
        forged_digest = file_sha256(forged)
        with self.connect() as conn:
            conn.execute(
                """
                UPDATE v102_migration_checkpoints
                SET schema_version='v1.0.2beta-1', backup_sha256=?
                WHERE batch_id=?
                """,
                (forged_digest, BATCH_ID),
            )
            conn.commit()
        database_before = file_sha256(self.database)

        with self.assertRaisesRegex(migrate_v102.MigrationError, "快照证明"):
            migrate_v102.repair_runtime_snapshots(
                self.database,
                forged,
                batch_id=BATCH_ID,
                backup_sha256=forged_digest,
            )

        self.assertEqual(file_sha256(self.database), database_before)

    def test_verify_rejects_unrepaired_checkpoint_version(self) -> None:
        self.apply_v102()
        with self.connect() as conn:
            conn.execute(
                """
                UPDATE v102_migration_checkpoints
                SET schema_version='v1.0.2beta-1'
                WHERE batch_id=?
                """,
                (BATCH_ID,),
            )
            conn.commit()

        result = migrate_v102.verify_database(self.database)

        self.assertFalse(result["ok"])
        self.assertIn("checkpoint_schema_versions", result)
        self.assertTrue(
            any("checkpoint_schema_versions" in error for error in result["errors"])
        )

    def test_runtime_readiness_rejects_corrupt_baseline_snapshot(self) -> None:
        self.apply_v102()
        with self.connect() as conn:
            conn.execute(
                """
                UPDATE v102_policy_state
                SET baseline_total_count=baseline_total_count+1
                WHERE member_id=1 AND group_area='蜂巢'
                """
            )
            conn.commit()

        with self.connect() as conn:
            with self.assertRaisesRegex(RuntimeError, "baseline snapshots"):
                migrate_v102._policy_schema.require_v102_ready(conn)

    def test_final_warning_history_creates_remove_pending_at_cutover(self) -> None:
        with self.connect() as conn:
            member_id = conn.execute(
                "SELECT id FROM members WHERE qq_number='60006'"
            ).fetchone()[0]
            conn.execute(
                """
                INSERT INTO violation_records(
                    member_id, group_area, violation_time, judgement, action,
                    is_countable, count_delta, is_test, created_at, updated_at
                ) VALUES(?, '蜂箱', '2026-07-15 10:00:00', '再犯',
                         '禁言10分钟', 1, 1, 0,
                         '2026-07-15 10:00:00', '2026-07-15 10:00:00')
                """,
                (member_id,),
            )
            conn.commit()

        self.apply_v102()

        with self.connect() as conn:
            cycle = conn.execute(
                """
                SELECT c.* FROM v102_policy_cycles c
                WHERE c.member_id=? AND c.group_area='蜂箱'
                """,
                (member_id,),
            ).fetchone()
            pending = conn.execute(
                """
                SELECT p.*, e.event_type FROM v102_pending_actions p
                JOIN v102_policy_events e ON e.id=p.caused_by_event_id
                WHERE p.member_id=? AND p.group_area='蜂箱'
                  AND p.action_type='remove_member'
                """,
                (member_id,),
            ).fetchone()
        self.assertEqual(cycle["cycle_type"], "final_warning")
        self.assertEqual(cycle["status"], "pending_decision")
        self.assertEqual(pending["status"], "pending")
        self.assertEqual(
            pending["event_type"], "historical_final_warning_violation_detected"
        )

    def test_withdraw_rebuilds_migrated_nonzero_baseline_cycle(self) -> None:
        with self.connect() as conn:
            record_id = int(
                conn.execute(
                    """
                    INSERT INTO violation_records(
                        member_id, group_area, violation_time, judgement, action,
                        is_countable, count_delta, is_test, created_at, updated_at
                    ) VALUES(4, '蜂巢', '2026-07-15 10:00:00', '历史违规',
                             '禁言10分钟', 1, 1, 0,
                             '2026-07-15 10:00:00', '2026-07-15 10:00:00')
                    """
                ).lastrowid
            )
            conn.commit()
        self.apply_v102()

        with self.connect() as conn:
            migrate_v102._deduction_policy.withdraw_violation_record(
                conn,
                record_id,
                effective_at="2026-08-03 12:00:00",
                reason="回归测试撤回",
            )
            conn.commit()
            state = conn.execute(
                """
                SELECT current_count_cache FROM member_group_states
                WHERE member_id=4 AND group_area='蜂巢'
                """
            ).fetchone()
            policy = conn.execute(
                """
                SELECT active_cycle_id, no_cycle_reason FROM v102_policy_state
                WHERE member_id=4 AND group_area='蜂巢'
                """
            ).fetchone()
            cycle = conn.execute(
                "SELECT * FROM v102_policy_cycles WHERE id=?",
                (policy["active_cycle_id"],),
            ).fetchone()

        self.assertEqual(state["current_count_cache"], 2)
        self.assertIsNotNone(policy["active_cycle_id"])
        self.assertIsNone(policy["no_cycle_reason"])
        self.assertEqual(cycle["cycle_type"], "normal")
        self.assertEqual(cycle["start_at"], CUTOVER)
        self.assertEqual(cycle["due_at"], "2026-08-16 17:00:00")

    def test_withdraw_replay_builds_baseline_cycle_at_cutover_watermark(self) -> None:
        self.apply_v102()
        record_ids = []
        with self.connect() as conn:
            for when in (
                "2026-08-03 10:00:00",
                "2026-08-04 10:00:00",
                "2026-08-05 10:00:00",
            ):
                record_id = int(
                    conn.execute(
                        """
                        INSERT INTO violation_records(
                            member_id, group_area, violation_time, judgement,
                            action, is_countable, count_delta, is_test,
                            created_at, updated_at
                        ) VALUES(1, '蜂巢', ?, '切换后违规', '禁言10分钟',
                                 1, 1, 0, ?, ?)
                        """,
                        (when, when, when),
                    ).lastrowid
                )
                record_ids.append(record_id)
                migrate_v102._deduction_policy.process_violation_record(
                    conn, record_id, ingest_time=when
                )
            migrate_v102._deduction_policy.withdraw_violation_record(
                conn,
                record_ids[-1],
                effective_at="2026-08-06 10:00:00",
                reason="回放水位线测试",
            )
            conn.commit()

            state = conn.execute(
                """
                SELECT current_count_cache FROM member_group_states
                WHERE member_id=1 AND group_area='蜂巢'
                """
            ).fetchone()
            active = conn.execute(
                """
                SELECT c.* FROM v102_policy_cycles c
                JOIN v102_policy_state p ON p.active_cycle_id=c.id
                WHERE p.member_id=1 AND p.group_area='蜂巢'
                """
            ).fetchone()
            generation = int(active["replay_generation"])
            normal = conn.execute(
                """
                SELECT * FROM v102_policy_cycles
                WHERE member_id=1 AND group_area='蜂巢'
                  AND replay_generation=? AND normal_light_count=2
                ORDER BY id DESC LIMIT 1
                """,
                (generation,),
            ).fetchone()

        self.assertEqual(state["current_count_cache"], 3)
        self.assertEqual(active["cycle_type"], "slow")
        self.assertEqual(active["due_at"], "2026-08-23 17:00:00")
        self.assertEqual(active["slow_light_count"], 0)
        self.assertEqual(active["slow_extended"], 0)
        self.assertIsNotNone(normal)

    def test_withdraw_replay_keeps_pre_cutover_backfill_out_of_cycle_counts(self) -> None:
        self.apply_v102()
        with self.connect() as conn:
            backfill_id = int(
                conn.execute(
                    """
                    INSERT INTO violation_records(
                        member_id, group_area, violation_time, judgement,
                        action, is_countable, count_delta, is_test,
                        created_at, updated_at
                    ) VALUES(1, '蜂巢', '2026-07-01 10:00:00', '切换后补录',
                             '禁言10分钟', 1, 1, 0,
                             '2026-08-03 10:00:00', '2026-08-03 10:00:00')
                    """
                ).lastrowid
            )
            migrate_v102._deduction_policy.process_violation_record(
                conn, backfill_id, ingest_time="2026-08-03 10:00:00"
            )
            later_id = int(
                conn.execute(
                    """
                    INSERT INTO violation_records(
                        member_id, group_area, violation_time, judgement,
                        action, is_countable, count_delta, is_test,
                        created_at, updated_at
                    ) VALUES(1, '蜂巢', '2026-08-04 10:00:00', '切换后违规',
                             '禁言10分钟', 1, 1, 0,
                             '2026-08-04 10:00:00', '2026-08-04 10:00:00')
                    """
                ).lastrowid
            )
            migrate_v102._deduction_policy.process_violation_record(
                conn, later_id, ingest_time="2026-08-04 10:00:00"
            )

            migrate_v102._deduction_policy.withdraw_violation_record(
                conn,
                later_id,
                effective_at="2026-08-05 10:00:00",
                reason="触发补录回放边界测试",
            )
            conn.commit()

            state = conn.execute(
                """
                SELECT current_count_cache FROM member_group_states
                WHERE member_id=1 AND group_area='蜂巢'
                """
            ).fetchone()
            active = conn.execute(
                """
                SELECT c.* FROM v102_policy_cycles c
                JOIN v102_policy_state p ON p.active_cycle_id=c.id
                WHERE p.member_id=1 AND p.group_area='蜂巢'
                """
            ).fetchone()

        self.assertEqual(state["current_count_cache"], 2)
        self.assertEqual(active["cycle_type"], "normal")
        self.assertEqual(active["start_at"], CUTOVER)
        self.assertEqual(active["due_at"], "2026-08-16 17:00:00")
        self.assertEqual(active["light_count"], 0)
        self.assertEqual(active["normal_light_count"], 0)

    def test_withdraw_rebuilds_final_warning_history_pending(self) -> None:
        with self.connect() as conn:
            member_id = int(
                conn.execute(
                    "SELECT id FROM members WHERE qq_number='60006'"
                ).fetchone()[0]
            )
            record_ids = []
            for sequence, when in enumerate(
                ("2026-07-15 10:00:00", "2026-07-20 10:00:00"), start=1
            ):
                record_ids.append(
                    int(
                        conn.execute(
                            """
                            INSERT INTO violation_records(
                                member_id, group_area, violation_time,
                                judgement, action, is_countable, count_delta,
                                is_test, created_at, updated_at
                            ) VALUES(?, '蜂箱', ?, ?, '禁言10分钟',
                                     1, 1, 0, ?, ?)
                            """,
                            (member_id, when, f"再犯{sequence}", when, when),
                        ).lastrowid
                    )
                )
            conn.commit()
        self.apply_v102()

        with self.connect() as conn:
            migrate_v102._deduction_policy.withdraw_violation_record(
                conn,
                record_ids[-1],
                effective_at="2026-08-03 12:00:00",
                reason="回归测试撤回",
            )
            conn.commit()
            cycle = conn.execute(
                """
                SELECT * FROM v102_policy_cycles
                WHERE member_id=? AND group_area='蜂箱'
                  AND status='pending_decision'
                ORDER BY id DESC LIMIT 1
                """,
                (member_id,),
            ).fetchone()
            pending = conn.execute(
                """
                SELECT p.*, e.event_type FROM v102_pending_actions p
                JOIN v102_policy_events e ON e.id=p.caused_by_event_id
                WHERE p.member_id=? AND p.group_area='蜂箱'
                  AND p.action_type='remove_member' AND p.status='pending'
                ORDER BY p.id DESC LIMIT 1
                """,
                (member_id,),
            ).fetchone()

        self.assertIsNotNone(cycle)
        self.assertEqual(cycle["cycle_type"], "final_warning")
        self.assertIsNotNone(pending)
        self.assertEqual(
            pending["event_type"], "historical_final_warning_violation_detected"
        )

    def apply(self) -> dict[str, object]:
        return self.apply_v102()

    def test_apply_calibrates_counts_audits_and_initializes_exact_cycle_types(self) -> None:
        violation_count_before = 1

        result = self.apply()

        self.assertEqual(result["batch_id"], BATCH_ID)
        self.assertEqual(result["source_sha256"], file_sha256(self.baseline))
        with self.connect() as conn:
            self.assertEqual(
                conn.execute("SELECT COUNT(*) FROM violation_records").fetchone()[0],
                violation_count_before,
            )
            self.assertEqual(
                conn.execute(
                    "SELECT qq_nickname FROM members WHERE qq_number='10001'"
                ).fetchone()[0],
                "旧昵称",
            )
            self.assertIsNone(
                conn.execute(
                    "SELECT qq_nickname FROM members WHERE qq_number='30003'"
                ).fetchone()[0]
            )
            checkpoint = conn.execute(
                "SELECT * FROM v102_migration_checkpoints WHERE batch_id=?",
                (BATCH_ID,),
            ).fetchone()
            self.assertEqual(checkpoint["status"], "applied")
            self.assertEqual(checkpoint["cutover_at"], CUTOVER)
            self.assertEqual(checkpoint["cutover_record_watermark"], 11)
            self.assertEqual(checkpoint["source_sha256"], file_sha256(self.baseline))
            self.assertEqual(checkpoint["backup_sha256"], self.last_backup_sha256)
            self.assertEqual(
                conn.execute(
                    "SELECT COUNT(*) FROM v102_baseline_audit WHERE batch_id=?",
                    (BATCH_ID,),
                ).fetchone()[0],
                7,
            )
            self.assertEqual(
                conn.execute(
                    "SELECT COUNT(*) FROM v102_policy_events WHERE event_type='baseline_migrated'"
                ).fetchone()[0],
                7,
            )
            policy_rows = conn.execute(
                "SELECT * FROM v102_policy_state ORDER BY member_id, group_area"
            ).fetchall()
            self.assertTrue(policy_rows)
            self.assertTrue(
                all(row["v102_operation_count"] == 0 for row in policy_rows)
            )

            hive = conn.execute(
                """
                SELECT s.*, p.baseline_adjustment, p.v102_operation_count
                FROM member_group_states s
                JOIN members m ON m.id=s.member_id
                JOIN v102_policy_state p
                  ON p.member_id=s.member_id AND p.group_area=s.group_area
                WHERE m.qq_number='10001' AND s.group_area='蜂巢'
                """
            ).fetchone()
            self.assertEqual(hive["total_count"], 3)
            self.assertEqual(hive["deduct_count"], 2)
            self.assertEqual(hive["current_count_cache"], 1)
            self.assertEqual(hive["baseline_adjustment"], 2)

            cycle_types = {
                (row["qq_number"], row["group_area"]): (
                    row["cycle_type"],
                    row["start_at"],
                    row["due_at"],
                )
                for row in conn.execute(
                    """
                    SELECT m.qq_number, c.group_area, c.cycle_type,
                           c.start_at, c.due_at
                    FROM v102_policy_cycles c
                    JOIN members m ON m.id=c.member_id
                    WHERE c.status='active'
                    """
                )
            }
            self.assertEqual(
                cycle_types[("10001", "蜂巢")],
                ("normal", CUTOVER, "2026-08-16 17:00:00"),
            )
            self.assertEqual(
                cycle_types[("10001", "蜂窝")],
                ("normal", CUTOVER, "2026-08-16 17:00:00"),
            )
            self.assertEqual(cycle_types[("40004", "蜂巢")][0], "slow")
            self.assertEqual(
                cycle_types[("60006", "蜂箱")],
                (
                    "final_warning",
                    "2026-07-01 00:00:00",
                    "2026-09-29 00:00:00",
                ),
            )
            self.assertEqual(cycle_types[("70007", "蜂箱")][0], "slow")

            zero_state = conn.execute(
                """
                SELECT p.* FROM v102_policy_state p
                JOIN members m ON m.id=p.member_id
                WHERE m.qq_number='30003' AND p.group_area='蜂巢'
                """
            ).fetchone()
            self.assertEqual(zero_state["no_cycle_reason"], "zero_count")
            terminal_state = conn.execute(
                """
                SELECT p.* FROM v102_policy_state p
                JOIN members m ON m.id=p.member_id
                WHERE m.qq_number='50005' AND p.group_area='蜂巢'
                """
            ).fetchone()
            self.assertEqual(
                terminal_state["no_cycle_reason"], "terminal_status"
            )

        verification = migrate_v102.verify_database(self.database)
        self.assertTrue(verification["ok"])
        self.assertEqual(verification["integrity_check"], "ok")
        self.assertEqual(verification["foreign_key_errors"], [])
        self.assertEqual(verification["invalid_policy_tags"], 0)
        self.assertEqual(verification["missing_unique_constraints"], [])

    def test_second_apply_fails_closed_without_changing_first_checkpoint(self) -> None:
        self.apply_v102()

        with self.assertRaisesRegex(migrate_v102.MigrationError, "已应用"):
            self.apply()

        with self.connect() as conn:
            self.assertEqual(
                conn.execute(
                    "SELECT COUNT(*) FROM v102_migration_checkpoints"
                ).fetchone()[0],
                1,
            )

    def test_invalid_backup_checksum_fails_before_database_write(self) -> None:
        snapshot, _ = self.create_pre_cutover_snapshot()
        database_before = file_sha256(self.database)

        with self.assertRaisesRegex(migrate_v102.MigrationError, "backup SHA-256"):
            migrate_v102.apply_migration(
                self.database,
                self.baseline,
                snapshot_database=snapshot,
                backup_sha256="not-a-sha",
                cutover_at=CUTOVER,
                batch_id=BATCH_ID,
            )

        self.assertEqual(file_sha256(self.database), database_before)

    def test_apply_rejects_backup_file_when_checksum_does_not_match(self) -> None:
        snapshot = self.root / "pre-cutover.db"
        shutil.copy2(self.database, snapshot)
        database_before = file_sha256(self.database)

        with self.assertRaisesRegex(migrate_v102.MigrationError, "摘要不匹配"):
            migrate_v102.apply_migration(
                self.database,
                self.baseline,
                snapshot_database=snapshot,
                backup_sha256="a" * 64,
                cutover_at=CUTOVER,
                batch_id=BATCH_ID,
            )

        self.assertEqual(file_sha256(self.database), database_before)

    def test_apply_rejects_backup_from_different_database(self) -> None:
        snapshot = self.root / "pre-cutover.db"
        shutil.copy2(self.database, snapshot)
        with sqlite3.connect(snapshot) as conn:
            conn.execute(
                """
                UPDATE member_group_states SET total_count=999
                WHERE member_id=1 AND group_area='蜂巢'
                """
            )
        digest = file_sha256(snapshot)
        database_before = file_sha256(self.database)

        with self.assertRaisesRegex(migrate_v102.MigrationError, "业务数据不一致"):
            migrate_v102.apply_migration(
                self.database,
                self.baseline,
                snapshot_database=snapshot,
                backup_sha256=digest,
                cutover_at=CUTOVER,
                batch_id=BATCH_ID,
            )

        self.assertEqual(file_sha256(self.database), database_before)

    def test_apply_rejects_hardlink_of_target_database_as_backup(self) -> None:
        snapshot = self.root / "hardlink-pre-cutover.db"
        os.link(self.database, snapshot)
        digest = file_sha256(snapshot)
        database_before = file_sha256(self.database)

        with self.assertRaisesRegex(migrate_v102.MigrationError, "同一文件"):
            migrate_v102.apply_migration(
                self.database,
                self.baseline,
                snapshot_database=snapshot,
                backup_sha256=digest,
                cutover_at=CUTOVER,
                batch_id=BATCH_ID,
            )

        self.assertEqual(file_sha256(self.database), database_before)

    def test_apply_compares_all_non_v102_business_tables(self) -> None:
        with self.connect() as conn:
            conn.execute(
                """
                CREATE TABLE member_policy_state(
                    id INTEGER PRIMARY KEY AUTOINCREMENT,
                    member_id INTEGER NOT NULL,
                    legacy_value TEXT NOT NULL
                )
                """
            )
            conn.execute(
                "INSERT INTO member_policy_state(member_id, legacy_value) VALUES(1, 'old')"
            )
            conn.commit()
        snapshot, digest = self.create_pre_cutover_snapshot()
        with self.connect() as conn:
            conn.execute(
                "UPDATE member_policy_state SET legacy_value='new' WHERE member_id=1"
            )
            conn.commit()
        database_before = file_sha256(self.database)

        with self.assertRaisesRegex(migrate_v102.MigrationError, "业务数据不一致"):
            migrate_v102.apply_migration(
                self.database,
                self.baseline,
                snapshot_database=snapshot,
                backup_sha256=digest,
                cutover_at=CUTOVER,
                batch_id=BATCH_ID,
            )

        self.assertEqual(file_sha256(self.database), database_before)

    def test_unsafe_batch_id_fails_before_database_write(self) -> None:
        snapshot, digest = self.create_pre_cutover_snapshot()
        database_before = file_sha256(self.database)

        with self.assertRaisesRegex(migrate_v102.MigrationError, "batch_id"):
            migrate_v102.apply_migration(
                self.database,
                self.baseline,
                snapshot_database=snapshot,
                backup_sha256=digest,
                cutover_at=CUTOVER,
                batch_id="../../unsafe batch",
            )

        self.assertEqual(file_sha256(self.database), database_before)

    def test_mid_apply_failure_rolls_back_schema_members_and_counts(self) -> None:
        with self.connect() as conn:
            conn.execute(
                """
                UPDATE member_group_states
                SET last_final_warning_time=NULL
                WHERE member_id=6 AND group_area='蜂箱'
                """
            )
            conn.commit()
            members_before = conn.execute(
                "SELECT id, qq_number, qq_nickname FROM members ORDER BY id"
            ).fetchall()
            states_before = conn.execute(
                """
                SELECT member_id, group_area, total_count, deduct_count,
                       current_count_cache
                FROM member_group_states
                ORDER BY member_id, group_area
                """
            ).fetchall()

        with self.assertRaisesRegex(
            migrate_v102.MigrationError,
            "last_final_warning_time",
        ):
            self.apply()

        with self.connect() as conn:
            self.assertEqual(
                conn.execute(
                    "SELECT id, qq_number, qq_nickname FROM members ORDER BY id"
                ).fetchall(),
                members_before,
            )
            self.assertEqual(
                conn.execute(
                    """
                    SELECT member_id, group_area, total_count, deduct_count,
                           current_count_cache
                    FROM member_group_states
                    ORDER BY member_id, group_area
                    """
                ).fetchall(),
                states_before,
            )
            self.assertIsNone(
                conn.execute(
                    "SELECT 1 FROM sqlite_master WHERE name='v102_policy_state'"
                ).fetchone()
            )
            self.assertIsNone(
                conn.execute(
                    "SELECT 1 FROM members WHERE qq_number='30003'"
                ).fetchone()
            )


class RollbackTests(MigrationTestCase):
    def test_rollback_rejects_missing_baseline_policy_state(self) -> None:
        self.apply_v102()
        with self.connect() as conn:
            conn.execute(
                """
                DELETE FROM v102_policy_state
                WHERE member_id=1 AND group_area='蜂巢'
                """
            )
            conn.commit()
        database_before = file_sha256(self.database)

        with self.assertRaisesRegex(migrate_v102.MigrationError, "快照"):
            migrate_v102.logical_rollback(
                self.database, BATCH_ID, rolled_back_at="2026-08-04 00:00:00"
            )

        self.assertEqual(file_sha256(self.database), database_before)

    def test_rollback_rejects_corrupt_baseline_snapshot(self) -> None:
        self.apply_v102()
        with self.connect() as conn:
            conn.execute(
                """
                UPDATE v102_policy_state
                SET baseline_total_count=baseline_total_count+1
                WHERE member_id=1 AND group_area='蜂巢'
                """
            )
            conn.commit()
        database_before = file_sha256(self.database)

        with self.assertRaisesRegex(migrate_v102.MigrationError, "快照"):
            migrate_v102.logical_rollback(
                self.database, BATCH_ID, rolled_back_at="2026-08-04 00:00:00"
            )

        self.assertEqual(file_sha256(self.database), database_before)

    def test_rollback_rejects_unrepaired_legacy_snapshot_checkpoint(self) -> None:
        self.apply_v102()
        with self.connect() as conn:
            conn.execute(
                """
                UPDATE v102_migration_checkpoints
                SET schema_version='v1.0.2beta-1'
                WHERE batch_id=?
                """,
                (BATCH_ID,),
            )
            conn.commit()
        database_before = file_sha256(self.database)

        with self.assertRaisesRegex(migrate_v102.MigrationError, "快照修复"):
            migrate_v102.logical_rollback(
                self.database, BATCH_ID, rolled_back_at="2026-08-04 00:00:00"
            )

        self.assertEqual(file_sha256(self.database), database_before)

    def test_rollback_restores_legacy_timer_fields_exactly(self) -> None:
        old_effective = "2026-05-01 08:00:00"
        old_deduct = "2026-05-15 08:00:00"
        with self.connect() as conn:
            conn.execute(
                """
                UPDATE member_group_states
                SET last_effective_violation_time=?, last_deduct_time=?
                WHERE member_id=1 AND group_area='蜂巢'
                """,
                (old_effective, old_deduct),
            )
            conn.commit()
        self.apply_v102()
        with self.connect() as conn:
            conn.execute(
                """
                UPDATE member_group_states
                SET last_effective_violation_time='2026-08-03 00:00:00',
                    last_deduct_time='2026-08-03 00:00:00'
                WHERE member_id=1 AND group_area='蜂巢'
                """
            )
            conn.commit()

        migrate_v102.logical_rollback(
            self.database, BATCH_ID, rolled_back_at="2026-08-04 00:00:00"
        )

        with self.connect() as conn:
            state = conn.execute(
                """
                SELECT last_effective_violation_time, last_deduct_time
                FROM member_group_states
                WHERE member_id=1 AND group_area='蜂巢'
                """
            ).fetchone()
        self.assertEqual(tuple(state), (old_effective, old_deduct))

    def test_rollback_preserves_timer_started_by_post_cutover_record(self) -> None:
        old_effective = "2026-05-01 08:00:00"
        old_deduct = "2026-05-15 08:00:00"
        new_time = "2026-08-03 00:00:00"
        with self.connect() as conn:
            conn.execute(
                """
                UPDATE member_group_states
                SET last_effective_violation_time=?, last_deduct_time=?
                WHERE member_id=1 AND group_area='蜂巢'
                """,
                (old_effective, old_deduct),
            )
            conn.commit()
        self.apply_v102()
        with self.connect() as conn:
            conn.execute(
                """
                INSERT INTO violation_records(
                    member_id, group_area, violation_time, judgement, action,
                    is_countable, count_delta, is_test, created_at, updated_at
                ) VALUES(1, '蜂巢', ?, '切换后违规', '禁言10分钟',
                         1, 1, 0, ?, ?)
                """,
                (new_time, new_time, new_time),
            )
            conn.execute(
                """
                UPDATE member_group_states
                SET last_effective_violation_time=?, last_deduct_time=?
                WHERE member_id=1 AND group_area='蜂巢'
                """,
                (new_time, new_time),
            )
            conn.commit()

        migrate_v102.logical_rollback(
            self.database, BATCH_ID, rolled_back_at="2026-08-04 00:00:00"
        )

        with self.connect() as conn:
            state = conn.execute(
                """
                SELECT last_effective_violation_time, last_deduct_time
                FROM member_group_states
                WHERE member_id=1 AND group_area='蜂巢'
                """
            ).fetchone()
        self.assertEqual(tuple(state), (new_time, new_time))

    def test_rollback_includes_runtime_initialized_scope(self) -> None:
        self.apply_v102()
        initialized = "2026-08-03 00:00:00"
        with self.connect() as conn:
            conn.execute(
                """
                INSERT INTO members(
                    qq_number, qq_nickname, aliases, created_at, updated_at
                ) VALUES('91010', '未覆盖旧成员', '[]', ?, ?)
                """,
                (initialized, initialized),
            )
            member_id = int(conn.execute("SELECT last_insert_rowid()").fetchone()[0])
            conn.execute(
                """
                INSERT INTO member_group_states(
                    member_id, group_area, status, total_count, deduct_count,
                    current_count_cache, last_effective_violation_time,
                    last_deduct_time, created_at, updated_at
                ) VALUES(?, '蜂巢', '正常', 5, 1, 4,
                         '2026-07-01 00:00:00', '2026-07-15 00:00:00', ?, ?)
                """,
                (member_id, initialized, initialized),
            )
            conn.execute(
                """
                INSERT INTO violation_records(
                    member_id, group_area, violation_time, judgement, action,
                    is_countable, count_delta, is_test, created_at, updated_at
                ) VALUES(?, '蜂巢', '2026-07-01 00:00:00', '历史违规',
                         '禁言10分钟', 1, 1, 0,
                         '2026-07-01 00:00:00', '2026-07-01 00:00:00')
                """,
                (member_id,),
            )
            migrate_v102._deduction_policy.ensure_policy_scope_snapshot(
                conn, member_id, "蜂巢", initialized
            )
            conn.execute(
                """
                UPDATE member_group_states
                SET deduct_count=3, current_count_cache=2,
                    last_deduct_time='2026-08-03 00:00:00'
                WHERE member_id=? AND group_area='蜂巢'
                """,
                (member_id,),
            )
            conn.execute(
                """
                UPDATE v102_policy_state
                SET policy_tag='slow', slow_level=2, v102_operation_count=2
                WHERE member_id=? AND group_area='蜂巢'
                """,
                (member_id,),
            )
            conn.commit()

        migrate_v102.logical_rollback(
            self.database, BATCH_ID, rolled_back_at="2026-08-04 00:00:00"
        )

        with self.connect() as conn:
            state = conn.execute(
                """
                SELECT total_count, deduct_count, current_count_cache,
                       last_effective_violation_time, last_deduct_time
                FROM member_group_states
                WHERE member_id=? AND group_area='蜂巢'
                """,
                (member_id,),
            ).fetchone()
            policy = conn.execute(
                """
                SELECT policy_tag, slow_level, v102_operation_count,
                       active_cycle_id, last_reason
                FROM v102_policy_state
                WHERE member_id=? AND group_area='蜂巢'
                """,
                (member_id,),
            ).fetchone()
        self.assertEqual(
            tuple(state),
            (5, 1, 4, "2026-07-01 00:00:00", "2026-07-15 00:00:00"),
        )
        self.assertEqual(tuple(policy), ("none", 0, 0, None, "logical_rollback"))

    def test_rollback_preserves_failed_status_job_without_policy_event(self) -> None:
        self.apply_v102()
        changed_at = "2026-08-03 00:00:00"
        with self.connect() as conn:
            member_id = conn.execute(
                "SELECT id FROM members WHERE qq_number='40004'"
            ).fetchone()[0]
            conn.execute(
                """
                UPDATE member_group_states
                SET status='已拉黑', locked=1, updated_at=?
                WHERE member_id=? AND group_area='蜂巢'
                """,
                (changed_at, member_id),
            )
            operation_log_id = int(
                conn.execute(
                    """
                    INSERT INTO operation_logs(
                        group_area, operation_type, source, target_member_id,
                        created_at, remark
                    ) VALUES('蜂巢', '更新状态', '人工', ?, ?, '联动失败演练')
                    """,
                    (member_id, changed_at),
                ).lastrowid
            )
            conn.execute(
                """
                INSERT INTO v102_status_bridge_jobs(
                    operation_log_id, member_id, group_area, target_status,
                    effective_at, idempotency_key, job_status, attempt_count,
                    last_error, created_at, updated_at
                ) VALUES(?, ?, '蜂巢', '已拉黑', ?, ?, 'failed', 1,
                         'RuntimeError: fixture', ?, ?)
                """,
                (
                    operation_log_id,
                    member_id,
                    changed_at,
                    f"status-log:{operation_log_id}",
                    changed_at,
                    changed_at,
                ),
            )
            conn.commit()

        migrate_v102.logical_rollback(
            self.database, BATCH_ID, rolled_back_at="2026-08-04 00:00:00"
        )

        with self.connect() as conn:
            state = conn.execute(
                """
                SELECT status, locked FROM member_group_states
                WHERE member_id=? AND group_area='蜂巢'
                """,
                (member_id,),
            ).fetchone()
            event_count = conn.execute(
                """
                SELECT COUNT(*) FROM v102_policy_events
                WHERE member_id=? AND group_area='蜂巢'
                  AND event_type='status_changed'
                """,
                (member_id,),
            ).fetchone()[0]
        self.assertEqual(event_count, 0)
        self.assertEqual(tuple(state), ("已拉黑", 1))

    def test_rollback_ignores_status_job_caused_by_withdrawn_record(self) -> None:
        self.apply_v102()
        changed_at = "2026-08-03 00:00:00"
        with self.connect() as conn:
            record_id = int(
                conn.execute(
                    """
                    INSERT INTO violation_records(
                        member_id, group_area, violation_time, judgement, action,
                        is_countable, count_delta, is_withdrawn, is_test,
                        created_at, updated_at
                    ) VALUES(1, '蜂巢', ?, '误记录', '禁言10分钟',
                             1, 1, 1, 0, ?, ?)
                    """,
                    (changed_at, changed_at, changed_at),
                ).lastrowid
            )
            conn.execute(
                """
                UPDATE member_group_states
                SET status='最后警告', locked=1,
                    last_final_warning_time=?, updated_at=?
                WHERE member_id=1 AND group_area='蜂巢'
                """,
                (changed_at, changed_at),
            )
            operation_log_id = int(
                conn.execute(
                    """
                    INSERT INTO operation_logs(
                        group_area, operation_type, source, target_member_id,
                        created_at, remark
                    ) VALUES('蜂巢', '最后警告', '人工', 1, ?, '因误记录触发')
                    """,
                    (changed_at,),
                ).lastrowid
            )
            conn.execute(
                """
                INSERT INTO v102_status_bridge_jobs(
                    operation_log_id, member_id, group_area, target_status,
                    caused_by_record_id, effective_at, idempotency_key,
                    job_status, created_at, updated_at
                ) VALUES(?, 1, '蜂巢', '最后警告', ?, ?, ?, 'applied', ?, ?)
                """,
                (
                    operation_log_id,
                    record_id,
                    changed_at,
                    f"status-log:{operation_log_id}",
                    changed_at,
                    changed_at,
                ),
            )
            conn.commit()

        migrate_v102.logical_rollback(
            self.database, BATCH_ID, rolled_back_at="2026-08-04 00:00:00"
        )

        with self.connect() as conn:
            state = conn.execute(
                """
                SELECT status, locked, last_final_warning_time
                FROM member_group_states
                WHERE member_id=1 AND group_area='蜂巢'
                """
            ).fetchone()
        self.assertEqual(tuple(state), ("正常", 0, None))

    def test_rollback_reverses_derived_status_and_preserves_manual_status(self) -> None:
        self.apply_v102()
        with self.connect() as conn:
            final_member_id = conn.execute(
                "SELECT id FROM members WHERE qq_number='60006'"
            ).fetchone()[0]
            manual_member_id = conn.execute(
                "SELECT id FROM members WHERE qq_number='40004'"
            ).fetchone()[0]
            conn.execute(
                """
                UPDATE member_group_states
                SET status='已质询', locked=0
                WHERE member_id=? AND group_area='蜂箱'
                """,
                (final_member_id,),
            )
            conn.execute(
                """
                UPDATE member_group_states
                SET status='已拉黑', locked=1
                WHERE member_id=? AND group_area='蜂巢'
                """,
                (manual_member_id,),
            )
            migrate_v102._insert_event(
                conn,
                member_id=manual_member_id,
                group_area="蜂巢",
                event_type="status_changed",
                effective_time="2026-08-03 00:00:00",
                event_priority=30,
                source_sequence=1,
                ingest_time="2026-08-03 00:00:00",
                idempotency_key="manual-status-after-cutover",
                payload={"status": "已拉黑"},
            )
            conn.commit()

        migrate_v102.logical_rollback(
            self.database, BATCH_ID, rolled_back_at="2026-08-04 00:00:00"
        )

        with self.connect() as conn:
            final_state = conn.execute(
                """
                SELECT status, locked FROM member_group_states
                WHERE member_id=? AND group_area='蜂箱'
                """,
                (final_member_id,),
            ).fetchone()
            manual_state = conn.execute(
                """
                SELECT status, locked FROM member_group_states
                WHERE member_id=? AND group_area='蜂巢'
                """,
                (manual_member_id,),
            ).fetchone()
        self.assertEqual(tuple(final_state), ("最后警告", 0))
        self.assertEqual(tuple(manual_state), ("已拉黑", 1))

    def test_rollback_orders_applied_status_jobs_by_effective_time(self) -> None:
        self.apply_v102()
        with self.connect() as conn:
            operation_ids = []
            for operation_type, created_at in (
                ("最后警告", "2026-08-10 00:00:00"),
                ("质询", "2026-08-11 00:00:00"),
            ):
                operation_ids.append(
                    int(
                        conn.execute(
                            """
                            INSERT INTO operation_logs(
                                group_area, operation_type, source,
                                target_member_id, created_at
                            ) VALUES('蜂巢', ?, '手动', 4, ?)
                            """,
                            (operation_type, created_at),
                        ).lastrowid
                    )
                )
            final_event = migrate_v102._deduction_policy.process_status_change(
                conn,
                member_id=4,
                group_area="蜂巢",
                status="最后警告",
                effective_at="2026-08-10 00:00:00",
                ingest_time="2026-08-10 00:00:00",
                idempotency_key=f"status-log:{operation_ids[0]}",
            )
            consulted_event = migrate_v102._deduction_policy.process_status_change(
                conn,
                member_id=4,
                group_area="蜂巢",
                status="已质询",
                effective_at="2026-08-05 00:00:00",
                ingest_time="2026-08-11 00:00:00",
                idempotency_key=f"status-log:{operation_ids[1]}",
            )
            conn.execute(
                """
                UPDATE member_group_states
                SET last_final_warning_time='2026-08-10 00:00:00'
                WHERE member_id=4 AND group_area='蜂巢'
                """
            )
            for operation_id, status, effective_at, event_id in (
                (
                    operation_ids[0],
                    "最后警告",
                    "2026-08-10 00:00:00",
                    final_event.event_id,
                ),
                (
                    operation_ids[1],
                    "已质询",
                    "2026-08-05 00:00:00",
                    consulted_event.event_id,
                ),
            ):
                conn.execute(
                    """
                    INSERT INTO v102_status_bridge_jobs(
                        operation_log_id, member_id, group_area, target_status,
                        effective_at, idempotency_key, job_status,
                        applied_event_id, created_at, updated_at
                    ) VALUES(?, 4, '蜂巢', ?, ?, ?, 'applied', ?, ?, ?)
                    """,
                    (
                        operation_id,
                        status,
                        effective_at,
                        f"status-log:{operation_id}",
                        event_id,
                        "2026-08-11 00:00:00",
                        "2026-08-11 00:00:00",
                    ),
                )
            conn.commit()
            before = conn.execute(
                """
                SELECT status FROM member_group_states
                WHERE member_id=4 AND group_area='蜂巢'
                """
            ).fetchone()["status"]
        self.assertEqual(before, "最后警告")

        migrate_v102.logical_rollback(
            self.database, BATCH_ID, rolled_back_at="2026-08-12 00:00:00"
        )

        with self.connect() as conn:
            state = conn.execute(
                """
                SELECT status, last_final_warning_time
                FROM member_group_states
                WHERE member_id=4 AND group_area='蜂巢'
                """
            ).fetchone()
        self.assertEqual(tuple(state), ("最后警告", "2026-08-10 00:00:00"))

    def test_logical_rollback_restores_old_baseline_and_keeps_new_records(self) -> None:
        self.apply_v102()
        with self.connect() as conn:
            member_id = conn.execute(
                "SELECT id FROM members WHERE qq_number='10001'"
            ).fetchone()[0]
            conn.execute(
                """
                INSERT INTO violation_records(
                    member_id, group_area, violation_time, judgement, action,
                    is_countable, count_delta, created_at, updated_at
                ) VALUES(?, '蜂巢', '2026-08-03 00:00:00', '切换后违规',
                         '禁言10分钟', 1, 1,
                         '2026-08-03 00:00:00', '2026-08-03 00:00:00')
                """,
                (member_id,),
            )
            conn.execute(
                """
                UPDATE member_group_states
                SET deduct_count=3, current_count_cache=1
                WHERE member_id=? AND group_area='蜂巢'
                """,
                (member_id,),
            )
            conn.execute(
                """
                UPDATE v102_policy_state SET v102_operation_count=1
                WHERE member_id=? AND group_area='蜂巢'
                """,
                (member_id,),
            )
            conn.commit()

        result = migrate_v102.logical_rollback(
            self.database, f" {BATCH_ID} ", rolled_back_at="2026-08-04 00:00:00"
        )

        self.assertEqual(result["status"], "rolled_back")
        with self.connect() as conn:
            member_id = conn.execute(
                "SELECT id FROM members WHERE qq_number='10001'"
            ).fetchone()[0]
            state = conn.execute(
                """
                SELECT * FROM member_group_states
                WHERE member_id=? AND group_area='蜂巢'
                """,
                (member_id,),
            ).fetchone()
            self.assertEqual(state["total_count"], 9)
            self.assertEqual(state["deduct_count"], 2)
            self.assertEqual(state["current_count_cache"], 7)
            self.assertEqual(
                conn.execute(
                    "SELECT COUNT(*) FROM violation_records WHERE member_id=?",
                    (member_id,),
                ).fetchone()[0],
                2,
            )
            policy = conn.execute(
                """
                SELECT * FROM v102_policy_state
                WHERE member_id=? AND group_area='蜂巢'
                """,
                (member_id,),
            ).fetchone()
            self.assertEqual(policy["v102_operation_count"], 0)
            self.assertIsNone(policy["active_cycle_id"])
            self.assertEqual(
                conn.execute(
                    "SELECT status FROM v102_migration_checkpoints WHERE batch_id=?",
                    (BATCH_ID,),
                ).fetchone()[0],
                "rolled_back",
            )
            self.assertEqual(
                conn.execute(
                    "SELECT COUNT(*) FROM v102_policy_cycles WHERE status='active'"
                ).fetchone()[0],
                0,
            )
            rollback_log = conn.execute(
                """
                SELECT remark FROM operation_logs
                WHERE operation_type='v1.0.2beta逻辑回滚'
                ORDER BY id DESC LIMIT 1
                """
            ).fetchone()
            self.assertEqual(rollback_log["remark"], f"batch_id={BATCH_ID}")

        verification = migrate_v102.verify_database(self.database)
        self.assertTrue(verification["ok"], verification["errors"])


class CliTests(MigrationTestCase):
    def test_dry_run_cli_emits_json_without_writing(self) -> None:
        before = file_sha256(self.database)
        command = [
            sys.executable,
            str(Path(migrate_v102.__file__)),
            "--database",
            str(self.database),
            "--baseline",
            str(self.baseline),
            "--dry-run",
        ]

        completed = subprocess.run(
            command,
            cwd=Path(migrate_v102.__file__).resolve().parents[1],
            check=False,
            capture_output=True,
            text=True,
        )

        self.assertEqual(completed.returncode, 0, completed.stderr)
        payload = json.loads(completed.stdout)
        self.assertEqual(payload["summary"]["baseline_rows"], 7)
        self.assertEqual(file_sha256(self.database), before)


if __name__ == "__main__":
    unittest.main()
