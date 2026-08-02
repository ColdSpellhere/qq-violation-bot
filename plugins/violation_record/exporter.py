import csv
from datetime import datetime, timedelta
from pathlib import Path
from typing import Any

from openpyxl import Workbook

from .admin_resolver import resolve_operator
from .config import EXPORT_DIR
from .db import connect, now_str


def _stamp() -> str:
    return datetime.now().strftime("%Y%m%d_%H%M%S")


def _range_from_intent(intent: dict[str, Any]) -> tuple[str, str | None, str | None]:
    query = intent.get("query") or {}
    raw = str(intent.get("_raw", ""))
    time_range = query.get("time_range")
    if not time_range:
        if "本月" in raw or "这个月" in raw or "当月" in raw:
            time_range = "current_month"
        elif "本周" in raw or "这周" in raw or "这个星期" in raw:
            time_range = "current_week"
        elif "昨天" in raw or "昨日" in raw:
            time_range = "yesterday"
        elif "今天" in raw or "今日" in raw:
            time_range = "today"
        elif "最近" in raw:
            time_range = "recent_days"
        else:
            time_range = "all"
    now = datetime.now().replace(second=0, microsecond=0)
    if time_range == "today":
        start = now.replace(hour=0, minute=0)
        return "今天", start.strftime("%Y-%m-%d %H:%M:%S"), now.strftime("%Y-%m-%d %H:%M:%S")
    if time_range == "yesterday":
        day = now - timedelta(days=1)
        return "昨天", day.replace(hour=0, minute=0).strftime("%Y-%m-%d %H:%M:%S"), day.replace(hour=23, minute=59).strftime("%Y-%m-%d %H:%M:%S")
    if time_range == "current_week":
        start = (now - timedelta(days=now.weekday())).replace(hour=0, minute=0)
        return "本周", start.strftime("%Y-%m-%d %H:%M:%S"), now.strftime("%Y-%m-%d %H:%M:%S")
    if time_range == "current_month":
        start = now.replace(day=1, hour=0, minute=0)
        return "本月", start.strftime("%Y-%m-%d %H:%M:%S"), now.strftime("%Y-%m-%d %H:%M:%S")
    if time_range == "recent_days":
        days = int(query.get("recent_days") or 14)
        start = now - timedelta(days=days)
        return f"最近{days}天", start.strftime("%Y-%m-%d %H:%M:%S"), now.strftime("%Y-%m-%d %H:%M:%S")
    return "全部", None, None


def _write_csv(path: Path, rows: list[dict[str, Any]]) -> None:
    if not rows:
        path.write_text("", encoding="utf-8")
        return
    with path.open("w", newline="", encoding="utf-8-sig") as f:
        writer = csv.DictWriter(f, fieldnames=list(rows[0].keys()))
        writer.writeheader()
        writer.writerows(rows)


def _write_xlsx(path: Path, rows: list[dict[str, Any]]) -> None:
    wb = Workbook()
    ws = wb.active
    ws.title = "data"
    if rows:
        headers = list(rows[0].keys())
        ws.append(headers)
        for row in rows:
            ws.append([row.get(h) for h in headers])
    wb.save(path)


def export_records(area: str | None = None, fmt: str = "xlsx", range_label: str = "全部", start: str | None = None, end: str | None = None) -> Path:
    EXPORT_DIR.mkdir(parents=True, exist_ok=True)
    with connect() as conn:
        where = ["v.is_withdrawn=0"]
        params: list[Any] = []
        if area:
            where.append("v.group_area=?")
            params.append(area)
        if start and end:
            where.append("v.violation_time BETWEEN ? AND ?")
            params.extend([start, end])
        where_sql = f"WHERE {' AND '.join(where)}" if where else ""
        rows = [dict(r) for r in conn.execute(
            f"""
            SELECT
                v.group_area AS 群聊,
                m.qq_nickname AS QQ昵称,
                m.qq_number AS QQ号,
                v.violation_time AS 违规时间,
                v.judgement AS 判定,
                v.action AS 处理措施,
                COALESCE(handler.nickname, '') AS 处理人,
                COALESCE(recorder.nickname, '') AS 记录人,
                COALESCE(v.remark, '无') AS 备注
            FROM violation_records v
            JOIN members m ON m.id=v.member_id
            LEFT JOIN admins handler ON handler.id=v.handler_admin_id
            LEFT JOIN admins recorder ON recorder.id=v.recorder_admin_id
            {where_sql}
            ORDER BY v.violation_time DESC
            """,
            params,
        ).fetchall()]
    range_part = "" if range_label == "全部" else f"_{range_label}"
    path = EXPORT_DIR / f"{area or '全部'}{range_part}_violation_records_{_stamp()}.{fmt}"
    _write_xlsx(path, rows) if fmt == "xlsx" else _write_csv(path, rows)
    return path


def export_logs(area: str | None = None, fmt: str = "xlsx") -> Path:
    EXPORT_DIR.mkdir(parents=True, exist_ok=True)
    with connect() as conn:
        if area:
            rows = [dict(r) for r in conn.execute("SELECT * FROM operation_logs WHERE group_area=? ORDER BY created_at DESC", (area,)).fetchall()]
        else:
            rows = [dict(r) for r in conn.execute("SELECT * FROM operation_logs ORDER BY created_at DESC").fetchall()]
    path = EXPORT_DIR / f"{area or '全部'}_operation_logs_{_stamp()}.{fmt}"
    _write_xlsx(path, rows) if fmt == "xlsx" else _write_csv(path, rows)
    return path


def weekly_report(fmt: str = "xlsx") -> Path:
    EXPORT_DIR.mkdir(parents=True, exist_ok=True)
    since = (datetime.now() - timedelta(days=7)).strftime("%Y-%m-%d %H:%M:%S")
    with connect() as conn:
        updates = [dict(r) for r in conn.execute("SELECT * FROM operation_logs WHERE created_at>=? ORDER BY created_at", (since,)).fetchall()]
        stats = [dict(r) for r in conn.execute(
            """
            SELECT s.group_area, m.qq_nickname, m.qq_number, s.status, s.total_count, s.deduct_count, s.current_count_cache
            FROM member_group_states s JOIN members m ON m.id=s.member_id
            ORDER BY s.group_area, s.current_count_cache DESC
            """
        ).fetchall()]
        has_policy = conn.execute(
            """
            SELECT 1 FROM sqlite_master
            WHERE type='table' AND name='v102_policy_events'
            """
        ).fetchone()
        policy_events: list[dict[str, Any]] = []
        policy_pending: list[dict[str, Any]] = []
        notification_attempts: list[dict[str, Any]] = []
        status_bridge_jobs: list[dict[str, Any]] = []
        if has_policy:
            policy_events = [
                dict(r)
                for r in conn.execute(
                    """
                    SELECT
                        e.id AS 事件编号,
                        e.group_area AS 群域,
                        m.qq_nickname AS QQ昵称,
                        m.qq_number AS QQ号,
                        e.event_type AS 事件类型,
                        e.effective_time AS 实际时间,
                        e.ingest_time AS 处理时间,
                        e.source_record_id AS 来源记录,
                        e.caused_by_event_id AS 因果事件,
                        e.reversed_by_event_id AS 冲正事件,
                        e.superseded_by_replay_id AS 替代回放事件,
                        e.replay_generation AS 回放代次,
                        e.payload_json AS 结构化详情,
                        e.rule_version AS 规则版本,
                        e.is_effective AS 当前有效,
                        COALESCE(
                            (
                                SELECT GROUP_CONCAT(DISTINCT o.status)
                                FROM v102_notification_outbox o
                                WHERE o.event_id=e.id
                            ),
                            '未排队'
                        ) AS 通知状态
                    FROM v102_policy_events e
                    JOIN members m ON m.id=e.member_id
                    WHERE e.created_at>=?
                    ORDER BY e.effective_time, e.event_priority,
                             e.source_sequence, e.id
                    """,
                    (since,),
                ).fetchall()
            ]
            policy_pending = [
                dict(r)
                for r in conn.execute(
                    """
                    SELECT
                        p.id AS 待办编号,
                        p.group_area AS 群域,
                        m.qq_nickname AS QQ昵称,
                        m.qq_number AS QQ号,
                        p.action_type AS 待办类型,
                        p.status AS 状态,
                        p.due_at AS 周期截止,
                        p.next_reminder_at AS 下次提醒,
                        p.reason AS 事由,
                        p.caused_by_event_id AS 来源事件,
                        p.decision_event_id AS 决定事件,
                        p.created_at AS 创建时间,
                        p.updated_at AS 更新时间
                    FROM v102_pending_actions p
                    JOIN members m ON m.id=p.member_id
                    WHERE p.created_at>=? OR p.updated_at>=? OR p.status='pending'
                    ORDER BY p.group_area, m.qq_number, p.id
                    """,
                    (since, since),
                ).fetchall()
            ]
            notification_attempts = [
                dict(r)
                for r in conn.execute(
                    """
                    SELECT
                        a.id AS 尝试编号,
                        a.outbox_id AS 通知编号,
                        o.event_id AS 事件编号,
                        o.group_area AS 群域,
                        m.qq_nickname AS QQ昵称,
                        m.qq_number AS QQ号,
                        o.message_type AS 通知类型,
                        o.reminder_slot AS 提醒时段,
                        a.attempt_number AS 尝试序号,
                        a.status AS 发送结果,
                        a.started_at AS 开始时间,
                        a.finished_at AS 完成时间,
                        a.detail AS 结果详情,
                        o.message_text AS 通知内容
                    FROM v102_notification_attempts a
                    JOIN v102_notification_outbox o ON o.id=a.outbox_id
                    LEFT JOIN members m ON m.id=o.member_id
                    WHERE a.created_at>=? OR a.updated_at>=?
                    ORDER BY a.started_at, a.outbox_id, a.attempt_number
                    """,
                    (since, since),
                ).fetchall()
            ]
            status_bridge_jobs = [
                dict(r)
                for r in conn.execute(
                    """
                    SELECT
                        j.id AS 作业编号,
                        j.group_area AS 群域,
                        m.qq_nickname AS QQ昵称,
                        m.qq_number AS QQ号,
                        j.target_status AS 目标状态,
                        j.effective_at AS 实际生效时间,
                        j.job_status AS 作业状态,
                        j.attempt_count AS 尝试次数,
                        j.applied_event_id AS 策略事件,
                        j.last_error AS 最近错误,
                        l.operation_type AS 业务操作,
                        l.operator_qq AS 操作人QQ,
                        l.operator_nickname AS 操作人昵称,
                        l.before_json AS 操作前,
                        l.after_json AS 操作后,
                        l.remark AS 事由,
                        j.created_at AS 创建时间,
                        j.updated_at AS 更新时间
                    FROM v102_status_bridge_jobs j
                    JOIN operation_logs l ON l.id=j.operation_log_id
                    JOIN members m ON m.id=j.member_id
                    WHERE j.created_at>=? OR j.updated_at>=?
                       OR j.job_status!='applied'
                    ORDER BY j.created_at, j.id
                    """,
                    (since, since),
                ).fetchall()
            ]
    path = EXPORT_DIR / f"weekly_report_{_stamp()}.{fmt}"
    if fmt == "xlsx":
        wb = Workbook()
        ws = wb.active
        ws.title = "本周更新"
        if updates:
            headers = list(updates[0].keys())
            ws.append(headers)
            for row in updates:
                ws.append([row.get(h) for h in headers])
        ws2 = wb.create_sheet("当前统计")
        if stats:
            headers = list(stats[0].keys())
            ws2.append(headers)
            for row in stats:
                ws2.append([row.get(h) for h in headers])
        if has_policy:
            ws3 = wb.create_sheet("减数策略日志")
            if policy_events:
                headers = list(policy_events[0].keys())
                ws3.append(headers)
                for row in policy_events:
                    ws3.append([row.get(h) for h in headers])
            ws4 = wb.create_sheet("减数待办")
            if policy_pending:
                headers = list(policy_pending[0].keys())
                ws4.append(headers)
                for row in policy_pending:
                    ws4.append([row.get(h) for h in headers])
            ws5 = wb.create_sheet("通知发送历史")
            if notification_attempts:
                headers = list(notification_attempts[0].keys())
                ws5.append(headers)
                for row in notification_attempts:
                    ws5.append([row.get(h) for h in headers])
            ws6 = wb.create_sheet("状态联动作业")
            if status_bridge_jobs:
                headers = list(status_bridge_jobs[0].keys())
                ws6.append(headers)
                for row in status_bridge_jobs:
                    ws6.append([row.get(h) for h in headers])
        wb.save(path)
    else:
        _write_csv(path, updates)
    return path


def export_by_intent(intent: dict[str, Any], operator_qq: str, operator_nickname: str | None, message_id: str | None) -> str:
    area = intent.get("group_area")
    raw = intent.get("_raw", "")
    fmt = "csv" if "csv" in raw.lower() else "xlsx"
    range_label, start, end = _range_from_intent(intent)
    try:
        if "日志" in raw:
            path = export_logs(area, fmt)
            op_type = "导出日志"
        else:
            path = export_records(area, fmt, range_label, start, end)
            op_type = "导出"
        with connect() as conn:
            operator = resolve_operator(operator_qq, operator_nickname)
            conn.execute(
                """
                INSERT INTO operation_logs(group_area, operation_type, source, operator_qq, operator_nickname,
                    target_member_id, before_json, after_json, message_id, created_at, remark)
                VALUES(?, ?, '手动', ?, ?, NULL, NULL, ?, ?, ?, ?)
                """,
                (area, op_type, (operator or {}).get("qq_number"), (operator or {}).get("nickname"), str(path), message_id, now_str(), f"导出文件；范围={range_label}"),
            )
        return f"已导出：{path}"
    except Exception as exc:
        return f"导出失败：{exc}"
