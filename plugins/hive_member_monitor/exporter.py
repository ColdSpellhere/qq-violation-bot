from __future__ import annotations

import os
import re
import tempfile
from datetime import datetime
from pathlib import Path
from typing import Iterable

from openpyxl import Workbook

from .store import MemberSnapshot, normalize_members


_ILLEGAL_XML_CONTROL_CHARACTERS = re.compile(r"[\x00-\x08\x0b\x0c\x0e-\x1f]")
_UNSAFE_FILE_NAME_CHARACTERS = re.compile(r"[\\/:*?\"<>|\[\]]")
_FORMULA_PREFIXES = ("=", "+", "-", "@")


def _safe_excel_text(value: object, *, fallback: str) -> str:
    """Return untrusted display text that Excel cannot interpret as a formula."""

    text = _ILLEGAL_XML_CONTROL_CHARACTERS.sub("", str(value or ""))
    if not text:
        text = fallback
    if text.lstrip(" \t\r\n").startswith(_FORMULA_PREFIXES):
        text = "'" + text
    return text


def normalize_group_label(value: object) -> str:
    label = _ILLEGAL_XML_CONTROL_CHARACTERS.sub("", str(value or "").strip())
    label = _UNSAFE_FILE_NAME_CHARACTERS.sub("_", label).strip(" .")
    return label[:20] or "群"


def export_member_list(
    members: Iterable[MemberSnapshot],
    *,
    output_dir: Path,
    group_label: str = "蜂巢",
    now: datetime | None = None,
) -> Path:
    """Write the requested two-column member workbook atomically."""

    normalized = normalize_members(members)
    if not normalized:
        raise ValueError("member export requires at least one valid member")

    output_dir = Path(output_dir)
    parent_existed = output_dir.exists()
    output_dir.mkdir(parents=True, exist_ok=True)
    if not parent_existed:
        os.chmod(output_dir, 0o700)

    timestamp = now or datetime.now()
    safe_group_label = normalize_group_label(group_label)
    workbook_title = f"{safe_group_label}群员名单"
    destination = output_dir / (
        f"{workbook_title}_{timestamp:%Y-%m-%d_%H-%M-%S}.xlsx"
    )
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = workbook_title
    sheet.append(("QQ号", "QQ名字"))
    for row_number, member in enumerate(normalized, start=2):
        qq_cell = sheet.cell(row=row_number, column=1, value=str(member.user_id))
        qq_cell.number_format = "@"
        name_cell = sheet.cell(
            row=row_number,
            column=2,
            value=_safe_excel_text(member.qq_name, fallback=member.user_id),
        )
        name_cell.data_type = "s"
    sheet.column_dimensions["A"].width = 18
    sheet.column_dimensions["B"].width = 24
    sheet.freeze_panes = "A2"

    file_descriptor, temporary_name = tempfile.mkstemp(
        prefix=".hive-members-",
        suffix=".xlsx.tmp",
        dir=output_dir,
    )
    os.close(file_descriptor)
    temporary_path = Path(temporary_name)
    try:
        os.chmod(temporary_path, 0o600)
        workbook.save(temporary_path)
        os.replace(temporary_path, destination)
        os.chmod(destination, 0o600)
    finally:
        workbook.close()
        temporary_path.unlink(missing_ok=True)
    return destination
